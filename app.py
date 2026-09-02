from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, render_template_string, request, redirect, flash, send_file, jsonify, session, Response, stream_with_context
import pandas as pd
import io
import math
import urllib.parse
import os
import json
from pathlib import Path
import tempfile
from werkzeug.utils import secure_filename
from supabase import create_client, Client, ClientOptions
from dotenv import load_dotenv
from datetime import datetime, timedelta, timezone
import re
import csv
from urllib.parse import urlparse
from functools import wraps
from utils.pdf_generator import generate_pdf
from utils.aws_upload import upload_pdf, delete_from_s3
from utils.filename_generator import generate_filename
import time
import fitz  # PyMuPDF — pip install pymupdf
import pytesseract  # pip install pytesseract
from PIL import Image
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import traceback
import uuid
import threading
import shutil
import base64
import html

_EASYOCR_READER = None
_EASYOCR_LOCK = threading.Lock()
_AML_DEBUG_DRIVERS = []

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "your-secret-key-change-this")
app.permanent_session_lifetime = timedelta(hours=8)

# ============================================================
# AUTH HELPERS
# ============================================================
    
# Cap PostgREST timeouts below Cloudflare's ~60s wait window so a slow
# Supabase query fails fast inside Python with a clear exception, instead
# of being silently killed by Cloudflare with a 522 "Connection timed out"
# page that breaks downstream `json.loads(error)` paths.
_SUPABASE_CLIENT_OPTS = ClientOptions(
    postgrest_client_timeout=30,   # select/insert/update/delete queries
    storage_client_timeout=30,    # file storage
    function_client_timeout=30,   # edge functions
)

def get_auth_supabase():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    return create_client(url, key, options=_SUPABASE_CLIENT_OPTS)

DEMO_ADMIN = {
    "id": 0,
    "email": "test123@gmail.com",
    "password": "test123",
    "display_name": "Test User - (Testing Only)",
    "allowed_pages": ["qc"],
    "is_admin": False,
    "role": "user",
    "is_active": True,
    "can_view_activity_log": True,
    "allowed_departments": ["ITC","AML", "Investment Scam","dashboard_management","Infringement", "Chargeback"],
    "created_at": "2025-01-01"
}

def parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in ("1", "true", "yes", "y", "on")


def fetch_user_by_email(email: str):
    if email.lower().strip() == DEMO_ADMIN["email"]:
        return DEMO_ADMIN
    try:
        client = get_auth_supabase()
        res = client.table("dashboard_users") \
            .select("*") \
            .eq("email", email.lower().strip()) \
            .eq("is_active", True) \
            .limit(1) \
            .execute()
        if res.data:
            return res.data[0]
        return None
    except Exception as e:
        print(f"[AUTH] fetch_user_by_email error: {e}")
        return None
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Welcome to Scam Intelligence GUI.", "error")
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated
def get_current_user():
    if "user_id" not in session:
        return None
    return {
        "id": session.get("user_id"),
        "email": session.get("email"),
        "display_name": session.get("display_name"),
        "allowed_pages": session.get("allowed_pages", []),
        "is_admin": session.get("is_admin", False),
        "role": session.get("role", "user"),
        "can_view_activity_log": session.get("can_view_activity_log", False),
        "allowed_departments": session.get("allowed_departments"),
    }
# ============================================================
# ACTIVITY LOG HELPER
# ============================================================
def log_activity(action_type, target_table=None, target_record_id=None,
                 field_name=None, old_value=None, new_value=None, extra_info=None):
    try:
        client = get_auth_supabase()
        client.table("activity_logs").insert({
            "user_id":          session.get("user_id"),
            "user_email":       session.get("email"),
            "display_name":     session.get("display_name"),
            "action_type":      action_type,
            "target_table":     target_table,
            "target_record_id": target_record_id,
            "field_name":       field_name,
            "old_value":        str(old_value) if old_value is not None else None,
            "new_value":        str(new_value) if new_value is not None else None,
            "extra_info":       extra_info,
        }).execute()
    except Exception as e:
        print(f"[ACTIVITY LOG] Failed to log activity: {e}")
# ============================================================
# Configuration
# ============================================================
PER_PAGE = 100
BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "sheet_mapping_config.json"
EXCEL_FOLDER_PATH = BASE_DIR / "excel_data"
EXCEL_FOLDER_PATH.mkdir(exist_ok=True)
BANK_NAME_MAPPING_PATH = EXCEL_FOLDER_PATH / "bank_name.xlsx"
IFSC_MAPPING_PATH = EXCEL_FOLDER_PATH / "ifsc_mapping.xlsx"

supabase: Client = create_client(
    os.environ.get("SUPABASE_URL"),
    os.environ.get("SUPABASE_KEY"),
    options=_SUPABASE_CLIENT_OPTS,
)
social_supabase: Client = supabase

PLATFORM_OPTIONS = [
    "Telegram", "WhatsApp", "Facebook", "Instagram",
    "Thread", "YouTube", "X"
]

SCAM_TYPE_OPTIONS = [
    "Investment Scam", "Carding Scam", "Shopping Scam",
    "Job Scam", "Subscription Scam", "Loan Scam",
    "Currency Exchange Scam", "Fake Account Selling Scam"
]

SOCIAL_PLATFORM_OPTIONS = [
    "Facebook", "Amazon", "Instagram", "Telegram", "WhatsApp",
    "Gmail Accounts", "Total Numbers"
]

WEBSITE_DIRECTORY_CATEGORY_OPTIONS = [
    "Job Scam", "Subscription Scam", "Fake Website Scam",
    "Loan Scam", "Government Scheme Scam", "Investment Scam",
    "LPG Booking Scam", "IPL Tickets Scam", "Carding Scam",
]

WEBSITE_DIRECTORY_SEARCH_FOR_OPTIONS = ["Web", "App"]

WEBSITE_ALLOTMENT_TABLE = "website_allotment"

ALLOTMENT_REMARK_OPTIONS = [
    "Found On GUI", "Website Not Working", "OTP-Based Login",
    "UPI Not Available", "System Under Maintenance", "Login Issues",
    "Payment Processing Error", "Deposit Section Issues", "Need New Credentials",
]

WEBSITE_DIRECTORY_COLUMNS = [
    "id", "date", "name", "url", "final_url", "invitation_code",
    "search_for", "group_app_name", "number", "email", "login_id",
    "password", "remark", "origin", "category",
    "automated_website", "payment_gateway", "inserted_at"
]

DEPARTMENT_OPTIONS = [
    "AML", "Investment Scam", "ITC", "Infringement", "Chargeback"
]

def can_access_lunch(user_session):
    """Check if user can access lunch break tracker"""
    allowed_pages = user_session.get("allowed_pages", [])
    return "lunch" in allowed_pages

def can_access_allotment(user_session):
    """Check if user can access Website Allotment page (either role)"""
    allowed_pages = user_session.get("allowed_pages", [])
    return "allotment" in allowed_pages or "allotment_admin" in allowed_pages

def is_allotment_admin(user_session):
    """Check if user is allowed to allot websites to others / see everyone's allotment"""
    allowed_pages = user_session.get("allowed_pages", [])
    return "allotment_admin" in allowed_pages


def is_superadmin(user_session):
    """True for users whose role is 'superadmin' — full access incl. user management, no dept filter."""
    return user_session.get("role") == "superadmin"


def is_admin_or_above(user_session):
    """True for users whose role is 'superadmin' or 'admin' — operational admin actions allowed."""
    return user_session.get("role") in ("superadmin", "admin")


ALL_EMPLOYEES = [
    "Parul Satsangi",
    "Kakul Pal",    
    "Rozma Khan",
    "Rishabh Yadav",
    "Nitin Kumar",
    "Vidhi Satsangi",
]

PLATFORM_ACCOUNT_STATUS = {
    "Facebook": ["Active", "Temporarily Block", "Restricted", "Permanent Block"],
    "Instagram": ["Active", "Temporarily Block", "Permanent Block"],
    "Telegram": ["Active", "Frozen", "Permanent Block"],
    "WhatsApp": ["Active", "Temporarily Block", "Permanent Block", "Restricted"],
    "Amazon": ["Active", "Temporarily Block", "Permanent Block"],
    "Gmail Accounts": ["Active", "Temporarily Block", "Permanent Block"],
    "Total Numbers": ["Active", "Temporarily Block", "Permanent Block"],
}
TEMPORARY_BLOCK_STATUS = "Temporarily Block"
LEGACY_TEMPORARY_BLOCK_STATUSES = {"Block", "Blocked"}


def normalize_social_account_status(value):
    """Canonicalize old social-account block status values.

    Older rows/UI builds used "Block" (and a few display helpers also tolerated
    "Blocked"). The current product label and database value is
    "Temporarily Block".
    """
    status = str(value or "").strip()
    if status in LEGACY_TEMPORARY_BLOCK_STATUSES:
        return TEMPORARY_BLOCK_STATUS
    return status or "Active"


def normalize_social_account_row(row):
    normalized = dict(row or {})
    normalized["account_status"] = normalize_social_account_status(normalized.get("account_status"))
    return normalized


def apply_social_status_filter(query, status):
    normalized_status = normalize_social_account_status(status)
    if normalized_status == TEMPORARY_BLOCK_STATUS:
        return query.in_("account_status", [TEMPORARY_BLOCK_STATUS, *LEGACY_TEMPORARY_BLOCK_STATUSES])
    return query.eq("account_status", normalized_status)

BS_INVESTMENT_COLUMNS = [
    "id", "bank_account_number", "bank_name", "upi_vpa", "screenshot",
    "search_for", "upi_bank_account_wallet", "handle", "payment_gateway_name",
    "scam_type", "ifsc_code", "upi_url", "website_url", "inserted_date",
    "input_user", "web_contact_no"
]

BS_INVESTMENT_SCAM_TYPE_OPTIONS = [
    "Investment Scam", "Loan Scam", "Subscription Scam", "Carding Scam","Government Scheme Scam",
    "Fake Website Scam", "Currency Exchange Scam", "Job Scam", "Shopping Scam","LPG Booking Scam", "IPL Tickets Scam", "ChaarDham Booking Scam"
]

BS_INVESTMENT_SEARCH_FOR_OPTIONS = [
    "Web", "Telegram", "WhatsApp", "Facebook",
    "Instagram", "YouTube", "X", "Thread"
]
BS_INVESTMENT_SM_SEARCH_FOR_VALUES = [
    "Telegram", "WhatsApp", "Facebook", "Instagram",
    "YouTube", "X", "Twitter", "Thread", "Threads",
    "Snapchat", "TikTok", "LinkedIn", "Pinterest", "Reddit"
]

BS_INVESTMENT_WALLET_OPTIONS = [
    "UPI", "Bank Account", "Wallet"
]

REQUIRED_COLUMNS = [
    'customer', 'package_name', 'channel_name', 'bank_account_number',
    'bank_name', 'upi_vpa', 'ac_holder_name', 'screenshot', 'platform',
    'search_for', 'status', 'upi_bank_account_wallet', 'priority', 'flag',
    'cessation', 'reviewed_status', 'handle', 'origin', 'payment_gateway_name',
    'category_of_website', 'screenshot_case_report_link',
    'payment_gateway_intermediate_url', 'neft_imps', 'transaction_method',
    'scam_type', 'ifsc_code', 'bank_branch_details', 'payment_gateway_url',
    'upi_url', 'website_url', 'inserted_date', 'reported_earlier',
    'approvd_status', 'feature_type', 'case_generated_time', 'web_contact_no'
]

SHEET_TYPES = {
    'upi': 'UPI_AML',
    'investment': 'Investment_Scam',
}
BANK_NAME_MAPPING = {}
IFSC_MAPPING = {}

ALLOWED_IMPORT_EXTENSIONS = {
    'csv', 'tsv', 'txt',
    'xlsx', 'xlsm', 'xlsb', 'xltx', 'xltm',
    'xls', 'xla', 'xlam',
    'ods', 'ots',
}

def is_allowed_file(filename):
    if not filename:
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in ALLOWED_IMPORT_EXTENSIONS
def read_data_file(file_path, file_ext):
    try:
        ext = file_ext.lower().lstrip('.')
        if ext == 'csv':
            for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    return pd.read_csv(file_path, encoding=encoding)
                except UnicodeDecodeError:
                    continue
            return pd.read_csv(file_path, encoding='latin-1', engine='python')
        if ext == 'tsv':
            for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                try:
                    return pd.read_csv(file_path, sep='\t', encoding=encoding)
                except UnicodeDecodeError:
                    continue
            return pd.read_csv(file_path, sep='\t', encoding='latin-1')
        if ext == 'txt':
            for sep in ['\t', ',', ';', '|']:
                for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
                    try:
                        df = pd.read_csv(file_path, sep=sep, encoding=encoding)
                        if len(df.columns) > 1:
                            return df
                    except Exception:
                        continue
            return pd.read_csv(file_path, encoding='latin-1', engine='python')
        if ext in ('xlsx', 'xlsm', 'xltx', 'xltm'):
            return pd.read_excel(file_path, engine='openpyxl')
        if ext == 'xlsb':
            return pd.read_excel(file_path, engine='pyxlsb')
        if ext in ('xls', 'xla', 'xlam'):
            try:
                return pd.read_excel(file_path, engine='xlrd')
            except Exception:
                return pd.read_excel(file_path, engine='openpyxl')
        if ext in ('ods', 'ots'):
            return pd.read_excel(file_path, engine='odf')
        try:
            return pd.read_excel(file_path)
        except Exception:
            return pd.read_csv(file_path, encoding='latin-1')
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        raise
def load_excel_data():
    global BANK_NAME_MAPPING, IFSC_MAPPING
    try:
        if BANK_NAME_MAPPING_PATH.exists():
            df_bank = pd.read_excel(BANK_NAME_MAPPING_PATH)
            BANK_NAME_MAPPING = {}
            key_col = next((c for c in df_bank.columns if any(k in str(c).lower() for k in ['key', 'handle', 'upi'])), df_bank.columns[0] if len(df_bank.columns) > 0 else None)
            bank_col = next((c for c in df_bank.columns if 'bank' in str(c).lower() and 'name' in str(c).lower()), df_bank.columns[1] if len(df_bank.columns) > 1 else None)
            if key_col and bank_col:
                for _, row in df_bank.iterrows():
                    k = str(row.get(key_col, '')).strip().lower()
                    v = str(row.get(bank_col, 'NA')).strip()
                    if k and k not in ['na', 'nan', 'none', 'null', '']:
                        BANK_NAME_MAPPING[k] = v
        if IFSC_MAPPING_PATH.exists():
            df_ifsc = pd.read_excel(IFSC_MAPPING_PATH)
            IFSC_MAPPING = {}
            prefix_col = next((c for c in df_ifsc.columns if any(k in str(c).lower() for k in ['ifsc', 'prefix', 'code'])), df_ifsc.columns[0] if len(df_ifsc.columns) > 0 else None)
            bank_col2 = next((c for c in df_ifsc.columns if 'bank' in str(c).lower() and 'name' in str(c).lower()), df_ifsc.columns[1] if len(df_ifsc.columns) > 1 else None)
            if prefix_col and bank_col2:
                for _, row in df_ifsc.iterrows():
                    k = str(row.get(prefix_col, '')).strip().upper()
                    v = str(row.get(bank_col2, 'NA')).strip()
                    if k and k.lower() not in ['na', 'nan', 'none', 'null', '']:
                        IFSC_MAPPING[k] = v
    except Exception as e:
        print(f"Error loading Excel data: {e}")
        BANK_NAME_MAPPING = {}
        IFSC_MAPPING = {}
load_excel_data()
def load_config():
    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        return create_default_config()
    except Exception as e:
        print(f"Error loading config: {e}")
        return create_default_config()
def create_default_config():
    default_config = {
        "sheet_mappings": {
            "upi": {
                "name": "UPI (AML)",
                "required_headers": ["UPI", "Screenshot", "Website URL", "Payment Gateway URL"],
                "column_mapping": {
                    "UPI": ["upi_vpa", "upi"],
                    "Screenshot": ["screenshot", "image", "proof"],
                    "Website URL": ["website_url", "url", "website"],
                    "Payment Gateway URL": ["payment_gateway_url", "payment_url", "gateway"],
                    "Transaction Method": ["transaction_method", "payment_method", "method"]
                }
            },
            "investment": {
                "name": "Investment Scam",
                "required_headers": ["UPI", "Account Holder Name", "Bank Account Number", "IFSC Code",
                                     "Website URL", "Payment Gateway URL",
                                     "Screenshot", "Contact Number", "Scam Type"],
                "column_mapping": {
                    "UPI": ["upi_vpa", "upi"],
                    "Account Holder Name": ["ac_holder_name", "account_holder", "holder_name", "customer"],
                    "Bank Account Number": ["bank_account_number", "account_number", "acc_no"],
                    "IFSC Code": ["ifsc_code", "ifsc", "bank_code"],
                    "Website URL": ["website_url", "url", "website"],
                    "Payment Gateway URL": ["payment_gateway_url", "payment_url", "gateway"],
                    "Transaction Method": ["transaction_method", "payment_method", "method"],
                    "Screenshot": ["screenshot", "image", "proof"],
                    "Contact Number": ["web_contact_no", "contact_number", "phone", "mobile"],
                    "Scam Type": ["scam_type", "type", "category"]
                }
            },
        },
        "global_settings": {
            "date_format": "%Y-%m-%d",
            "na_values": ["NA", "N/A", "", "null", "NULL", "None", "undefined"],
            "allowed_extensions": list(ALLOWED_IMPORT_EXTENSIONS),
            "max_file_size_mb": 50
        }
    }
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(default_config, f, indent=2)
    return default_config
def get_sheet_headers(sheet_type):
    config = load_config()
    if not config:
        return []
    sheet_config = config['sheet_mappings'].get(sheet_type)
    if not sheet_config:
        return []
    return sheet_config.get('required_headers', [])
def standardize_headers(headers, sheet_type):
    config = load_config()
    if not config:
        return headers
    sheet_config = config['sheet_mappings'].get(sheet_type)
    if not sheet_config:
        return headers
    standardized = []
    column_mapping = sheet_config.get('column_mapping', {})
    for header in headers:
        header_lower = str(header).lower().strip()
        mapped = False
        for target_col, source_cols in column_mapping.items():
            for source_col in source_cols:
                if header_lower == source_col.lower():
                    standardized.append(target_col)
                    mapped = True
                    break
            if mapped:
                break
        if not mapped:
            for target_col, source_cols in column_mapping.items():
                for source_col in source_cols:
                    if source_col.lower() in header_lower or header_lower in source_col.lower():
                        standardized.append(target_col)
                        mapped = True
                        break
                if mapped:
                    break
        if not mapped:
            standardized.append(header)
    return standardized
def clean_value(value):
    if pd.isna(value) or value in ["NA", "", None, "null", "NULL", "None", "undefined"]:
        return "NA"
    value_str = str(value).strip()
    if value_str.lower() in ("nan", "none", "null", "na", "n/a", "undefined", ""):
        return "NA"
    # .0 remove karo numeric strings se (e.g. "8815336405.0" -> "8815336405")
    if value_str.endswith('.0'):
        try:
            int_val = int(float(value_str))
            value_str = str(int_val)
        except (ValueError, OverflowError):
            pass
    value_str = ''.join(char for char in value_str if ord(char) < 0x10000)
    return value_str
def extract_handle(upi_vpa):
    upi_vpa = clean_value(upi_vpa)
    if upi_vpa == "NA":
        return "NA"
    if '@' in upi_vpa:
        handle_part = upi_vpa.split('@')[1]
        if '.' in handle_part:
            handle_part = handle_part.split('.')[0]
        handle = handle_part.lower().strip()
        if not handle or handle in ('nan', 'none', 'null', 'na', 'n/a', ''):
            return "NA"
        return handle
    return "NA"
def get_bank_name_from_handle(handle, ifsc_code=None):
    if handle != "NA" and handle:
        handle_lower = handle.lower().strip()
        if handle_lower in BANK_NAME_MAPPING:
            return BANK_NAME_MAPPING[handle_lower]
        for key, value in BANK_NAME_MAPPING.items():
            if key in handle_lower or handle_lower in key:
                return value
        common_mappings = {
            'okaxis': 'NA'
        }
        for pattern, bank_name in common_mappings.items():
            if pattern in handle_lower:
                return bank_name
    if ifsc_code and ifsc_code != "NA":
        try:
            ifsc_prefix = ifsc_code[:4].upper()
            if ifsc_prefix in IFSC_MAPPING:
                return IFSC_MAPPING[ifsc_prefix]
            common_ifsc = {
                'SBIN': 'State Bank of India', 'ICIC': 'ICICI Bank',
                'HDFC': 'HDFC Bank', 'UTIB': 'Axis Bank', 'CNRB': 'Canara Bank',
                'BARB': 'Bank of Baroda', 'BKID': 'Bank of India',
                'PUNB': 'Punjab National Bank', 'UBIN': 'Union Bank of India',
                'INDB': 'IndusInd Bank', 'YESB': 'Yes Bank',
                'KARB': 'Karnataka Bank', 'FDRL': 'Federal Bank',
                'IDFB': 'IDFC First Bank', 'RATN': 'RBL Bank'
            }
            if ifsc_prefix in common_ifsc:
                return common_ifsc[ifsc_prefix]
        except Exception as e:
            print(f"Error in IFSC lookup for {ifsc_code}: {e}")
    return "NA"
def extract_search_for_from_url(url):
    url_value = clean_value(url)
    if url_value == "NA":
        return "Platform"
    try:
        parsed_url = urlparse(url_value.lower())
        domain = parsed_url.netloc
        if not domain:
            return "Platform"
        domain_without_www = domain[4:] if domain.startswith('www.') else domain
        platform_domains = {
            't.me': 'Telegram', 'wa.me': 'WhatsApp',
            'chat.whatsapp.com': 'WhatsApp', 'facebook.com': 'Facebook',
            'instagram.com': 'Instagram', 'telegram.org': 'Telegram','web.telegram.org': 'Telegram','telegram.me': 'Telegram',
            'threads.com': 'Thread', 'youtube.com': 'YouTube', 'x.com': 'X'
        }
        if domain in platform_domains:
            return platform_domains[domain]
        if domain_without_www in platform_domains:
            return platform_domains[domain_without_www]
        if url_value.lower().startswith(('https://', 'http://')):
            return "Web"
        return "Platform"
    except Exception:
        return "Platform"

def bulk_lookup_origin_category(urls: list) -> dict:
    result = {}
    unique_urls = [u for u in set(urls) if u and u.upper() not in ("NA", "N/A", "")]
    if not unique_urls:
        return result
    # Build domain index for fallback
    domain_map = {}
    for u in unique_urls:
        try:
            netloc = urlparse(u).netloc
            if netloc:
                domain = netloc[4:] if netloc.startswith("www.") else netloc
                if domain not in domain_map:
                    domain_map[domain] = u
        except Exception:
            pass

    CHUNK = 200
    try:
        all_rows = []
        offset = 0
        while True:
            resp = supabase.table("website_directory") \
                .select("url,final_url,origin,category") \
                .order("id", desc=False) \
                .range(offset, offset + CHUNK - 1).execute()
            rows = resp.data or []
            all_rows.extend(rows)
            if len(rows) < CHUNK:
                break
            offset += CHUNK
        # Build lookup maps from DB
        url_exact   = {}   # url_lower → (origin, category)
        domain_db   = {}   # domain    → (origin, category)
        for row in all_rows:
            origin   = (row.get("origin")   or "NA").strip() or "NA"
            category = (row.get("category") or "NA").strip() or "NA"
            for col in ("url", "final_url"):
                val = (row.get(col) or "").strip().lower()
                if val and val not in ("na", "n/a", ""):
                    url_exact[val] = (origin, category)
                    try:
                        netloc = urlparse(val).netloc
                        if netloc:
                            dom = netloc[4:] if netloc.startswith("www.") else netloc
                            if dom not in domain_db:
                                domain_db[dom] = (origin, category)
                    except Exception:
                        pass
        # Match each requested URL
        for u in unique_urls:
            u_lower = u.lower().strip()
            if u_lower in url_exact:
                result[u_lower] = url_exact[u_lower]
                continue
            # Try http/https swap
            for old, new in [("https://", "http://"), ("http://", "https://")]:
                if u_lower.startswith(old):
                    alt = new + u_lower[len(old):]
                    if alt in url_exact:
                        result[u_lower] = url_exact[alt]
                        break
            if u_lower in result:
                continue
            # Domain fallback
            try:
                netloc = urlparse(u_lower).netloc
                if netloc:
                    dom = netloc[4:] if netloc.startswith("www.") else netloc
                    if dom in domain_db:
                        result[u_lower] = domain_db[dom]
            except Exception:
                pass
    except Exception as e:
        print(f"[bulk_lookup_origin_category] Error: {e}")
    return result
def extract_case_time_and_date_from_npci_url(url):
    if not url or url == "NA":
        return "NA", "NA"
    match = re.search(r'npci-(\d{10})_', url)
    if not match:
        return "NA", "NA"
    try:
        ts = int(match.group(1))
        utc_dt = datetime.fromtimestamp(ts, timezone.utc).replace(tzinfo=None)
        ist_dt = utc_dt + timedelta(hours=5, minutes=30)
        return ist_dt.strftime("%Y-%m-%d %H:%M:%S"), ist_dt.strftime("%Y-%m-%d")
    except Exception:
        return "NA", "NA"
def generate_screenshot_urls(screenshot_url):
    screenshot_value = clean_value(screenshot_url)
    if screenshot_value == "NA":
        return "NA"
    try:
        parsed_url = urlparse(screenshot_value)
        path = parsed_url.path
        if not path:
            return "NA"
        filename = path.split('/')[-1]
        if not filename:
            return "NA"
        if '-' in filename:
            parts = filename.split('-', 1)
            if len(parts) == 2:
                _, rest_of_filename = parts
                urls = []
                for new_prefix in ['mfilterit', 'npci', 'without_header']:
                    new_filename = f"{new_prefix}-{rest_of_filename}"
                    new_path = '/'.join(path.split('/')[:-1] + [new_filename])
                    new_url = f"{parsed_url.scheme}://{parsed_url.netloc}{new_path}"
                    urls.append(new_url)
                return ','.join(urls)
        return screenshot_value
    except Exception as e:
        return screenshot_value
def extract_payment_gateway_name(upi_url, website_url):
    upi_url_value = clean_value(upi_url)
    website_url_value = clean_value(website_url)
    if upi_url_value == "NA":
        return "NA"
    try:
        parsed_upi = urlparse(upi_url_value)
        upi_domain = parsed_upi.netloc
        if not upi_domain:
            path = parsed_upi.path.lstrip('/')
            domain_part = path.split('/')[0]
            upi_domain = domain_part if '.' in domain_part else None
            if not upi_domain:
                return "NA"
        upi_domain_clean = upi_domain[4:] if upi_domain.startswith('www.') else upi_domain
        if website_url_value == "NA":
            return upi_domain
        parsed_website = urlparse(website_url_value)
        website_domain = parsed_website.netloc
        if not website_domain:
            path = parsed_website.path.lstrip('/')
            domain_part = path.split('/')[0]
            website_domain = domain_part if '.' in domain_part else None
            if not website_domain:
                return upi_domain
        website_domain_clean = website_domain[4:] if website_domain.startswith('www.') else website_domain
        return "NA" if upi_domain_clean == website_domain_clean else upi_domain
    except Exception as e:
        return "NA"
def validate_input_columns(df, sheet_type):
    """
    Check ki uploaded file ke columns EXACTLY match karte hain
    sheet ke 'required_headers' (Download Template) se.
    Returns: (is_valid, missing_cols, extra_cols)
    """
    config = load_config()
    sheet_config = config['sheet_mappings'].get(sheet_type, {})
    required_headers = sheet_config.get('required_headers', [])
    column_mapping = sheet_config.get('column_mapping', {})

    header_aliases = {}
    for target_col in required_headers:
        aliases = set()
        aliases.add(target_col.lower().strip())
        for sc in column_mapping.get(target_col, []):
            aliases.add(sc.lower().strip())
        header_aliases[target_col] = aliases

    input_cols_lower = [str(c).lower().strip() for c in df.columns]

    missing_cols = []
    for target_col, aliases in header_aliases.items():
        found = any(
            ic == alias or alias in ic or ic in alias
            for ic in input_cols_lower
            for alias in aliases
        )
        if not found:
            missing_cols.append(target_col)

    all_known_aliases = set()
    for aliases in header_aliases.values():
        all_known_aliases.update(aliases)

    extra_cols = []
    for col, col_lower in zip(df.columns, input_cols_lower):
        matched = any(
            col_lower == alias or alias in col_lower or col_lower in alias
            for alias in all_known_aliases
        )
        if not matched:
            extra_cols.append(str(col))

    is_valid = (len(missing_cols) == 0 and len(extra_cols) == 0)
    return is_valid, missing_cols, extra_cols
def process_sheet_data(df, sheet_type):
    result_df = pd.DataFrame(columns=REQUIRED_COLUMNS)
    if df.empty:
        return result_df, {'total_values': 0, 'unique_upi_ids': 0, 'unique_bank_accounts': 0, 'unique_websites': 0}
    input_headers = list(df.columns)
    standardized_headers = standardize_headers(input_headers, sheet_type)
    df.columns = standardized_headers
    unique_upi_ids = set()
    unique_bank_accounts = set()
    unique_websites = set()
    # ── Pre-fetch origin/category for all website URLs in one bulk call ──
    _website_url_col = None
    for h in standardized_headers:
        if h == "Website URL":
            _website_url_col = h
            break
    _wd_cache = {}
    if _website_url_col and _website_url_col in df.columns:
        _all_urls = [
            clean_value(df.iloc[i][_website_url_col])
            for i in range(len(df))
        ]
        _valid_urls = [u for u in _all_urls if u != "NA"]
        if _valid_urls:
            _wd_cache = bulk_lookup_origin_category(_valid_urls)
    for idx in range(len(df)):
        row_data = {col: "NA" for col in REQUIRED_COLUMNS}
        row_data['case_generated_time'] = "NA"
        row_data['inserted_date'] = "NA"
        for std_header in standardized_headers:
            value = df.iloc[idx][std_header]
            cleaned_value = clean_value(value)
            if std_header == "UPI":
                row_data['upi_vpa'] = cleaned_value
                if cleaned_value != "NA":
                    unique_upi_ids.add(cleaned_value)
            elif std_header == "Account Holder Name":
                row_data['ac_holder_name'] = cleaned_value
            elif std_header == "Bank Account Number":
                row_data['bank_account_number'] = cleaned_value
                if cleaned_value != "NA":
                    unique_bank_accounts.add(cleaned_value)
            elif std_header == "IFSC Code":
                row_data['ifsc_code'] = cleaned_value
            elif std_header == "Website URL":
                row_data['website_url'] = cleaned_value
                if cleaned_value != "NA":
                    unique_websites.add(cleaned_value)
            elif std_header == "Payment Gateway URL":
                row_data['payment_gateway_url'] = cleaned_value
            elif std_header == "Transaction Method":
                row_data['transaction_method'] = cleaned_value
            elif std_header == "Screenshot":
                row_data['_original_screenshot'] = cleaned_value
                case_time, inserted_date = extract_case_time_and_date_from_npci_url(cleaned_value)
                row_data['case_generated_time'] = case_time
                row_data['inserted_date'] = inserted_date
                row_data['screenshot'] = generate_screenshot_urls(cleaned_value)
            elif std_header == "Contact Number":
                row_data['web_contact_no'] = cleaned_value
            elif std_header == "Scam Type":
                row_data['scam_type'] = cleaned_value
            elif std_header == "Category":
                row_data['category_of_website'] = cleaned_value
        if sheet_type == 'upi':
            row_data.update({
                'customer': "Mystery Shopping", 'package_name': "com.mysteryshopping",
                'channel_name': "Organic Search", 'status': "Active", 'priority': "High",
                'flag': "1", 'cessation': "Open", 'reviewed_status': "1",
                'reported_earlier': "No", 'approvd_status': "1",
                'feature_type': "BS Money Laundering", 'platform': "NA",
                'neft_imps': "NA", 'bank_branch_details': "NA", 'scam_type': "NA",'transaction_method': "NA"
            })
            row_data['upi_bank_account_wallet'] = "UPI" if row_data['upi_vpa'] != "NA" else "Bank Account"
            if row_data['website_url'] != "NA":
                _key = row_data['website_url'].lower().strip()
                origin, category = _wd_cache.get(_key, ("NA", "NA"))
                row_data['origin'] = origin
                row_data['category_of_website'] = category
            else:
                row_data['origin'] = "NA"
                row_data['category_of_website'] = "NA"

        elif sheet_type == 'investment':
            row_data.update({
                'customer': "Mystery Shopping", 'package_name': "com.mysteryshopping",
                'channel_name': "Organic Search", 'status': "Active", 'priority': "High",
                'flag': "1", 'cessation': "Open", 'reviewed_status': "1",
                'reported_earlier': "No", 'approvd_status': "1",
                'feature_type': "BS Investment Scam", 'platform': "NA",
                'neft_imps': "NA", 'bank_branch_details': "NA",'transaction_method': "NA"
            })
            row_data['upi_bank_account_wallet'] = "UPI" if row_data['upi_vpa'] != "NA" else "Bank Account"
            if row_data['scam_type'] != "NA" and row_data['category_of_website'] == "NA":
                row_data['category_of_website'] = row_data['scam_type']
            if row_data['website_url'] != "NA":
                _key = row_data['website_url'].lower().strip()
                origin, _ = _wd_cache.get(_key, ("NA", "NA"))
                row_data['origin'] = origin
            else:
                row_data['origin'] = "NA"
        handle = extract_handle(row_data['upi_vpa'])
        row_data['handle'] = handle
        row_data['bank_name'] = get_bank_name_from_handle(handle, row_data['ifsc_code'])
        row_data['search_for'] = extract_search_for_from_url(row_data['website_url'])
        if sheet_type != 'messaging' and row_data['category_of_website'] == "NA":
            if row_data['website_url'] != "NA":
                _key = row_data['website_url'].lower().strip()
                _, category = _wd_cache.get(_key, ("NA", "NA"))
                row_data['category_of_website'] = category
        row_data['screenshot_case_report_link'] = row_data.get('screenshot', "NA")
        row_data.pop('_original_screenshot', None)
        payment_gateway_url = row_data.get('payment_gateway_url', "NA")
        if payment_gateway_url != "NA":
            row_data['payment_gateway_intermediate_url'] = payment_gateway_url
            row_data['upi_url'] = payment_gateway_url
            row_data['payment_gateway_name'] = extract_payment_gateway_name(
                row_data['upi_url'], row_data['website_url']
            )
        else:
            row_data['payment_gateway_intermediate_url'] = "NA"
            row_data['upi_url'] = "NA"
            row_data['payment_gateway_name'] = "NA"
        result_df.loc[idx] = [row_data.get(col, "NA") for col in REQUIRED_COLUMNS]
        result_df = result_df.fillna("NA")
        result_df = result_df.fillna("NA")
        result_df = result_df.replace("nan", "NA")
        result_df = result_df.replace("NaN", "NA")
    return result_df, {
        'total_values': len(result_df),
        'unique_upi_ids': len(unique_upi_ids),
        'unique_bank_accounts': len(unique_bank_accounts),
        'unique_websites': len(unique_websites)
    }
# ============================================================
# Helper function to extract clean display name
# ============================================================
def get_clean_display_name(display_name):
    """Extract display name without parentheses content"""
    if not display_name:
        return "User"
    clean_name = re.sub(r'\s*\([^)]*\)', '', display_name).strip()
    return clean_name if clean_name else display_name

def redirect_to_allowed_page(allowed_pages):
    """Allowed_pages ke first entry ke hisaab se sahi URL pe redirect karta hai."""
    if not allowed_pages:
        return redirect("/login")
    first_page = allowed_pages[0]
    if first_page == "website_directory":
        return redirect("/website-directory")
    if first_page == "qc":
        return redirect("/qc-gui")
    if first_page == "lunch":
        return redirect("/lunch-break")
    if first_page == "case_report":
        return redirect("/case-report")
    if first_page in ("allotment", "allotment_admin"):
        return redirect("/scam-website-allotment")
    if first_page == "dashboard_management":
        return redirect("/dashboard-management")
    return redirect(f"/?page={first_page}")
# ============================================================
# LOGIN / LOGOUT
# ============================================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect("/")
    error = None
    prefill_email = ""
    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        prefill_email = email
        if not email or not password:
            error = "Please enter both email and password."
        else:
            user = fetch_user_by_email(email)
            if user and user.get("password") == password:
                session.permanent = True
                session["user_id"]               = user["id"]
                session["email"]                 = user["email"]
                session["display_name"]          = user["display_name"]
                session["allowed_pages"]         = user.get("allowed_pages") or []
                session["is_admin"]              = parse_bool(user.get("is_admin", False))
                session["role"]                  = (user.get("role")
                                                  or ("admin" if parse_bool(user.get("is_admin", False)) else "user"))
                session["can_view_activity_log"] = parse_bool(user.get("can_view_activity_log", False))
                session["allowed_departments"] = user.get("allowed_departments") or None
                allowed = session["allowed_pages"]
                first_page = allowed[0] if allowed else "scraping"
                return redirect(f"/?page={first_page}")
            else:
                error = "Invalid email or password. Please try again."
    return render_template("login.html", error=error, prefill_email=prefill_email)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")
# ============================================================
# USER ACTIVITY LOG ROUTES
# ============================================================
@app.route("/get-user-activity-log", methods=["GET"])
@login_required
def get_user_activity_log():
    if not session.get("can_view_activity_log"):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        client = get_auth_supabase()
        is_super       = is_superadmin(session)
        allowed_depts = session.get("allowed_departments")  # None = no dept restriction
        allowed_pages = session.get("allowed_pages", [])

        ual_date_from = request.args.get("date_from", "").strip()
        ual_date_to   = request.args.get("date_to", "").strip()

        query = client.table("activity_logs").select("*")
        if ual_date_from:
            query = query.gte("created_at", ual_date_from + " 00:00:00")
        if ual_date_to:
            query = query.lte("created_at", ual_date_to + " 23:59:59")
        resp = query.order("created_at", desc=True).limit(500).execute()
        all_logs = resp.data or []
        PAGE_TABLE_MAP = {
            "scraping":        "scrapping_data",
            "social":          "social_media_accounts",
            "investment":      "BS_Investment_Scam",
            "allotment":       "website_allotment",
            "allotment_admin": "website_allotment",
        }

        # Superadmin — sab kuch dikhao (but still filter by allowed_pages)
        if is_super:
            allowed_tables = set(PAGE_TABLE_MAP[p] for p in allowed_pages if p in PAGE_TABLE_MAP)
            if "website_directory" in allowed_pages:
                allowed_tables.add("website_directory")

            def _admin_log_allowed(log):
                target_table = log.get("target_table") or ""
                return target_table in allowed_tables

            logs = list(filter(_admin_log_allowed, all_logs))
            return jsonify({"success": True, "logs": logs})

        # Non-admin — allowed_pages ke basis pe tables set karo
        allowed_tables = set()
        for page in allowed_pages:
            if page in PAGE_TABLE_MAP:
                allowed_tables.add(PAGE_TABLE_MAP[page])
        # website_directory sirf tab dikhao jab investment allow ho
        if "website_directory" in allowed_pages:
            allowed_tables.add("website_directory")

        current_email = session.get("email", "")

        def _log_allowed(log):
            target_table = log.get("target_table") or ""
            action_type  = log.get("action_type") or ""

            # Agar table allowed nahi hai toh hide karo
            if not target_table or target_table not in allowed_tables:
                return False

            # Department restriction bhi hai toh extra check
            if allowed_depts and target_table == "social_media_accounts":
                # Apni khud ki activity hamesha dikhao
                if log.get("user_email") == current_email:
                    return True
                # Doosron ki import activity hide karo
                if action_type == "import":
                    return False
                # Field update mein department check karo
                if action_type == "field_update":
                    extra = log.get("extra_info") or {}
                    if isinstance(extra, str):
                        try:
                            import json as _j
                            extra = _j.loads(extra)
                        except Exception:
                            extra = {}
                    dept = extra.get("department", "")
                    return bool(dept and dept in allowed_depts)
                return False
            return True
        logs = list(filter(_log_allowed, all_logs))
        print(f"[UAL DEBUG] allowed_pages={allowed_pages}")
        print(f"[UAL DEBUG] allowed_tables={allowed_tables}")
        print(f"[UAL DEBUG] total logs={len(all_logs)}, filtered={len(logs)}")
        # Sample target_tables from all_logs
        sample_tables = list(set(l.get("target_table") for l in all_logs[:20]))
        print(f"[UAL DEBUG] sample target_tables={sample_tables}")
        return jsonify({"success": True, "logs": logs})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/export-user-activity-log", methods=["GET"])
@login_required
def export_user_activity_log():
    """Export user activity log as CSV"""
    if not session.get("can_view_activity_log"):
        flash("Access denied.", "error")
        return redirect("/")
    try:
        client = get_auth_supabase()
        resp = client.table("activity_logs") \
            .select("*") \
            .order("created_at", desc=True) \
            .execute()
        logs = resp.data or []
        if not logs:
            flash("No activity logs to export.", "error")
            return redirect("/")
        df = pd.DataFrame(logs)
        column_mapping = {
            'id': 'ID',
            'user_id': 'User ID',
            'user_email': 'User Email',
            'display_name': 'Login User Name',
            'action_type': 'Action Type',
            'target_table': 'Target Table',
            'target_record_id': 'Target Record ID',
            'field_name': 'Field Name',
            'old_value': 'Previous Value',
            'new_value': 'Updated Value',
            'extra_info': 'Extra Info',
            'created_at': 'Timestamp'
        }
        available_columns = [col for col in column_mapping.keys() if col in df.columns]
        df = df[available_columns]
        df = df.rename(columns=column_mapping)
        if 'Timestamp' in df.columns:
            df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"user_activity_log_{timestamp}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            download_name=filename,
            as_attachment=True,
            mimetype="text/csv"
        )
    except Exception as e:
        flash(f"Export Error: {str(e)}", "error")
        return redirect("/")
# ============================================================
# SCRAPING TRACKER STATS
# ============================================================
@app.route("/scraping-tracker-stats", methods=["GET"])
@login_required
def scraping_tracker_stats():
    try:
        CHUNK = 1000
        rows = []
        offset = 0
        is_super       = is_superadmin(session)
        current_clean_name = get_clean_display_name(session.get("display_name", ""))
        while True:
            _q = supabase.table("scrapping_data").select("scam_type,platform")
            if not is_super:
                _q = _q.eq("name", current_clean_name)
            resp = _q \
                .order("id", desc=False) \
                .range(offset, offset + CHUNK - 1) \
                .execute()
            chunk = resp.data or []
            rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        scam_counts = {}
        platform_counts = {}
        scam_platform_breakdown = {}
        for row in rows:
            st = (row.get("scam_type") or "Unknown").strip()
            if not st or st in ("NA", "N/A", ""):
                st = "Unknown"
            p = (row.get("platform") or "Unknown").strip()
            if not p or p in ("NA", "N/A", ""):
                p = "Unknown"
            scam_counts[st] = scam_counts.get(st, 0) + 1
            platform_counts[p] = platform_counts.get(p, 0) + 1
            if st not in scam_platform_breakdown:
                scam_platform_breakdown[st] = {}
            scam_platform_breakdown[st][p] = scam_platform_breakdown[st].get(p, 0) + 1
        total = len(rows)
        return jsonify({
            "success": True,
            "stats": {
                "scam_counts": scam_counts,
                "platform_counts": platform_counts,
                "scam_platform_breakdown": scam_platform_breakdown,
                "total": total
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# MAIN DASHBOARD ROUTE
# ============================================================
@app.route("/", methods=["GET"])
@login_required
def index():
    user = get_current_user()
    allowed_pages = session.get("allowed_pages", [])
    page_type = request.args.get("page", "").strip()
    if not page_type or page_type not in allowed_pages:
        page_type = allowed_pages[0] if allowed_pages else "scraping"
    # Dedicated pages — apne route pe redirect karo
    if page_type == "qc":
        return redirect("/qc-gui")
    if page_type == "website_directory":
        return redirect("/website-directory")
    if page_type == "lunch_break":
        return redirect("/lunch-break")
    if page_type == "case_report":
        return redirect("/case-report")
    if page_type in ("allotment", "allotment_admin"):
        return redirect("/scam-website-allotment")
    search_query = request.args.get("search", "").strip()
    scam_filter = request.args.get("scam_type", "").strip()
    platform_filter = request.args.get("platform", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    date_filter = request.args.get("date_filter", "").strip()
    page = int(request.args.get("page_num", 1))
    social_search = request.args.get("social_search", "").strip()
    social_platform = request.args.get("social_platform", "").strip()
    social_permanent_block = request.args.get("permanent_block", "").strip()
    social_status_filter = request.args.get("social_status", "").strip()
    social_department_filter = request.args.get("social_department", "").strip()
    inv_search = request.args.get("inv_search", "").strip()
    inv_scam_type = request.args.get("inv_scam_type", "").strip()
    inv_search_for = request.args.get("inv_search_for", "").strip()
    inv_wallet = request.args.get("inv_wallet", "").strip()
    inv_date_from = request.args.get("inv_date_from", "").strip()
    inv_date_to = request.args.get("inv_date_to", "").strip()
    items = []
    total_rows = 0
    total_pages = 1
    if page_type == "scraping":
        try:
            query = supabase.table("scrapping_data").select("*", count='exact')
            if not is_superadmin(session):
                current_clean_name = get_clean_display_name(session.get("display_name", ""))
                query = query.eq("name", current_clean_name)
            if search_query:
                like_term = f"%{search_query}%"
                query = query.or_(f"name.ilike.{like_term},platform.ilike.{like_term},post_url.ilike.{like_term},chat_number.ilike.{like_term},group_name.ilike.{like_term},chat_link.ilike.{like_term},scam_type.ilike.{like_term}")
            if scam_filter:
                query = query.eq("scam_type", scam_filter)
            if platform_filter:
                query = query.eq("platform", platform_filter)
            if date_from:
                query = query.gte("inserted_date", date_from)
            if date_to:
                query = query.lte("inserted_date", date_to)
            if date_filter and not date_from and not date_to:
                query = query.eq("inserted_date", date_filter)
            share_status_filter = request.args.get("share_status", "").strip()
            if share_status_filter:
                query = query.eq("share_status", share_status_filter)
            query = query.order("id", desc=True)
            offset = (page - 1) * PER_PAGE
            query = query.range(offset, offset + PER_PAGE - 1)
            response = query.execute()
            items = response.data or []
            total_rows = response.count or 0
            total_pages = max(1, math.ceil(total_rows / PER_PAGE)) if total_rows else 1
        except Exception as e:
            print(f"[DEBUG] Scraping error: {e}")
            items = []
            total_rows = 0
            total_pages = 1
            flash(f"Error fetching scraping data: {str(e)}", "error")
    elif page_type == "social":
        try:
            query = social_supabase.table("social_media_accounts").select("*", count='exact')
            allowed_depts = session.get("allowed_departments")
            if allowed_depts:  # None = see all, list = restricted
                if len(allowed_depts) == 1:
                    query = query.eq("department", allowed_depts[0])
                else:
                    query = query.in_("department", allowed_depts)
            if social_search:
                like_term = f"%{social_search}%"
                query = query.or_(
                    f"login_user.ilike.{like_term},"
                    f"number.ilike.{like_term},"
                    f"full_name.ilike.{like_term},"
                    f"page_name.ilike.{like_term},"
                    f"platform.ilike.{like_term},"
                    f"account_status.ilike.{like_term},"
                    f"owned_by.ilike.{like_term},"
                    f"department.ilike.{like_term},"
                    f"sim_operator.ilike.{like_term},"
                    f"mail_id.ilike.{like_term},"
                    f"account_id.ilike.{like_term},"
                    f"number_type.ilike.{like_term},"
                    f"login_device.ilike.{like_term},"
                    f"sim_inserted_device.ilike.{like_term}"
                )
            if social_platform and social_platform != "":
                query = query.eq("platform", social_platform)
            if social_department_filter:
                query = query.eq("department", social_department_filter)
            if social_permanent_block == "true":
                query = query.eq("account_status", "Permanent Block")
            else:
                if social_status_filter:
                    query = apply_social_status_filter(query, social_status_filter)
                else:
                    query = query.neq("account_status", "Permanent Block")
            query = query.order("id", desc=False)
            offset = (page - 1) * PER_PAGE
            query = query.range(offset, offset + PER_PAGE - 1)
            response = query.execute()
            items = [normalize_social_account_row(row) for row in (response.data or [])]
            total_rows = response.count or 0
            total_pages = max(1, math.ceil(total_rows / PER_PAGE)) if total_rows else 1
            print(f"[DEBUG] Social items: {len(items)}, total: {total_rows}")
        except Exception as e:
            print(f"[DEBUG] Social error: {e}")
            items = []
            total_rows = 0
            total_pages = 1
            flash(f"Error fetching social media data: {str(e)}", "error")
    elif page_type == "investment":
        try:
            if inv_search:
                query = supabase.table("BS_Investment_Scam").select("*")
            else:
                query = supabase.table("BS_Investment_Scam").select("*", count='exact')
            if inv_search:
                like_term = f"%{inv_search}%"
                query = query.or_(
                    f"Bank_account_number.ilike.{like_term},"
                    f"Upi_vpa.ilike.{like_term},"
                    f"Handle.ilike.{like_term},"
                    f"Website_url.ilike.{like_term},"
                    f"Web_contact_no.ilike.{like_term},"
                    f"Input_user.ilike.{like_term}"
                )
            if inv_scam_type:
                query = query.eq("Scam_type", inv_scam_type)
            if inv_search_for:
                query = query.eq("Search_for", inv_search_for)
            if inv_wallet:
                query = query.eq("Upi_bank_account_wallet", inv_wallet)
            if inv_date_from:
                query = query.gte("Inserted_date", inv_date_from)
            if inv_date_to:
                query = query.lte("Inserted_date", inv_date_to)
            query = query.order("Id", desc=True)
            offset = (page - 1) * PER_PAGE
            query = query.range(offset, offset + PER_PAGE - 1)
            response = query.execute()
            raw = response.data or []
            items = [{k.lower(): v for k, v in row.items()} for row in raw]
            if inv_search:
                total_rows = offset + len(raw)
                total_pages = page + 1 if len(raw) == PER_PAGE else max(1, page)
            else:
                total_rows = response.count or 0
                total_pages = max(1, math.ceil(total_rows / PER_PAGE)) if total_rows else 1
            print(f"[DEBUG] Investment items: {len(items)}, total: {total_rows}")
        except Exception as e:
            print(f"[DEBUG] Investment error: {e}")
            items = []
            total_rows = 0
            total_pages = 1
            flash(f"Error fetching BS Investment Scam data: {str(e)}", "error")
    # Get clean display name for template
    clean_display_name = get_clean_display_name(session.get("display_name", "User"))
    return render_template(
        "index.html",
        page_type=page_type,
        items=items,
        search_query=search_query,
        scam_filter=scam_filter,
        platform_filter=platform_filter,
        date_filter=date_filter,
        date_from=date_from,
        date_to=date_to,
        social_search=social_search,
        social_platform=social_platform,
        social_permanent_block=social_permanent_block,
        social_status_filter=social_status_filter,
        social_department_filter=social_department_filter,
        inv_search=inv_search,
        inv_scam_type=inv_scam_type,
        inv_search_for=inv_search_for,
        inv_wallet=inv_wallet,
        inv_date_from=inv_date_from,
        inv_date_to=inv_date_to,
        page_num=page,
        total_pages=total_pages,
        total_rows=total_rows,
        platform_options=PLATFORM_OPTIONS,
        scam_type_options=SCAM_TYPE_OPTIONS,
        social_platform_options=SOCIAL_PLATFORM_OPTIONS,
        bs_investment_scam_type_options=BS_INVESTMENT_SCAM_TYPE_OPTIONS,
        bs_investment_search_for_options=BS_INVESTMENT_SEARCH_FOR_OPTIONS,
        bs_investment_wallet_options=BS_INVESTMENT_WALLET_OPTIONS,
        department_options=DEPARTMENT_OPTIONS,
        current_user=user,
        allowed_pages=allowed_pages,
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        can_view_activity_log=session.get("can_view_activity_log", False),
        is_admin=session.get("is_admin", False),
        role=session.get("role", "user"),
    )
# ============================================================
# BS Investment Scam Tracker Stats
# ============================================================
@app.route("/investment-tracker-stats", methods=["GET"])
@login_required
def investment_tracker_stats():
    try:
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        CHUNK = 1000
        all_rows = []
        offset = 0
        while True:
            q = supabase.table("BS_Investment_Scam").select("Input_user,Search_for,Scam_type,Inserted_date,Upi_vpa,Bank_account_number,Upi_bank_account_wallet")
            if date_from: q = q.gte("Inserted_date", date_from)
            if date_to: q = q.lte("Inserted_date", date_to)
            resp = q.order("Id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        rows = [{k.lower(): v for k, v in r.items()} for r in all_rows]
        upi_set = set()
        bank_set = set()
        for r in rows:
            wallet = (r.get("upi_bank_account_wallet") or "").strip()
            upi_vpa = (r.get("upi_vpa") or "").strip()
            bank_acc = (r.get("bank_account_number") or "").strip()
            if wallet == "UPI" and upi_vpa and upi_vpa.upper() not in ("NA", "N/A", ""):
                upi_set.add(upi_vpa)
            if wallet == "Bank Account" and bank_acc and bank_acc.upper() not in ("NA", "N/A", ""):
                bank_set.add(bank_acc)
        users_count = {}
        for r in rows:
            user = (r.get("input_user") or "Unknown").strip()
            sf = (r.get("search_for") or "Unknown").strip()
            if user not in users_count: users_count[user] = {}
            users_count[user][sf] = users_count[user].get(sf, 0) + 1
        scam_type_counts = {}
        for r in rows:
            user = (r.get("input_user") or "Unknown").strip()
            st = (r.get("scam_type") or "Unknown").strip()
            if user not in scam_type_counts: scam_type_counts[user] = {}
            scam_type_counts[user][st] = scam_type_counts[user].get(st, 0) + 1
        total_counts = {}
        for r in rows:
            st = (r.get("scam_type") or "Unknown").strip()
            sf = (r.get("search_for") or "Unknown").strip()
            if st not in total_counts: total_counts[st] = {}
            total_counts[st][sf] = total_counts[st].get(sf, 0) + 1
        return jsonify({
            "success": True,
            "total_rows": len(rows),
            "unique_upi_count": len(upi_set),
            "unique_bank_count": len(bank_set),
            "users_count": users_count,
            "scam_type_counts": scam_type_counts,
            "total_counts": total_counts
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/investment-last-date", methods=["GET"])
@login_required
def investment_last_date():
    try:
        resp = supabase.table("BS_Investment_Scam") \
            .select("Inserted_date") \
            .order("Inserted_date", desc=True) \
            .limit(1) \
            .execute()
        if resp.data:
            raw = resp.data[0].get("Inserted_date") or ""
            # Normalise to YYYY-MM-DD
            date_str = str(raw).split("T")[0].strip()
            return jsonify({"success": True, "last_date": date_str})
        return jsonify({"success": True, "last_date": None})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

def _csv_text(rows, fieldnames, include_bom=False, include_header=False):
    output = io.StringIO(newline="")
    if include_bom:
        output.write("\ufeff")
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    if include_header:
        writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def _stream_supabase_csv(build_query, order_column, download_name, chunk_size=1000, row_transform=None):
    def generate():
        offset = 0
        fieldnames = None

        while True:
            resp = build_query().order(order_column, desc=False).range(offset, offset + chunk_size - 1).execute()
            rows = resp.data or []
            if row_transform:
                rows = [row_transform(row) for row in rows]

            if rows:
                if fieldnames is None:
                    fieldnames = list(rows[0].keys())
                    yield _csv_text(rows, fieldnames, include_bom=True, include_header=True)
                else:
                    yield _csv_text(rows, fieldnames)

            if len(rows) < chunk_size:
                break
            offset += chunk_size

        if fieldnames is None:
            yield "\ufeff"

    return Response(
        stream_with_context(generate()),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={download_name}",
            "Cache-Control": "no-cache",
        },
    )

# ============================================================
# BS Investment Scam Export
# ============================================================
@app.route("/investment-export", methods=["GET"])
@login_required
def investment_export():
    try:
        inv_search = request.args.get("inv_search", "").strip()
        inv_scam_type = request.args.get("inv_scam_type", "").strip()
        inv_search_for = request.args.get("inv_search_for", "").strip()
        inv_wallet = request.args.get("inv_wallet", "").strip()
        inv_date_from = request.args.get("inv_date_from", "").strip()
        inv_date_to = request.args.get("inv_date_to", "").strip()
        inv_export_columns = [
            "Id", "Customer", "Package_name", "Channel_name", "Bank_account_number", "Bank_name", "Upi_vpa",
            "Ac_holder_name", "Screenshot", "Platform", "Search_for", "Status", "Upi_bank_account_wallet",
            "Priority", "Flag", "Cessation", "Reviewed_status", "Handle", "Origin", "Payment_gateway_name",
            "Category_of_website", "Screenshot_case_report_link", "Payment_gateway_intermediate_url", "Neft_imps",
            "Transaction_method", "Scam_type", "Ifsc_code", "Bank_branch_details", "Payment_gateway_url",
            "Upi_url", "Website_url", "Inserted_date", "Reported_earlier", "Input_user", "Approvd_status",
            "Approved_by", "Qc_remarks", "Inserted_datetime", "Case_generated_time", "Upi_found_status",
            "Feature_type", "Approved_date", "Video_url", "Web_contact_no"
        ]

        def _build_inv_query():
            q = supabase.table("BS_Investment_Scam").select(",".join(inv_export_columns))
            if inv_search:
                like_term = f"%{inv_search}%"
                q = q.or_(f"Bank_account_number.ilike.{like_term},Upi_vpa.ilike.{like_term},Handle.ilike.{like_term},Website_url.ilike.{like_term},Web_contact_no.ilike.{like_term},Input_user.ilike.{like_term}")
            if inv_scam_type: q = q.eq("Scam_type", inv_scam_type)
            if inv_search_for: q = q.eq("Search_for", inv_search_for)
            if inv_wallet: q = q.eq("Upi_bank_account_wallet", inv_wallet)
            if inv_date_from: q = q.gte("Inserted_date", inv_date_from)
            if inv_date_to: q = q.lte("Inserted_date", inv_date_to)
            return q

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return _stream_supabase_csv(_build_inv_query, "Id", f"bs_investment_scam_{timestamp}.csv")
    except Exception as e:
        flash(f"Export Error: {str(e)}", "error")
        return redirect("/?page=investment")
# ============================================================
# TRACKER STATS
# ============================================================
@app.route("/tracker-stats", methods=["GET"])
@login_required
def tracker_stats():
    try:
        platforms = ["Facebook", "Amazon", "Instagram", "Telegram", "WhatsApp", "Gmail Accounts", "Total Numbers"]
        platform_counts = {}
        platform_status_counts = {}
        perm_block_counts = {}
        perm_block_total = 0
        platform_dept_counts = {}
        platform_number_type_counts = {}
        total_numbers_recharged_count = 0
        CHUNK = 1000
        today = datetime.now().date()
        for platform in platforms:
            try:
                all_rows = []
                offset = 0
                total_count = 0
                while True:
                    _q = social_supabase.table("social_media_accounts") \
                        .select("account_status,department,number_type,recharge_date", count='exact') \
                        .eq("platform", platform)
                    allowed_depts = session.get("allowed_departments")
                    if allowed_depts:
                        if len(allowed_depts) == 1:
                            _q = _q.eq("department", allowed_depts[0])
                        else:
                            _q = _q.in_("department", allowed_depts)
                    resp = _q.range(offset, offset + CHUNK - 1).execute()
                    if offset == 0: total_count = resp.count or 0
                    chunk = resp.data or []
                    all_rows.extend(chunk)
                    if len(chunk) < CHUNK: break
                    offset += CHUNK
                platform_counts[platform] = total_count
                status_map = {}
                pb_count = 0
                dept_map = {}
                num_type_map = {}
                for item in all_rows:
                    status = normalize_social_account_status(item.get('account_status'))
                    dept = (item.get('department') or 'Unknown').strip()
                    num_type = (item.get('number_type') or 'Unknown').strip()
                    if not dept or dept in ('NA', 'N/A', ''):
                        dept = 'Unknown'
                    if not num_type or num_type in ('NA', 'N/A', ''):
                        num_type = 'Unknown'
                    if status == 'Permanent Block':
                        pb_count += 1
                    else:
                        status_map[status] = status_map.get(status, 0) + 1
                    if status != 'Permanent Block':
                        dept_map[dept] = dept_map.get(dept, 0) + 1
                        num_type_map[num_type] = num_type_map.get(num_type, 0) + 1
                    if platform == "Total Numbers":
                        recharge_date = item.get('recharge_date')
                        if recharge_date and str(recharge_date).upper() not in ('NA', 'N/A', 'NONE', 'NULL'):
                            try:
                                recharge_dt = datetime.strptime(str(recharge_date)[:10], "%Y-%m-%d").date()
                                if recharge_dt <= today and recharge_dt + timedelta(days=28) >= today:
                                    total_numbers_recharged_count += 1
                            except Exception:
                                pass
                platform_status_counts[platform] = status_map
                perm_block_counts[platform] = pb_count
                perm_block_total += pb_count
                platform_dept_counts[platform] = dept_map
                platform_number_type_counts[platform] = num_type_map
            except Exception as e:
                print(f"[tracker_stats] error for {platform}: {e}")
                platform_counts[platform] = 0
                platform_status_counts[platform] = {}
                perm_block_counts[platform] = 0
        try:
            total_response = social_supabase.table("social_media_accounts").select("id", count='exact').execute()
            total_accounts = total_response.count or 0
        except Exception:
            total_accounts = sum(platform_counts.values())
        return jsonify({
            "success": True,
            "stats": {
                "platform_counts": platform_counts,
                "platform_status_counts": platform_status_counts,
                "total_accounts": total_accounts,
                "perm_block_counts": perm_block_counts,
                "perm_block_total": perm_block_total,
                "platform_dept_counts": platform_dept_counts,
                "platform_number_type_counts": platform_number_type_counts,
                "total_numbers_recharged_count": total_numbers_recharged_count,
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/get-number-type-counts", methods=["GET"])
@login_required
def get_number_type_counts():
    try:
        is_super       = is_superadmin(session)
        allowed_depts = session.get("allowed_departments")
        number_types = ["Prepaid", "Postpaid", "Disposable Number"]
        counts = {}
        for nt in number_types:
            q = social_supabase.table("social_media_accounts") \
                .select("id", count='exact') \
                .eq("number_type", nt) \
                .neq("account_status", "Permanent Block")
            if not is_super and allowed_depts:
                if len(allowed_depts) == 1:
                    q = q.eq("department", allowed_depts[0])
                else:
                    q = q.in_("department", allowed_depts)
            resp = q.execute()
            counts[nt] = resp.count or 0
        return jsonify({"success": True, "counts": counts})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# SOCIAL IMPORT — with activity logging
# ============================================================
@app.route("/social-import", methods=["POST"])
@login_required
def social_import():
    try:
        file = request.files.get("file")
        if not file or file.filename == '':
            flash("No file selected", "error")
            return redirect("/?page=social")
        if not is_allowed_file(file.filename):
            flash(f"Unsupported file type.", "error")
            return redirect("/?page=social")
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
        df = read_data_file(temp_path, file_ext)
        df.columns = df.columns.astype(str).str.strip()
        df = df.fillna('')
        ALL_SOCIAL_COLUMNS = [
            'owned_by', 'login_user', 'number', 'login_device', 'sim_inserted_device',
            'account_status', 'review_status', 'number_type', 'blocked_date', 'unblock_date',
            'account_create_date', 'sim_operator', 'full_name', 'recharge_date', 'sim_buy_date',
            'account_type', 'mail_id', 'account_id', 'password', 'page_name', 'platform', 'department',
        ]
        file_columns = list(df.columns)
        matched_columns = [col for col in file_columns if col in ALL_SOCIAL_COLUMNS and col != 'id']
        if not matched_columns:
            flash("Import Error: No matching column names found.", "error")
            os.remove(temp_path)
            return redirect("/?page=social")
        try:
            max_id_response = social_supabase.table("social_media_accounts").select("id").order("id", desc=True).limit(1).execute()
            next_id = int(max_id_response.data[0]['id']) + 1 if max_id_response.data else 1
        except Exception:
            next_id = None
        DATE_COLUMNS = {'blocked_date', 'unblock_date', 'account_create_date', 'recharge_date', 'sim_buy_date'}
        def sanitize_value(col, value):
            if value is None:
                return None if col in DATE_COLUMNS else "NA"
            try:
                if pd.isna(value):
                    return None if col in DATE_COLUMNS else "NA"
            except (TypeError, ValueError):
                pass
            v = str(value).strip()
            if col in DATE_COLUMNS:
                if not v or v.upper() in ('NA', 'N/A', 'NAT', 'NONE', 'NULL', 'UNDEFINED', '-', 'N.A', 'N.A.', ''):
                    return None
                if ' ' in v: v = v.split(' ')[0]
                if 'T' in v: v = v.split('T')[0]
                return v
            if col == 'account_status':
                return normalize_social_account_status(v) if v else "NA"
            return v if v else "NA"
        records = []
        for i, (_, row) in enumerate(df.iterrows()):
            record = {}
            if next_id is not None:
                record['id'] = next_id + i
            for col in matched_columns:
                record[col] = sanitize_value(col, row[col])
            records.append(record)
        social_supabase.table("social_media_accounts").insert(records).execute()
        log_activity(
            action_type="import",
            target_table="social_media_accounts",
            extra_info={"file_name": filename, "records_count": len(records)}
        )
        flash(f"File Imported Successfully! {len(records)} records added.", "success")
        os.remove(temp_path)
    except Exception as e:
        flash(f"Import Error: {str(e)}", "error")
    return redirect("/?page=social")

@app.route("/social-export", methods=["GET"])
@login_required
def social_export():
    try:
        social_search = request.args.get("social_search", "").strip()
        social_platform = request.args.get("social_platform", "").strip()
        social_permanent_block = request.args.get("permanent_block", "").strip()
        social_status_filter = request.args.get("social_status", "").strip()
        social_department_filter = request.args.get("social_department", "").strip()

        def _build_social_query():
            q = social_supabase.table("social_media_accounts").select("*")
            allowed_depts = session.get("allowed_departments")
            if allowed_depts:
                if len(allowed_depts) == 1:
                    q = q.eq("department", allowed_depts[0])
                else:
                    q = q.in_("department", allowed_depts)
            if social_search:
                like_term = f"%{social_search}%"
                q = q.or_(
                    f"login_user.ilike.{like_term},number.ilike.{like_term},"
                    f"full_name.ilike.{like_term},page_name.ilike.{like_term},"
                    f"platform.ilike.{like_term},account_status.ilike.{like_term},"
                    f"owned_by.ilike.{like_term},department.ilike.{like_term},"
                    f"sim_operator.ilike.{like_term},mail_id.ilike.{like_term},"
                    f"account_id.ilike.{like_term},number_type.ilike.{like_term},"
                    f"login_device.ilike.{like_term},sim_inserted_device.ilike.{like_term}"
                )
            if social_platform and social_platform not in ["", "All Platforms"]:
                q = q.eq("platform", social_platform)
            if social_department_filter:
                q = q.eq("department", social_department_filter)
            if social_permanent_block == "true":
                q = q.eq("account_status", "Permanent Block")
            elif social_status_filter:
                q = apply_social_status_filter(q, social_status_filter)
            else:
                q = q.neq("account_status", "Permanent Block")
            return q

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return _stream_supabase_csv(_build_social_query, "id", f"social_media_accounts_{timestamp}.csv", row_transform=normalize_social_account_row)
    except Exception as e:
        flash(f"Export Error: {str(e)}", "error")
        return redirect("/?page=social")

@app.route("/get-sheet-headers/<sheet_type>", methods=["GET"])
@login_required
def get_sheet_headers_route(sheet_type):
    try:
        headers = get_sheet_headers(sheet_type)
        config = load_config()
        sheet_name = config['sheet_mappings'][sheet_type]['name'] if config else sheet_type
        return jsonify({"success": True, "sheet_name": sheet_name, "headers": headers, "headers_count": len(headers)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/download-template/<sheet_type>", methods=["GET"])
@login_required
def download_template(sheet_type):
    try:
        headers = get_sheet_headers(sheet_type)
        if not headers:
            flash("No headers found for this sheet type", "error")
            return redirect("/?page=sheet")
        output = io.StringIO()
        csv.writer(output).writerow(headers)
        output.seek(0)
        config = load_config()
        sheet_name = config['sheet_mappings'][sheet_type]['name'].replace(' ', '_') if config else sheet_type
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            download_name=f"{sheet_name}_Input_Template.csv",
            as_attachment=True, mimetype="text/csv"
        )
    except Exception as e:
        flash(f"Error generating template: {str(e)}", "error")
        return redirect("/?page=sheet")

@app.route("/preview-sheet", methods=["POST"])
@login_required
def preview_sheet():
    try:
        sheet_type = request.form.get("sheet_type")
        file = request.files.get("file")
        if not sheet_type:
            return jsonify({"success": False, "error": "Please select a sheet type"})
        if not file or file.filename == '':
            return jsonify({"success": False, "error": "Please select a file"})
        if not is_allowed_file(file.filename):
            return jsonify({"success": False, "error": "Unsupported file type."})
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        try:
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
            df = read_data_file(temp_path, file_ext)
            if df.empty:
                return jsonify({"success": False, "error": "The uploaded file is empty"})

            is_valid, missing_cols, extra_cols = validate_input_columns(df, sheet_type)
            if not is_valid:
                os.remove(temp_path)
                error_parts = []
                if missing_cols:
                    error_parts.append(f"Missing column(s): {', '.join(missing_cols)}")
                if extra_cols:
                    error_parts.append(f"Extra/unexpected column(s): {', '.join(extra_cols)}")
                return jsonify({
                    "success": False,
                    "error": "Uploaded file does not match the input template. " + " | ".join(error_parts) +
                             ". Please use the Download Template option and upload the file in the same format."
                })

            config = load_config()
            sheet_config = config['sheet_mappings'][sheet_type]
            result_df, preview_metrics = process_sheet_data(df, sheet_type)
            os.remove(temp_path)
            return jsonify({
                "success": True,
                "sheet_name": sheet_config['name'],
                "total_values": preview_metrics['total_values'],
                "unique_upi_ids": preview_metrics['unique_upi_ids'],
                "unique_bank_accounts": preview_metrics['unique_bank_accounts'],
                "unique_websites": preview_metrics['unique_websites'],
                "total_columns": len(result_df.columns),
                "columns": list(result_df.columns),
                "preview_rows": result_df.fillna('').head(50).to_dict(orient='records'),
                "input_headers": list(df.columns),
                "output_headers": list(result_df.columns)
            })
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"success": False, "error": f"Error processing file: {str(e)}"})
    except Exception as e:
        return jsonify({"success": False, "error": f"Error previewing sheet: {str(e)}"})

@app.route("/generate-sheet", methods=["POST"])
@login_required
def generate_sheet():
    try:
        sheet_type = request.form.get("sheet_type")
        file = request.files.get("file")
        if not sheet_type:
            flash("Please select a sheet type", "error")
            return redirect("/?page=sheet")
        if not file or file.filename == '':
            flash("Please select a file", "error")
            return redirect("/?page=sheet")
        if not is_allowed_file(file.filename):
            flash("Unsupported file type.", "error")
            return redirect("/?page=sheet")
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        try:
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
            df = read_data_file(temp_path, file_ext)
            if df.empty:
                flash("The uploaded file is empty", "error")
                return redirect("/?page=sheet")

            is_valid, missing_cols, extra_cols = validate_input_columns(df, sheet_type)
            if not is_valid:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                error_parts = []
                if missing_cols:
                    error_parts.append(f"Missing column(s): {', '.join(missing_cols)}")
                if extra_cols:
                    error_parts.append(f"Extra/unexpected column(s): {', '.join(extra_cols)}")
                flash("Uploaded file does not match the input template. " + " | ".join(error_parts) +
                      ". Please use the Download Template option and upload the file in the same format.", "error")
                return redirect("/?page=sheet")

            result_df, _ = process_sheet_data(df, sheet_type)
            output = io.StringIO()
            result_df.to_csv(output, index=False, encoding='utf-8-sig')
            today_date = datetime.now().strftime("%Y-%m-%d")
            sheet_name_clean = SHEET_TYPES.get(sheet_type, sheet_type)
            os.remove(temp_path)
            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                download_name=f"{sheet_name_clean}_{today_date}.csv",
                as_attachment=True, mimetype="text/csv"
            )
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            flash(f"Error processing file: {str(e)}", "error")
            return redirect("/?page=sheet")
    except Exception as e:
        flash(f"Error generating sheet: {str(e)}", "error")
        return redirect("/?page=sheet")

def _build_sheet_csv_response(sheet_type, file_storage):
    if not sheet_type:
        raise ValueError("Please select a sheet type")
    if not file_storage or file_storage.filename == '':
        raise ValueError("Please select a file")
    if not is_allowed_file(file_storage.filename):
        raise ValueError("Unsupported file type.")

    filename = secure_filename(file_storage.filename)
    temp_path = os.path.join(tempfile.gettempdir(), filename)
    file_storage.save(temp_path)
    try:
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
        df = read_data_file(temp_path, file_ext)
        if df.empty:
            raise ValueError("The uploaded file is empty")

        is_valid, missing_cols, extra_cols = validate_input_columns(df, sheet_type)
        if not is_valid:
            error_parts = []
            if missing_cols:
                error_parts.append(f"Missing column(s): {', '.join(missing_cols)}")
            if extra_cols:
                error_parts.append(f"Extra/unexpected column(s): {', '.join(extra_cols)}")
            raise ValueError("Uploaded file does not match the input template. " + " | ".join(error_parts))

        result_df, _ = process_sheet_data(df, sheet_type)
        output = io.StringIO()
        result_df.to_csv(output, index=False, encoding='utf-8-sig')
        return output.getvalue(), result_df, filename
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

def _build_aml_gui_import_csv(sheet_type, result_df):
    return None, result_df

def _sheet_import_download_payload(csv_text, filename):
    return {
        "filename": filename,
        "content_b64": base64.b64encode(csv_text.encode("utf-8-sig")).decode("ascii"),
    }

@app.route("/sheet-import-to-aml-gui", methods=["POST"])
@login_required
def sheet_import_to_aml_gui():
    download_payload = None
    try:
        sheet_type = request.form.get("sheet_type")
        file = request.files.get("file")
        print(f"[SHEET AML IMPORT] start sheet_type={sheet_type} file={getattr(file, 'filename', None)}", flush=True)
        csv_text, result_df, _ = _build_sheet_csv_response(sheet_type, file)
        import_csv_text, import_df = _build_aml_gui_import_csv(sheet_type, result_df)
        if import_csv_text:
            csv_text = import_csv_text
        import_filename = f"aml_gui_import_{uuid.uuid4().hex}.csv"
        download_payload = _sheet_import_download_payload(csv_text, import_filename)
        print(f"[SHEET AML IMPORT] CSV generated filename={import_filename} rows={len(import_df)} cols={len(import_df.columns)}", flush=True)
        import_engine = os.environ.get("SHEET_AML_IMPORT_ENGINE", "http").strip().lower()
        if import_engine == "selenium":
            page_preview = _import_sheet_csv_to_aml_gui_headed(csv_text, import_filename)
        else:
            resp = _import_sheet_csv_to_aml_gui(csv_text, import_filename)
            page_preview = _validate_aml_sheet_import_response(resp.text, resp.url)
        print(f"[SHEET AML IMPORT] {import_engine if import_engine == 'selenium' else 'http'} import submitted", flush=True)

        return jsonify({
            "success": True,
            "message": "CSV generated and submitted to AML GUI.",
            "rows": len(import_df),
            "debug_preview": page_preview,
            "download": download_payload,
        })
    except Exception as e:
        print(f"[SHEET AML IMPORT] ERROR: {e}", flush=True)
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "download": download_payload})

# ============================================================
# BS Investment Scam Import
# ============================================================
BS_INVESTMENT_IMPORT_COLUMNS = [
    "Id", "Bank_account_number", "Bank_name", "Upi_vpa", "Screenshot",
    "Search_for", "Upi_bank_account_wallet", "Handle", "Payment_gateway_name",
    "Scam_type", "Ifsc_code", "Upi_url", "Website_url", "Inserted_date",
    "Input_user", "Web_contact_no", "Customer", "Package_name", "Channel_name",
    "Ac_holder_name", "Platform", "Status", "Priority", "Flag", "Cessation",
    "Reviewed_status", "Origin", "Category_of_website", "Screenshot_case_report_link",
    "Payment_gateway_intermediate_url", "Neft_imps", "Transaction_method",
    "Bank_branch_details", "Payment_gateway_url", "Reported_earlier",
    "Approvd_status", "Approved_by", "Qc_remarks", "Inserted_datetime",
    "Case_generated_time", "Upi_found_status", "Feature_type", "Approved_date",
    "Video_url"
]

BS_INVESTMENT_IMPORT_ALIASES = {
    "Bank_account_number": ["bank_account_number", "bank account number", "account_number", "account number", "acc_no", "acc no"],
    "Bank_name": ["bank_name", "bank name"],
    "Upi_vpa": ["upi_vpa", "upi vpa", "upi", "vpa"],
    "Screenshot": ["screenshot", "image", "proof"],
    "Search_for": ["search_for", "search for"],
    "Upi_bank_account_wallet": ["upi_bank_account_wallet", "upi bank account wallet", "wallet", "payment type"],
    "Handle": ["handle", "upi handle"],
    "Payment_gateway_name": ["payment_gateway_name", "payment gateway name", "gateway name"],
    "Scam_type": ["scam_type", "scam type", "type", "category"],
    "Ifsc_code": ["ifsc_code", "ifsc code", "ifsc", "bank_code"],
    "Upi_url": ["upi_url", "upi url"],
    "Website_url": ["website_url", "website url", "url", "website"],
    "Inserted_date": ["inserted_date", "inserted date", "date"],
    "Input_user": ["input_user", "input user", "user", "username"],
    "Web_contact_no": ["web_contact_no", "web contact no", "contact number", "phone", "mobile"],
    "Customer": ["customer"],
    "Package_name": ["package_name", "package name"],
    "Channel_name": ["channel_name", "channel name"],
    "Ac_holder_name": ["ac_holder_name", "account holder name", "account_holder", "holder_name", "customer name"],
    "Platform": ["platform"],
    "Status": ["status"],
    "Priority": ["priority"],
    "Flag": ["flag"],
    "Cessation": ["cessation"],
    "Reviewed_status": ["reviewed_status", "reviewed status"],
    "Origin": ["origin"],
    "Category_of_website": ["category_of_website", "category of website", "category"],
    "Screenshot_case_report_link": ["screenshot_case_report_link", "screenshot case report link", "case report link"],
    "Payment_gateway_intermediate_url": ["payment_gateway_intermediate_url", "payment gateway intermediate url"],
    "Neft_imps": ["neft_imps", "neft imps"],
    "Transaction_method": ["transaction_method", "transaction method", "payment method", "method"],
    "Bank_branch_details": ["bank_branch_details", "bank branch details"],
    "Payment_gateway_url": ["payment_gateway_url", "payment gateway url", "payment_url", "gateway"],
    "Reported_earlier": ["reported_earlier", "reported earlier"],
    "Approvd_status": ["approvd_status", "approved_status", "approved status"],
    "Approved_by": ["approved_by", "approved by"],
    "Qc_remarks": ["qc_remarks", "qc remarks"],
    "Inserted_datetime": ["inserted_datetime", "inserted datetime"],
    "Case_generated_time": ["case_generated_time", "case generated time"],
    "Upi_found_status": ["upi_found_status", "upi found status"],
    "Feature_type": ["feature_type", "feature type"],
    "Approved_date": ["approved_date", "approved date"],
    "Video_url": ["video_url", "video url"],
}

def normalize_import_header(value):
    return re.sub(r'[^a-z0-9]+', ' ', str(value).lower()).strip()

def find_import_column(df_columns, target_col):
    normalized = {normalize_import_header(col): col for col in df_columns}
    candidates = [target_col] + BS_INVESTMENT_IMPORT_ALIASES.get(target_col, [])
    for candidate in candidates:
        key = normalize_import_header(candidate)
        if key in normalized:
            return normalized[key]
    return None

def normalize_import_date(value, default_value=None):
    cleaned = clean_value(value)
    if cleaned == "NA":
        return default_value if default_value is not None else "NA"
    try:
        parsed = pd.to_datetime(cleaned, errors='coerce')
        if not pd.isna(parsed):
            return parsed.strftime("%Y-%m-%d")
    except Exception:
        pass
    return cleaned.split(" ")[0] if " " in cleaned else cleaned

def normalize_import_datetime(value, default_value=None):
    cleaned = clean_value(value)
    if cleaned == "NA":
        return default_value if default_value is not None else "NA"
    try:
        parsed = pd.to_datetime(cleaned, errors='coerce')
        if not pd.isna(parsed):
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    return cleaned

@app.route("/investment-import", methods=["POST"])
@login_required
def investment_import():
    temp_path = None
    try:
        file = request.files.get("file")
        if not file or file.filename == '':
            flash("No file selected", "error")
            return redirect("/?page=investment")
        if not is_allowed_file(file.filename):
            flash("Unsupported file type.", "error")
            return redirect("/?page=investment")

        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
        df = read_data_file(temp_path, file_ext)
        if df.empty:
            flash("The uploaded file is empty", "error")
            return redirect("/?page=investment")

        df.columns = df.columns.astype(str).str.strip()
        df = df.fillna('')
        source_columns = list(df.columns)
        column_lookup = {
            target: find_import_column(source_columns, target)
            for target in BS_INVESTMENT_IMPORT_COLUMNS
            if target != "Id"
        }
        matched_columns = [col for col in column_lookup.values() if col]
        if not matched_columns:
            flash("Import Error: No matching investment column names found.", "error")
            return redirect("/?page=investment")

        try:
            max_id_response = supabase.table("BS_Investment_Scam").select("Id").order("Id", desc=True).limit(1).execute()
            next_id = int(max_id_response.data[0]["Id"]) + 1 if max_id_response.data else 1
        except Exception:
            next_id = 1

        today = datetime.now().strftime("%Y-%m-%d")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        input_user_default = get_clean_display_name(session.get("display_name", "User"))
        records = []

        for i, (_, row) in enumerate(df.iterrows()):
            record = {col: "NA" for col in BS_INVESTMENT_IMPORT_COLUMNS}
            record["Id"] = next_id + i
            for target_col, source_col in column_lookup.items():
                if source_col:
                    record[target_col] = clean_value(row[source_col])

            record["Inserted_date"] = normalize_import_date(record.get("Inserted_date"), today)
            record["Inserted_datetime"] = normalize_import_datetime(record.get("Inserted_datetime"), now_str)
            if record["Input_user"] == "NA":
                record["Input_user"] = input_user_default

            if record["Screenshot"] != "NA":
                case_time, screenshot_date = extract_case_time_and_date_from_npci_url(record["Screenshot"])
                record["Screenshot"] = generate_screenshot_urls(record["Screenshot"])
                if record["Case_generated_time"] == "NA":
                    record["Case_generated_time"] = case_time
                if record["Inserted_date"] == "NA" and screenshot_date != "NA":
                    record["Inserted_date"] = screenshot_date
                if record["Screenshot_case_report_link"] == "NA":
                    record["Screenshot_case_report_link"] = record["Screenshot"]

            if record["Handle"] == "NA":
                record["Handle"] = extract_handle(record["Upi_vpa"])
            if record["Bank_name"] == "NA":
                record["Bank_name"] = get_bank_name_from_handle(record["Handle"], record["Ifsc_code"])
            if record["Search_for"] == "NA":
                record["Search_for"] = extract_search_for_from_url(record["Website_url"])
            if record["Upi_bank_account_wallet"] == "NA":
                record["Upi_bank_account_wallet"] = "UPI" if record["Upi_vpa"] != "NA" else ("Bank Account" if record["Bank_account_number"] != "NA" else "NA")
            if record["Payment_gateway_url"] == "NA" and record["Upi_url"] != "NA":
                record["Payment_gateway_url"] = record["Upi_url"]
            if record["Payment_gateway_intermediate_url"] == "NA" and record["Payment_gateway_url"] != "NA":
                record["Payment_gateway_intermediate_url"] = record["Payment_gateway_url"]
            if record["Upi_url"] == "NA" and record["Payment_gateway_url"] != "NA":
                record["Upi_url"] = record["Payment_gateway_url"]
            if record["Payment_gateway_name"] == "NA":
                record["Payment_gateway_name"] = extract_payment_gateway_name(record["Upi_url"], record["Website_url"])

            defaults = {
                "Customer": "Mystery Shopping",
                "Package_name": "com.mysteryshopping",
                "Channel_name": "Organic Search",
                "Platform": "NA",
                "Status": "Active",
                "Priority": "High",
                "Flag": "1",
                "Cessation": "Open",
                "Reviewed_status": "1",
                "Neft_imps": "NA",
                "Reported_earlier": "No",
                "Approvd_status": "1",
                "Feature_type": "BS Investment Scam",
            }
            for col, value in defaults.items():
                if record[col] == "NA":
                    record[col] = value

            if record["Category_of_website"] == "NA" and record["Scam_type"] != "NA":
                record["Category_of_website"] = record["Scam_type"]
            records.append(record)

        supabase.table("BS_Investment_Scam").insert(records).execute()
        log_activity(
            action_type="import",
            target_table="BS_Investment_Scam",
            extra_info={"file_name": filename, "records_count": len(records)}
        )
        flash(f"Investment file imported successfully! {len(records)} records added.", "success")
    except Exception as e:
        flash(f"Investment Import Error: {str(e)}", "error")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
    return redirect("/?page=investment")

@app.route("/get-excel-headers", methods=["GET"])
@login_required
def get_excel_headers():
    try:
        bank_headers = list(pd.read_excel(BANK_NAME_MAPPING_PATH).columns) if BANK_NAME_MAPPING_PATH.exists() else []
        return jsonify({
            "success": True,
            "bank_name_mapping_headers": bank_headers,
            "bank_name_mapping_count": len(BANK_NAME_MAPPING)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/get-ifsc-headers", methods=["GET"])
@login_required
def get_ifsc_headers():
    try:
        ifsc_headers = list(pd.read_excel(IFSC_MAPPING_PATH).columns) if IFSC_MAPPING_PATH.exists() else []
        return jsonify({"success": True, "ifsc_mapping_headers": ifsc_headers, "ifsc_mapping_count": len(IFSC_MAPPING)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/reload-data", methods=["POST"])
@login_required
def reload_data():
    try:
        load_excel_data()
        return jsonify({"success": True, "message": f"Data reloaded! Bank: {len(BANK_NAME_MAPPING)}, IFSC: {len(IFSC_MAPPING)}"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# SCRAPING DATA IMPORT — with activity logging
# ============================================================
@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "file" not in request.files:
        flash("No file uploaded", "error")
        return redirect("/?page=scraping")
    file = request.files["file"]
    if not file or file.filename == '':
        flash("No file selected", "error")
        return redirect("/?page=scraping")
    if not is_allowed_file(file.filename):
        flash("Unsupported file type.", "error")
        return redirect("/?page=scraping")
    try:
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
        df = read_data_file(temp_path, file_ext)
        df.columns = df.columns.astype(str).str.strip()
        df = df.fillna('')
        required_cols = [
            "name", "platform", "post_url", "chat_number", "group_name",
            "chat_link", "inserted_date", "chat_status", "assigned_to",
            "assigned_at_datetime", "inserted_datetime", "priority",
            "extra_field_1", "extra_field_2", "extra_field_3",
            "extra_field_4", "extra_field_5", "screenshot", "scam_type"
        ]
        for col in required_cols:
            if col not in df.columns:
                df[col] = "NA"
        if 'inserted_date' not in df.columns or df['inserted_date'].isna().all():
            df['inserted_date'] = datetime.now().strftime("%Y-%m-%d")
        if 'inserted_datetime' not in df.columns or df['inserted_datetime'].isna().all():
            df['inserted_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        records = df[required_cols].to_dict(orient='records')
        supabase.table("scrapping_data").insert(records).execute()
        log_activity(
            action_type="import",
            target_table="scrapping_data",
            extra_info={"file_name": filename, "records_count": len(records)}
        )
        flash(f"File Imported Successfully! {len(records)} records added.", "success")
        os.remove(temp_path)
    except Exception as e:
        flash(f"Import Error: {str(e)}", "error")
    return redirect("/?page=scraping")

@app.route("/export")
@login_required
def export():
    try:
        search_query = request.args.get("search", "").strip()
        scam_filter = request.args.get("scam_type", "").strip()
        platform_filter = request.args.get("platform", "").strip()
        date_filter = request.args.get("date_filter", "").strip()
        date_from = request.args.get("date_from", "").strip()
        date_to = request.args.get("date_to", "").strip()
        share_status_filter = request.args.get("share_status", "").strip()

        def _build_query():
            q = supabase.table("scrapping_data").select("*")
            if not is_superadmin(session):
                current_clean_name = get_clean_display_name(session.get("display_name", ""))
                q = q.eq("name", current_clean_name)
            if search_query:
                like_term = f"%{search_query}%"
                q = q.or_(f"name.ilike.{like_term},platform.ilike.{like_term},post_url.ilike.{like_term},chat_number.ilike.{like_term},group_name.ilike.{like_term},chat_link.ilike.{like_term},scam_type.ilike.{like_term}")
            if scam_filter: q = q.eq("scam_type", scam_filter)
            if platform_filter: q = q.eq("platform", platform_filter)
            if date_from: q = q.gte("inserted_date", date_from)
            if date_to: q = q.lte("inserted_date", date_to)
            if date_filter and not date_from and not date_to: q = q.eq("inserted_date", date_filter)
            if share_status_filter: q = q.eq("share_status", share_status_filter)
            return q

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return _stream_supabase_csv(_build_query, "id", f"scam_reports_{timestamp}.csv")
    except Exception as e:
        flash(f"Export Error: {str(e)}", "error")
        return redirect("/?page=scraping")

@app.route("/parse-raw-file", methods=["POST"])
@login_required
def parse_raw_file():
    try:
        file = request.files.get("file")
        if not file or file.filename == '':
            return jsonify({"success": False, "error": "No file"})
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'csv'
        df = read_data_file(temp_path, file_ext)
        df = df.fillna('')
        os.remove(temp_path)
        headers = list(df.columns)
        rows = df.head(5000).to_dict(orient='records')
        return jsonify({
            "success": True,
            "headers": headers,
            "rows": rows,
            "total_rows": len(df)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

AML_GUI_BASE_URL = "https://aml-gui.chargebackzero.com/"
AML_DATALIST_URL = urllib.parse.urljoin(AML_GUI_BASE_URL, "datalist_npci.php")
AML_IMPORT_URL = os.environ.get("AML_IMPORT_URL", "").strip()
AML_GUI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36"
)

def _parse_csv_response_to_summary(resp):
    content_type = resp.headers.get("Content-Type", "")
    try:
        df = pd.read_csv(io.BytesIO(resp.content), dtype=str).fillna("")
    except Exception as exc:
        preview = (resp.text or "")[:500].replace("\n", " ").replace("\r", " ")
        raise RuntimeError(f"Export did not return readable CSV. content_type={content_type} preview={preview}") from exc
    headers = [str(c) for c in df.columns]
    rows = df.head(5000).to_dict(orient="records")
    return headers, rows, len(df)

def _download_aml_datalist_csv(date_from, date_to, keyword="icuser", transaction_method="All", status="Approve"):
    session_obj = aml_login_requests(max_captcha_attempts=7, require_report_form=False)
    headers = {"Referer": AML_DATALIST_URL, "Origin": AML_GUI_BASE_URL.rstrip("/")}
    session_obj.get(AML_DATALIST_URL, timeout=60)
    print(f"[AML SUMMARY] Applying HTTP filter {date_from}..{date_to} keyword={keyword}", flush=True)
    session_obj.post(
        AML_DATALIST_URL,
        data={
            "clear_data": "",
            "Keyword_Filter": keyword,
            "from_date": date_from,
            "to_date": date_to,
            "transaction_method": transaction_method,
            "status": status,
            "filter": "filter",
        },
        headers=headers,
        timeout=300,
    )
    print("[AML SUMMARY] Exporting filtered CSV over HTTP", flush=True)
    resp = session_obj.post(
        AML_DATALIST_URL,
        data={"exportConfigFields[]": "All", "Export": "Export"},
        headers=headers,
        timeout=900,
        stream=False,
    )
    resp.raise_for_status()
    return _parse_csv_response_to_summary(resp)

def _extract_form_action(html_text, base_url):
    forms = re.findall(r"<form\b[^>]*>.*?</form>", html_text or "", flags=re.IGNORECASE | re.DOTALL)
    for form in forms:
        if re.search(r"type=[\"']file[\"']", form, flags=re.IGNORECASE):
            match = re.search(r"action=[\"']([^\"']*)[\"']", form, flags=re.IGNORECASE)
            return urllib.parse.urljoin(base_url, match.group(1)) if match else base_url
    return None

def _extract_file_form_html(html_text):
    forms = re.findall(r"<form\b[^>]*>.*?</form>", html_text or "", flags=re.IGNORECASE | re.DOTALL)
    return next((form for form in forms if re.search(r"type=[\"']file[\"']", form, flags=re.IGNORECASE)), html_text or "")

def _extract_attrs(tag_text):
    attrs = {}
    for match in re.finditer(
        r"([A-Za-z_:][-A-Za-z0-9_:.]*)\s*=\s*(?:([\"'])(.*?)\2|([^\s\"'=<>`]+))",
        tag_text or "",
        flags=re.DOTALL,
    ):
        attrs[match.group(1).lower()] = html.unescape((match.group(3) if match.group(2) else match.group(4)) or "")
    return attrs

def _set_form_value(data, name, value):
    if not name:
        return
    if name in data:
        if isinstance(data[name], list):
            data[name].append(value)
        else:
            data[name] = [data[name], value]
    else:
        data[name] = value

def _extract_import_form_payload(html_text, service_name=None):
    form_html = _extract_file_form_html(html_text)
    data = {}
    file_field = "file"
    submit_added = False

    for match in re.finditer(r"<input\b([^>]*)>", form_html, flags=re.IGNORECASE | re.DOTALL):
        attrs = _extract_attrs(match.group(1))
        input_type = (attrs.get("type") or "text").lower()
        name = attrs.get("name")
        if not name:
            continue
        if input_type == "file":
            file_field = name
            continue
        if input_type in ("reset", "image"):
            continue
        if input_type in ("submit", "button"):
            if not submit_added and name:
                _set_form_value(data, name, attrs.get("value", ""))
                submit_added = True
            continue
        if input_type in ("checkbox", "radio"):
            _set_form_value(data, name, attrs.get("value") or "on")
        else:
            _set_form_value(data, name, attrs.get("value", ""))

    for select_match in re.finditer(r"<select\b([^>]*)>(.*?)</select>", form_html, flags=re.IGNORECASE | re.DOTALL):
        attrs = _extract_attrs(select_match.group(1))
        name = attrs.get("name")
        if not name:
            continue
        option_values = []
        for option_match in re.finditer(r"<option\b([^>]*)>(.*?)</option>", select_match.group(2) or "", flags=re.IGNORECASE | re.DOTALL):
            opt_attrs = _extract_attrs(option_match.group(1))
            value = opt_attrs.get("value")
            if value is None:
                value = re.sub(r"<[^>]+>", "", option_match.group(2) or "").strip()
            value = html.unescape(value or "").strip()
            if value:
                option_values.append(value)
        if option_values:
            data[name] = option_values

    for textarea_match in re.finditer(r"<textarea\b([^>]*)>(.*?)</textarea>", form_html, flags=re.IGNORECASE | re.DOTALL):
        attrs = _extract_attrs(textarea_match.group(1))
        name = attrs.get("name")
        if name:
            _set_form_value(data, name, html.unescape(textarea_match.group(2) or ""))

    service_name = (service_name or os.environ.get("AML_IMPORT_SERVICE_NAME") or "").strip()
    if service_name and not str(data.get("serviceNameHidden", "")).strip():
        data["serviceNameHidden"] = service_name
    if "importCSVFileButton" not in data:
        data["importCSVFileButton"] = "IMPORT"

    return data, file_field

def _summarize_import_form_controls(html_text):
    form_html = _extract_file_form_html(html_text)
    controls = []
    for match in re.finditer(r"<input\b([^>]*)>", form_html, flags=re.IGNORECASE | re.DOTALL):
        attrs = _extract_attrs(match.group(1))
        input_type = (attrs.get("type") or "text").lower()
        controls.append({
            "tag": "input",
            "type": input_type,
            "name": attrs.get("name", ""),
            "value": attrs.get("value", "")[:80],
        })
    for select_match in re.finditer(r"<select\b([^>]*)>(.*?)</select>", form_html, flags=re.IGNORECASE | re.DOTALL):
        attrs = _extract_attrs(select_match.group(1))
        options = []
        for option_match in re.finditer(r"<option\b([^>]*)>(.*?)</option>", select_match.group(2) or "", flags=re.IGNORECASE | re.DOTALL):
            opt_attrs = _extract_attrs(option_match.group(1))
            value = opt_attrs.get("value")
            if value is None:
                value = re.sub(r"<[^>]+>", "", option_match.group(2) or "").strip()
            text = re.sub(r"<[^>]+>", "", option_match.group(2) or "").strip()
            options.append({"value": html.unescape(value or "")[:80], "text": html.unescape(text or "")[:80]})
        controls.append({
            "tag": "select",
            "name": attrs.get("name", ""),
            "multiple": "multiple" in (select_match.group(1) or "").lower(),
            "options": options[:20],
        })
    return controls[:80]

def _validate_aml_sheet_import_response(response_text, response_url):
    response_text = response_text or ""
    response_preview = re.sub(r"\s+", " ", response_text)[:500]
    if re.search(r"name=[\"']usernamee[\"']", response_text, flags=re.IGNORECASE):
        raise RuntimeError("AML import returned login page; session expired or login failed")

    success_match = re.search(
        r"(success(?:fully)?|imported\s+successfully|uploaded\s+successfully|inserted\s+successfully|record[s]?\s+(?:added|inserted)|data\s+import(?:ed)?\s+successfully)",
        response_text,
        flags=re.IGNORECASE,
    )
    # Success confirmation le liye pehle check karo — agar mil gaya toh import
    # pakka successful hai, chahe page mein kahin "required"/"error" jaisa koi
    # unrelated word (menu, footer, disclaimer) bhi ho.
    if success_match:
        return response_preview

    error_match = re.search(
        r"(oops|no\s+column\s+present|fatal\s+error|import\s+failed|upload\s+failed|invalid\s+file|not\s+imported|not\s+uploaded|duplicate\s+entr(?:y|ies)|wrong\s+format|column\s+mismatch)",
        response_text,
        flags=re.IGNORECASE,
    )
    if error_match:
        raise RuntimeError(f"AML import rejected the CSV near: {response_preview}")

    raise RuntimeError(
        "AML import did not return a success confirmation. "
        f"response_url={response_url} preview={response_preview}"
    )

def _find_aml_import_page(session_obj):
    if AML_IMPORT_URL:
        return AML_IMPORT_URL
    page = session_obj.get(AML_DATALIST_URL, timeout=60)
    page.raise_for_status()
    links = re.findall(r"<a\b[^>]*href=[\"']([^\"']+)[\"'][^>]*>", page.text or "", flags=re.IGNORECASE)
    for href in links:
        low = href.lower()
        if any(token in low for token in ("import", "upload", "csv")):
            return urllib.parse.urljoin(AML_DATALIST_URL, href)
    if links:
        return urllib.parse.urljoin(AML_DATALIST_URL, links[0])
    raise RuntimeError("AML import page link not found on datalist page")

def _import_sheet_csv_to_aml_gui(csv_text, filename):
    session_obj = aml_login_requests(max_captcha_attempts=7, require_report_form=False)
    import_page_url = _find_aml_import_page(session_obj)
    print(f"[SHEET AML IMPORT] import_page={import_page_url}", flush=True)
    import_page = session_obj.get(import_page_url, timeout=60)
    import_page.raise_for_status()
    post_url = _extract_form_action(import_page.text, import_page_url) or import_page_url
    _, _, _, platform_value = get_aml_credentials()
    data, file_field = _extract_import_form_payload(import_page.text, service_name=platform_value)
    selected_fields = sum(len(v) if isinstance(v, list) else 1 for v in data.values())
    print(f"[SHEET AML IMPORT] form file_field={file_field} selected_fields={selected_fields} serviceNameHidden={data.get('serviceNameHidden', '')}", flush=True)
    print(f"[SHEET AML IMPORT] csv_headers={csv_text.splitlines()[0][:1000] if csv_text else ''}", flush=True)
    print(f"[SHEET AML IMPORT] form_controls={json.dumps(_summarize_import_form_controls(import_page.text))[:3000]}", flush=True)
    files = {file_field: (filename or "aml_gui_import.csv", io.BytesIO(csv_text.encode("utf-8-sig")), "text/csv")}
    resp = session_obj.post(
        post_url,
        data=data,
        files=files,
        timeout=300,
        allow_redirects=True,
        headers={"Referer": import_page_url, "Origin": AML_GUI_BASE_URL.rstrip("/")},
    )
    resp.raise_for_status()
    response_text = resp.text or ""
    response_preview = re.sub(r"\s+", " ", response_text)[:500]
    print(f"[SHEET AML IMPORT] response_url={resp.url} content_type={resp.headers.get('Content-Type', '')} preview={response_preview}", flush=True)
    _validate_aml_sheet_import_response(response_text, resp.url)
    return resp

def _import_sheet_csv_to_aml_gui_headed(csv_text, filename):
    temp_csv_path = os.path.join(tempfile.gettempdir(), filename or f"aml_gui_import_{uuid.uuid4().hex}.csv")
    with open(temp_csv_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(csv_text)

    driver = aml_build_driver(headless=True)
    _AML_DEBUG_DRIVERS.append(driver)
    wait = WebDriverWait(driver, 40)
    try:
        if not aml_login(driver, max_captcha_attempts=7):
            raise RuntimeError("AML GUI headed login failed after captcha retries")

        driver.get(AML_DATALIST_URL)
        time.sleep(2)
        try:
            import_btn = wait.until(EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(normalize-space(.), 'Import') or contains(normalize-space(.), 'Update')]"
                "|//a[contains(normalize-space(.), 'Import') or contains(normalize-space(.), 'Update')]"
            )))
            driver.execute_script("arguments[0].click();", import_btn)
        except Exception as exc:
            print(f"[SHEET AML IMPORT HEADED] Import/Update click failed, falling back to modal JS: {exc}", flush=True)
            driver.execute_script("""
                const modal = document.querySelector('#importModal');
                if (modal) {
                    modal.classList.add('show');
                    modal.style.display = 'block';
                    modal.removeAttribute('aria-hidden');
                }
            """)
        time.sleep(1)

        file_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
        file_input.send_keys(temp_csv_path)

        form = file_input.find_element(By.XPATH, "ancestor::form[1]")
        form_debug = driver.execute_script("""
            const form = arguments[0];
            return Array.from(form.elements).map(el => ({
                tag: el.tagName,
                type: el.type || '',
                name: el.name || '',
                value: el.type === 'file' ? el.value.split('\\\\').pop() : el.value,
                multiple: !!el.multiple,
                options: el.tagName === 'SELECT'
                    ? Array.from(el.options || []).map(o => ({value:o.value, text:o.text, selected:o.selected}))
                    : []
            }));
        """, form)
        print(f"[SHEET AML IMPORT HEADED] form_controls={json.dumps(form_debug)[:3000]}", flush=True)
        selected_count = driver.execute_script("""
            const form = arguments[0];
            return 0;
        """, form)
        print(f"[SHEET AML IMPORT HEADED] selected import fields={selected_count}", flush=True)
        try:
            submit_btn = form.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
            driver.execute_script("arguments[0].click();", submit_btn)
        except Exception:
            driver.execute_script("arguments[0].submit();", form)

        time.sleep(5)
        page_text = driver.find_element(By.TAG_NAME, "body").text
        page_preview = re.sub(r"\s+", " ", page_text or "")[:500]
        print(f"[SHEET AML IMPORT HEADED] current_url={driver.current_url} preview={page_preview}", flush=True)
        if re.search(r"(Oops!|No column present|alert-danger|Fatal error|SQLSTATE|specified twice|not imported|failed|invalid)", page_text or "", flags=re.IGNORECASE):
            raise RuntimeError(f"AML GUI import failed on page: {page_preview}")

        _close_debug_driver(driver, temp_csv_path)
        return page_preview
    except Exception:
        try:
            debug_id = uuid.uuid4().hex[:8]
            debug_dir = os.path.join(tempfile.gettempdir(), "aml_sheet_import_debug")
            os.makedirs(debug_dir, exist_ok=True)
            driver.save_screenshot(os.path.join(debug_dir, f"sheet_import_{debug_id}.png"))
            with open(os.path.join(debug_dir, f"sheet_import_{debug_id}.html"), "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(f"[SHEET AML IMPORT HEADED] debug saved id={debug_id} dir={debug_dir}", flush=True)
        except Exception as dbg_exc:
            print(f"[SHEET AML IMPORT HEADED] debug save failed: {dbg_exc}", flush=True)
        _close_debug_driver(driver, temp_csv_path)
        raise

def _close_debug_driver(driver, temp_csv_path=None):
    try:
        driver.quit()
    except Exception:
        pass
    try:
        if driver in _AML_DEBUG_DRIVERS:
            _AML_DEBUG_DRIVERS.remove(driver)
    except Exception:
        pass
    if temp_csv_path and os.path.exists(temp_csv_path):
        try:
            os.remove(temp_csv_path)
        except Exception:
            pass

@app.route("/summary-source-from-gui", methods=["POST"])
@login_required
def summary_source_from_gui():
    if "sheet" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Didnt have Sheet Page Access"}), 403
    data = request.get_json() or {}
    date_from = str(data.get("date_from", "")).strip()
    date_to = str(data.get("date_to", "")).strip()
    keyword = "icuser"
    if not date_from or not date_to:
        return jsonify({"success": False, "error": "From date and To date required"})

    try:
        print(f"[AML SUMMARY] Start date_from={date_from} date_to={date_to} keyword={keyword}", flush=True)
        headers, rows, total_rows = _download_aml_datalist_csv(date_from, date_to, keyword=keyword)
        print(f"[AML SUMMARY] Completed HTTP export rows={total_rows} headers={len(headers)}", flush=True)
        return jsonify({
            "success": True,
            "headers": headers,
            "rows": rows,
            "total_rows": total_rows,
            "source": "AML GUI",
            "keyword": keyword,
            "date_from": date_from,
            "date_to": date_to,
        })
    except Exception as e:
        print(f"[AML SUMMARY] ERROR: {e}", flush=True)
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})

@app.route("/health", methods=["GET"])
def health_check():
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "excel_data_loaded": {
            "bank_name_mapping": len(BANK_NAME_MAPPING),
            "ifsc_mapping": len(IFSC_MAPPING)
        }
    })
# ============================================================
# SAVE SOCIAL FIELD — with old value fetch + activity logging
# ============================================================
@app.route("/save-social-field", methods=["POST"])
@login_required
def save_social_field():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        account_id = data.get('id')
        field = data.get('field')
        value = data.get('value', '').strip()
        if not account_id or not field:
            return jsonify({"success": False, "error": "Missing id or field"})
        EDITABLE_FIELDS = ['login_user', 'number', 'login_device', 'sim_inserted_device',
                           'account_status', 'review_status', 'number_type', 'blocked_date',
                           'unblock_date', 'recharge_date', 'sim_buy_date', 'sim_operator',
                           'owned_by', 'full_name', 'account_create_date','password']
        if field not in EDITABLE_FIELDS:
            return jsonify({"success": False, "error": f"Field '{field}' is not editable"})
        old_value = None
        platform = None
        old_number = None
        try:
            select_fields = ",".join(dict.fromkeys([field, "platform", "number"]))
            old_resp = social_supabase.table("social_media_accounts") \
                .select(select_fields).eq("id", account_id).limit(1).execute()
            if old_resp.data:
                old_value = old_resp.data[0].get(field)
                platform = old_resp.data[0].get('platform')
                old_number = old_resp.data[0].get('number')
        except Exception as e:
            print(f"[ACTIVITY LOG] Could not fetch old value: {e}")
        DATE_FIELDS = {'blocked_date', 'unblock_date', 'recharge_date', 'account_create_date', 'sim_buy_date'}
        if field in DATE_FIELDS:
            save_value = None if (not value or value.upper() in ('NA', 'N/A', 'NONE', 'NULL', '')) else value
        elif field == 'account_status':
            save_value = normalize_social_account_status(value) if value else "NA"
        else:
            save_value = value if value else "NA"
        update_payload = {field: save_value}
        if field == 'account_status' and save_value == 'Permanent Block':
            update_payload['blocked_date'] = datetime.now().strftime("%Y-%m-%d")
        response = social_supabase.table("social_media_accounts").update(update_payload).eq("id", account_id).execute()
        if hasattr(response, 'data'):
            if response.data:
                TOTAL_NUMBER_SYNC_FIELDS = {
                    'owned_by', 'number', 'sim_inserted_device', 'number_type',
                    'sim_operator', 'recharge_date', 'sim_buy_date'
                }
                if platform == "Total Numbers" and field in TOTAL_NUMBER_SYNC_FIELDS and old_number:
                    social_supabase.table("social_media_accounts") \
                        .update(update_payload) \
                        .eq("number", old_number) \
                        .execute()
                extra_info = {}
                if platform:
                    extra_info['platform'] = platform
                # Also store department so activity log can be filtered
                try:
                    dept_resp = social_supabase.table("social_media_accounts") \
                        .select("department").eq("id", account_id).limit(1).execute()
                    if dept_resp.data and dept_resp.data[0].get("department"):
                        extra_info['department'] = dept_resp.data[0]['department']
                except Exception:
                    pass
                log_activity(
                    action_type="field_update",
                    target_table="social_media_accounts",
                    target_record_id=account_id,
                    field_name=field,
                    old_value=old_value,
                    new_value=save_value,
                    extra_info=extra_info if extra_info else None
                )
                return jsonify({"success": True, "message": "Saved successfully", "updated_row": response.data[0]})
            verify = social_supabase.table("social_media_accounts").select("id").eq("id", account_id).execute()
            if verify.data:
                return jsonify({"success": False, "error": "Update failed — check Supabase API key permissions"})
            return jsonify({"success": False, "error": f"Row {account_id} not found"})
        return jsonify({"success": False, "error": "No response from Supabase"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/get-permanent-block-accounts", methods=["GET"])
@login_required
def get_permanent_block_accounts():
    try:
        search   = request.args.get("search",   "").strip()
        platform = request.args.get("platform", "").strip()
        query = social_supabase.table("social_media_accounts") \
            .select("id,owned_by,number,login_user,login_device,sim_inserted_device,blocked_date,account_create_date,platform,department") \
            .eq("account_status", "Permanent Block")
        # ── Department filter ───────────────────────────────────────────
        # Always enforce for non-superadmins; superadmins see all
        is_super       = is_superadmin(session)
        allowed_depts = session.get("allowed_departments")  # None = unrestricted
        if not is_super and allowed_depts:
            if len(allowed_depts) == 1:
                query = query.eq("department", allowed_depts[0])
            else:
                query = query.in_("department", allowed_depts)
        # If is_super OR allowed_depts is None → no department filter applied
        # ── Platform filter ────────────────────────────────────────────
        if platform:
            query = query.eq("platform", platform)
        # ── Search filter ──────────────────────────────────────────────
        if search:
            like_term = f"%{search}%"
            query = query.or_(
                f"owned_by.ilike.{like_term},"
                f"number.ilike.{like_term},"
                f"login_device.ilike.{like_term},"
                f"platform.ilike.{like_term}"
            )
        query = query.order("id", desc=False)
        response = query.execute()
        accounts = []
        for item in (response.data or []):
            item["login_user"] = "NA"
            item["sim_inserted_device"] = "NA"
            b_date_str      = item.get("blocked_date")        or ""
            create_date_str = item.get("account_create_date") or ""
            active_duration = "N/A"
            if b_date_str and create_date_str:
                try:
                    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                        try:
                            days = (
                                datetime.strptime(b_date_str[:10], fmt) -
                                datetime.strptime(create_date_str[:10], fmt)
                            ).days
                            active_duration = f"{days} days" if days >= 0 else "N/A"
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
            accounts.append({
                "id":              item.get("id"),
                "owned_by":        item.get("owned_by")    or "N/A",
                "number":          item.get("number")      or "N/A",
                "login_device":    item.get("login_device") or "N/A",
                "platform":        item.get("platform")    or "N/A",
                "department":      item.get("department")  or "N/A",
                "blocked_date":    b_date_str              or "N/A",
                "active_duration": active_duration,
            })
        return jsonify({"success": True, "accounts": accounts, "count": len(accounts)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ============================================================
# UPI VALIDITY CHECKER (integrated from Tampermonkey script)
# Background job + live progress + export
# ============================================================
UPI_CHECK_API_URL = "https://upi-api-new.onrender.com/check_upi"

DEFAULT_UPI_HANDLES = [
    "airtel", "apl", "freecharge", "ikwik", "mbk", "mbkns", "naviaxis","nye","nyes","superyes",
    "okbizaxis", "ptaxis", "pthdfc", "ptsbi", "ptyes", "slc", "upi",
    "yapl", "kotakbank", "jupiter", "icici", "mbank", "jupiteraxis",
    "axl", "ibl", "ybl", "fam", "mairtel", "jio"
]
# Sirf inhi handles ko -1 se -7 tak suffix milta hai
SUFFIX_UPI_HANDLES = ["axl", "ibl", "ybl"]

UPI_CHECK_JOBS = {}
UPI_CHECK_JOBS_LOCK = threading.Lock()
MAX_UPI_NUMBERS_PER_JOB = 3

def build_upi_candidates(number, custom_handles=None):
    custom_handles = custom_handles or []
    candidates = []
    for h in DEFAULT_UPI_HANDLES:
        candidates.append(f"{number}@{h}")
        if h in SUFFIX_UPI_HANDLES:
            for i in range(1, 8):
                candidates.append(f"{number}-{i}@{h}")
    for h in custom_handles:
        h = str(h).strip().lower()
        if h:
            candidates.append(f"{number}@{h}")
    return candidates

def verify_upi_external(upi, retries=2):
    """Per-number sequential + retry + longer timeout — too much parallel
    hammering on the free Render API caused failures that silently became
    'Unknown'."""
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(
                UPI_CHECK_API_URL,
                json={"upi": upi},
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            if resp.status_code != 200:
                last_err = f"HTTP {resp.status_code}"
                time.sleep(1.5)
                continue
            data = resp.json()
            status = (
                data.get("status") or data.get("Status")
                or data.get("result") or data.get("message")
                or json.dumps(data)
            )
            return str(status), None
        except Exception as e:
            last_err = str(e)
            time.sleep(1.5)
            continue
    return None, last_err or "request failed"

def _run_upi_check_worker(job_id, number, custom_handles):
    """Ek number ke saare candidates ko sequentially check karta hai.
    Multiple numbers alag-alag threads mein parallel chalte hain, lekin har
    number ke andar requests sequential rehti hain (API overload avoid)."""
    candidates = build_upi_candidates(number, custom_handles)
    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
        if not job:
            return
        job["total"] += len(candidates)

    for upi in candidates:
        with UPI_CHECK_JOBS_LOCK:
            job = UPI_CHECK_JOBS.get(job_id)
            if not job or job.get("stop"):
                if job:
                    job["current"][number] = None
                return
            job["current"][number] = upi

        status_text, error = verify_upi_external(upi)
        if error and not status_text:
            status_text = f"Error: {error}"
            normalized = "Error"
        else:
            t = (status_text or "").lower()
            if "invalid" in t or "not" in t:
                normalized = "Invalid"
            elif "valid" in t:
                normalized = "Valid"
            else:
                normalized = "Unknown"

        result = {
            "number": number,
            "upi": upi,
            "handle": upi.split("@")[1] if "@" in upi else "",
            "status": status_text,
            "normalized": normalized,
            "already_reported": False,
            "report_count": 0,
            "report_ids": [],
            "report_dates": [],
            "report_user": None,
        }

        if normalized == "Valid":
            try:
                resp = supabase.table("BS_Investment_Scam") \
                    .select("Id,Upi_vpa,Inserted_date,Scam_type,Input_user") \
                    .ilike("Upi_vpa", upi) \
                    .limit(10).execute()
                found = resp.data or []
                result["already_reported"] = len(found) > 0
                result["report_count"] = len(found)
                result["report_ids"] = [str(x.get("Id")) for x in found]
                result["report_dates"] = [x.get("Inserted_date") for x in found]
                result["report_user"] = found[0].get("Input_user") if found else None
            except Exception as e:
                print(f"[UPI CHECK] duplicate lookup error for {upi}: {e}")

        with UPI_CHECK_JOBS_LOCK:
            job = UPI_CHECK_JOBS.get(job_id)
            if not job:
                return
            job["results"].append(result)
            job["completed"] += 1

    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
        if job:
            job["current"][number] = None

def _run_upi_check_job(job_id, numbers, custom_handles):
    threads = []
    for number in numbers:
        t = threading.Thread(
            target=_run_upi_check_worker,
            args=(job_id, number, custom_handles),
            daemon=True
        )
        threads.append(t)
        t.start()
    for t in threads:
        t.join()

    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
        if job:
            job["status"] = "stopped" if job.get("stop") else "done"

@app.route("/start-upi-check", methods=["POST"])
@login_required
def start_upi_check():
    try:
        data = request.get_json() or {}
        # Backward compatible: single "number" ya "numbers" array dono chalega
        raw_numbers = data.get("numbers")
        if not raw_numbers:
            single = str(data.get("number", "")).strip()
            raw_numbers = [single] if single else []
        numbers = []
        for n in raw_numbers:
            n = str(n).strip()
            if n and n not in numbers:
                numbers.append(n)
        numbers = numbers[:MAX_UPI_NUMBERS_PER_JOB]
        custom_handles = data.get("custom_handles", []) or []
        if not numbers:
            return jsonify({"success": False, "error": "At least one mobile number is required"})

        job_id = uuid.uuid4().hex
        with UPI_CHECK_JOBS_LOCK:
            UPI_CHECK_JOBS[job_id] = {
                "status": "running",
                "total": 0,
                "completed": 0,
                "current": {n: None for n in numbers},
                "results": [],
                "numbers": numbers,
                "stop": False,
            }
        thread = threading.Thread(
            target=_run_upi_check_job,
            args=(job_id, numbers, custom_handles),
            daemon=True
        )
        thread.start()
        return jsonify({"success": True, "job_id": job_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/upi-check-status/<job_id>", methods=["GET"])
@login_required
def upi_check_status(job_id):
    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
        if not job:
            return jsonify({"success": False, "error": "Job not found"})
        return jsonify({
            "success": True,
            "status": job["status"],
            "total": job["total"],
            "completed": job["completed"],
            "current": job["current"],
            "numbers": job["numbers"],
            "results": job["results"],
        })

@app.route("/stop-upi-check/<job_id>", methods=["POST"])
@login_required
def stop_upi_check(job_id):
    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
        if job:
            job["stop"] = True
    return jsonify({"success": True})

@app.route("/export-upi-check/<job_id>", methods=["GET"])
@login_required
def export_upi_check(job_id):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    with UPI_CHECK_JOBS_LOCK:
        job = UPI_CHECK_JOBS.get(job_id)
    if not job:
        return jsonify({"success": False, "error": "Job not found"}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UPI Check Results"
    headers = ["Number", "UPI ID", "Status", "Already Reported", "Report Count", "Report Dates"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    blue        = PatternFill(start_color="FFD6EAF8", end_color="FFD6EAF8", fill_type="solid")
    light_green = PatternFill(start_color="FFC8F7C5", end_color="FFC8F7C5", fill_type="solid")
    light_red   = PatternFill(start_color="FFFFC9C9", end_color="FFFFC9C9", fill_type="solid")

    for r in job["results"]:
        is_valid = r["normalized"] == "Valid"
        is_invalid = r["normalized"] == "Invalid"
        is_reported = bool(r.get("already_reported"))

        if is_valid and is_reported:
            already_reported_text = "Valid and Reported UPI"
        elif is_valid and not is_reported:
            already_reported_text = "Valid and Not Reported Yet"
        elif is_invalid:
            already_reported_text = "Invalid UPI"
        else:
            already_reported_text = "No"

        row = [
            r.get("number", ""), r["upi"], r["status"],
            already_reported_text,
            r.get("report_count", 0),
            ", ".join([str(d) for d in (r.get("report_dates") or []) if d]),
        ]
        ws.append(row)
        excel_row = ws.max_row
        if is_valid and is_reported:
            ws.cell(row=excel_row, column=4).fill = light_green
        elif is_valid and not is_reported:
            ws.cell(row=excel_row, column=4).fill = blue
        elif is_invalid:
            ws.cell(row=excel_row, column=4).fill = light_red

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    numbers_str = "_".join(job.get("numbers", ["upi"]))
    return send_file(
        output,
        download_name=f"upi_check_{numbers_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/check-duplicates", methods=["POST"])
@login_required
def check_duplicates():
    try:
        data = request.get_json()
        entries = data.get("entries", [])
        if not entries:
            return jsonify({"success": False, "error": "No entries provided"})
        results = []
        for entry in entries:
            val = str(entry.get("value", "")).strip()
            typ = entry.get("type", "upi")
            if not val or val.upper() in ("NA", "N/A", "", "NONE"):
                continue
            try:
                if typ == "upi":
                    res = supabase.table("BS_Investment_Scam")\
                        .select("Id, Upi_vpa, Inserted_date, Scam_type, Input_user")\
                        .ilike("Upi_vpa", val)\
                        .limit(10).execute()
                else:
                    res = supabase.table("BS_Investment_Scam")\
                        .select("Id, Bank_account_number, Inserted_date, Scam_type, Input_user")\
                        .ilike("Bank_account_number", val)\
                        .limit(10).execute()
                found = res.data or []
                results.append({
                    "value": val,
                    "type": typ,
                    "status": "DUPLICATE" if found else "NEW",
                    "count": len(found),
                    "earliest_date": found[0].get("Inserted_date") if found else None,
                    "latest_date": found[-1].get("Inserted_date") if len(found) > 1 else None,
                    "scam_type": found[0].get("Scam_type") if found else None,
                    "input_user": found[0].get("Input_user") if found else None,
                    "record_ids": [str(r.get("Id")) for r in found]
                })
            except Exception as e:
                results.append({
                    "value": val, "type": typ,
                    "status": "ERROR", "count": 0,
                    "error": str(e)
                })
        total = len(results)
        duplicates = sum(1 for r in results if r["status"] == "DUPLICATE")
        new_entries = sum(1 for r in results if r["status"] == "NEW")
        return jsonify({
            "success": True,
            "results": results,
            "summary": {
                "total": total,
                "duplicates": duplicates,
                "new": new_entries,
                "errors": total - duplicates - new_entries
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

import urllib.request

@app.route("/get-session-info", methods=["GET"])
@login_required
def get_session_info():
    return jsonify({
        "email":         session.get("email", ""),
        "display_name":  session.get("display_name", ""),
        "is_admin":      session.get("is_admin", False),
        "role":          session.get("role", "user"),
        "allowed_depts": session.get("allowed_departments"),
        "allowed_pages": session.get("allowed_pages", []),
    })

@app.route("/getDepartmentData", methods=["GET"])
@login_required
def get_department_data_proxy():
    """Proxy for external MIS API to avoid CORS issues"""
    try:
        user_mail  = session.get("email", "")
        department = "Anti_Money_Laundering"
        role       = "Team_Lead"
        external_url = (
            f"https://mis-iw3m.onrender.com/getDepartmentData"
            f"?user_mail={urllib.parse.quote(user_mail)}"
            f"&department={urllib.parse.quote(department)}"
            f"&role={urllib.parse.quote(role)}"
        )
        req = urllib.request.Request(external_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return jsonify(data)
    except Exception as e:
        print(f"[MIS PROXY] Error: {e}")
        return jsonify([])

@app.route("/insert-social-record", methods=["POST"])
@login_required
def insert_social_record():
    try:
        if "social" not in session.get("allowed_pages", []):
            return jsonify({"success": False, "error": "Didnt have social page access"}), 403
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        DATE_FIELDS = {'blocked_date', 'unblock_date', 'account_create_date', 'recharge_date', 'sim_buy_date'}
        ALLOWED_FIELDS = [
            'platform', 'department', 'owned_by', 'login_user', 'number',
            'login_device', 'sim_inserted_device', 'account_status', 'review_status',
            'number_type', 'blocked_date', 'unblock_date', 'account_create_date',
            'sim_operator', 'full_name', 'recharge_date', 'sim_buy_date',
            'account_type', 'mail_id', 'account_id', 'password', 'page_name'
        ]
        record = {}
        for field in ALLOWED_FIELDS:
            val = str(data.get(field, '')).strip()
            if field in DATE_FIELDS:
                record[field] = val if val and val.upper() not in ('NA', 'N/A', 'NONE', 'NULL', '') else None
            else:
                record[field] = val if val else "NA"
        # platform is required
        if not record.get('platform') or record['platform'] == 'NA':
            return jsonify({"success": False, "error": "Platform is required"})
        if record.get('number') and record['number'] != 'NA':
            existing_perm = social_supabase.table("social_media_accounts") \
                .select("id") \
                .eq("platform", record['platform']) \
                .eq("number", record['number']) \
                .eq("account_status", "Permanent Block") \
                .limit(1) \
                .execute()
            if existing_perm.data:
                return jsonify({
                    "success": False,
                    "error": "Already exist in Permanent Block Account"
                })
        # get next id
        try:
            max_id_resp = social_supabase.table("social_media_accounts") \
                .select("id").order("id", desc=True).limit(1).execute()
            record['id'] = int(max_id_resp.data[0]['id']) + 1 if max_id_resp.data else 1
        except Exception:
            pass
        resp = social_supabase.table("social_media_accounts").insert(record).execute()
        if resp.data:
            inserted = resp.data[0]
            log_activity(
                action_type="import",
                target_table="social_media_accounts",
                extra_info={"file_name": "manual_insert", "records_count": 1}
            )
            return jsonify({"success": True, "record": inserted})
        return jsonify({"success": False, "error": "Insert failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# ADD THIS ROUTE TO app.py  (paste before the if __name__ == "__main__": block)
# ============================================================
@app.route("/insert-scraping-record", methods=["POST"])
@login_required
def insert_scraping_record():
    try:
        if "scraping" not in session.get("allowed_pages", []):
            return jsonify({"success": False, "error": "Didnt have scrapping page access"}), 403
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        rows = data.get("rows", [])
        if not rows:
            return jsonify({"success": False, "error": "No rows provided"})
        ALLOWED_FIELDS = [
            "name", "platform", "post_url", "chat_number", "group_name",
            "scam_type", "share_status", "screenshot",
            "chat_status", "assigned_to", "assigned_at_datetime",
            "inserted_datetime", "priority", "inserted_date",
            "extra_field_1", "extra_field_2", "extra_field_3",
            "extra_field_4", "extra_field_5"
        ]
        records = []
        today = datetime.now().strftime("%Y-%m-%d")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in rows:
            record = {}
            for field in ALLOWED_FIELDS:
                val = str(row.get(field, "")).strip()
                record[field] = val if val else "NA"
            record["name"] = get_clean_display_name(session.get("display_name", "User"))
            # defaults
            if record.get("inserted_date") in ("", "NA"):
                record["inserted_date"] = today
            if record.get("inserted_datetime") in ("", "NA"):
                record["inserted_datetime"] = now_str
            if record.get("screenshot") in ("", "NA"):
                record["screenshot"] = "NA"
            if record.get("share_status") in ("", "NA"):
                record["share_status"] = "Pending"
            if record.get("chat_status") in ("", "NA"):
                record["chat_status"] = "NA"
            if record.get("priority") in ("", "NA"):
                record["priority"] = "NA"
            for ef in ["extra_field_1","extra_field_2","extra_field_3","extra_field_4","extra_field_5"]:
                if record.get(ef) in ("", "NA"):
                    record[ef] = "NA"
            records.append(record)
        resp = supabase.table("scrapping_data").insert(records).execute()
        if resp.data:
            log_activity(
                action_type="import",
                target_table="scrapping_data",
                extra_info={"file_name": "manual_insert", "records_count": len(records)}
            )
            return jsonify({"success": True, "records": resp.data, "count": len(resp.data)})
        return jsonify({"success": False, "error": "Insert returned no data"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
def _extract_facebook_profile_id(*values):
    for value in values:
        raw = str(value or "").strip()
        if not raw or raw.upper() in ("NA", "N/A"):
            continue
        try:
            parsed = urllib.parse.urlparse(raw if raw.lower().startswith(("http://", "https://")) else "https://" + raw)
            host = (parsed.netloc or "").lower()
            if "facebook.com" not in host and "fb.com" not in host:
                continue
            qs = urllib.parse.parse_qs(parsed.query)
            fb_id = (qs.get("id") or [""])[0].strip()
            if fb_id and re.fullmatch(r"\d{5,}", fb_id):
                return fb_id
        except Exception:
            pass
        m = re.search(r"(?:[?&]id=)(\d{5,})", raw)
        if m:
            return m.group(1)
    return ""

def _normalize_scraping_duplicate_url(value):
    raw = str(value or "").strip()
    if not raw or raw.upper() in ("NA", "N/A"):
        return ""
    raw = re.sub(r"[\s#/)]+$", "", raw)
    try:
        parsed = urllib.parse.urlparse(raw if raw.lower().startswith(("http://", "https://")) else "https://" + raw)
        host = (parsed.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = urllib.parse.unquote(parsed.path or "")
        path = re.sub(r"[\s#/)]+$", "", path).rstrip("/")
        query = urllib.parse.urlencode(sorted(urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)))
        return f"{host}{path}{'?' + query if query else ''}"
    except Exception:
        return re.sub(r"[\s#/)]+$", "", raw.lower()).removeprefix("https://").removeprefix("http://").removeprefix("www.").rstrip("/")

def _scraping_url_lookup_term(value):
    normalized = _normalize_scraping_duplicate_url(value)
    if not normalized:
        return ""
    try:
        parsed = urllib.parse.urlparse("https://" + normalized)
        parts = [p for p in (parsed.path or "").split("/") if p]
        if parts:
            return re.sub(r"[^A-Za-z0-9_.-]", "", parts[-1])[:80]
        qs = urllib.parse.parse_qs(parsed.query)
        if qs:
            return re.sub(r"[^A-Za-z0-9_.-]", "", next(iter(qs.values()))[0])[:80]
    except Exception:
        pass
    return re.sub(r"[^A-Za-z0-9_.-]", "", normalized.split("?")[0].split("/")[-1])[:80]

def _normalize_chat_number(value):
    """Return a stripped, lowercase chat number, or '' if NA / missing.

    NA, N/A, None, blank/whitespace are all treated as "no chat number info".
    Used so duplicate detection compares chat numbers case-insensitively.
    """
    raw = str(value or "").strip()
    if not raw or raw.upper() in ("NA", "N/A"):
        return ""
    return raw.lower()


def _chat_numbers_match(input_cn, existing_cn):
    """Return True only if both chat numbers are concrete and equal (case-insensitive).

    Returns False when either side is NA / missing. That is the whole point:
    if post URL or group name already exists in the database but its chat number
    is missing OR different from the new row's chat number, the new row is NOT a
    duplicate and the data should still be inserted.
    """
    a = _normalize_chat_number(input_cn)
    b = _normalize_chat_number(existing_cn)
    return bool(a) and bool(b) and a == b


@app.route("/check-scraping-duplicates", methods=["POST"])
@login_required
def check_scraping_duplicates():
    try:
        data = request.get_json()
        entries = data.get("entries", [])
        if not entries:
            return jsonify({"success": True, "results": []})
        results = []
        facebook_id_counts = {}
        normalized_url_counts = {}
        for entry in entries:
            platform = str(entry.get("platform", "")).strip()
            normalized_url = _normalize_scraping_duplicate_url(entry.get("post_url", ""))
            if normalized_url:
                normalized_url_counts[normalized_url] = normalized_url_counts.get(normalized_url, 0) + 1
            if platform.lower() != "facebook":
                continue
            fb_id = _extract_facebook_profile_id(entry.get("post_url", ""), entry.get("group_name", ""))
            if fb_id:
                facebook_id_counts[fb_id] = facebook_id_counts.get(fb_id, 0) + 1
        for entry in entries:
            gn = str(entry.get("group_name", "")).strip()
            cn = str(entry.get("chat_number", "")).strip()
            post_url = str(entry.get("post_url", "")).strip()
            platform = str(entry.get("platform", "")).strip()
            is_facebook = platform.lower() == "facebook"
            facebook_id = _extract_facebook_profile_id(post_url, gn) if is_facebook else ""
            normalized_url = _normalize_scraping_duplicate_url(post_url)
            # Dono NA hain toh skip
            gn_empty = not gn or gn.upper() in ("NA", "N/A", "")
            cn_empty = not cn or cn.upper() in ("NA", "N/A", "")
            if gn_empty and cn_empty and not facebook_id and not normalized_url:
                results.append({"status": "NEW", "count": 0})
                continue
            try:
                found = []
                local_facebook_duplicate = bool(facebook_id and facebook_id_counts.get(facebook_id, 0) > 1)
                local_url_duplicate = bool(normalized_url and normalized_url_counts.get(normalized_url, 0) > 1)
                if not gn_empty and not cn_empty:
                    # Dono available — AND match
                    res = supabase.table("scrapping_data") \
                        .select("id, group_name, post_url, chat_number, inserted_date") \
                        .ilike("group_name", gn) \
                        .ilike("chat_number", cn) \
                        .limit(10).execute()
                    found = res.data or []
                elif not gn_empty:
                    # Sirf group_name
                    res = supabase.table("scrapping_data") \
                        .select("id, group_name, post_url, chat_number, inserted_date") \
                        .ilike("group_name", gn) \
                        .limit(10).execute()
                    found = res.data or []
                elif not cn_empty:
                    # Sirf chat_number
                    res = supabase.table("scrapping_data") \
                        .select("id, group_name, post_url, chat_number, inserted_date") \
                        .ilike("chat_number", cn) \
                        .limit(10).execute()
                    found = res.data or []
                if is_facebook and facebook_id and not found:
                    # Facebook-profile ID match hua; chat number bhi match hona
                    # chahiye — alag chat_number ke saath URL match duplicate nahi hai
                    # aur naya row insert hona chahiye.
                    res = supabase.table("scrapping_data") \
                        .select("id, group_name, post_url, chat_number, inserted_date") \
                        .eq("platform", "Facebook") \
                        .or_(f"post_url.ilike.%id={facebook_id}%,group_name.ilike.%id={facebook_id}%") \
                        .limit(25).execute()
                    candidates = res.data or []
                    found = [
                        r for r in candidates
                        if _chat_numbers_match(cn, r.get("chat_number"))
                    ]
                if normalized_url and not found:
                    lookup_term = _scraping_url_lookup_term(post_url)
                    if lookup_term:
                        res = supabase.table("scrapping_data") \
                            .select("id, group_name, post_url, chat_number, inserted_date") \
                            .or_(f"post_url.ilike.%{lookup_term}%,group_name.ilike.%{lookup_term}%") \
                            .limit(25).execute()
                        candidates = res.data or []
                        # Same post_url/group_name hone ke baad bhi chat_number
                        # match hona chahiye — URL match alone duplicate nahi hai.
                        # Agar existing record ka chat_number alag ya NA hai,
                        # toh naya row insert allow karo.
                        found = [
                            row for row in candidates
                            if _chat_numbers_match(cn, row.get("chat_number"))
                            and normalized_url in (
                                _normalize_scraping_duplicate_url(row.get("post_url", "")),
                                _normalize_scraping_duplicate_url(row.get("group_name", "")),
                            )
                        ]
                results.append({
                    "status": "DUPLICATE" if found or local_facebook_duplicate or local_url_duplicate else "NEW",
                    "count": len(found),
                    "earliest_date": found[0].get("inserted_date") if found else None,
                })
            except Exception as e:
                results.append({"status": "ERROR", "count": 0, "error": str(e)})
        return jsonify({"success": True, "results": results})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/check-chat-number", methods=["POST"])
@login_required
def check_chat_number():
    try:
        data = request.get_json()
        chat_number = str(data.get("chat_number", "")).strip()
        if not chat_number or chat_number.upper() in ("NA", "N/A", ""):
            return jsonify({"exists": False})
        res = supabase.table("scrapping_data") \
            .select("id, inserted_date, name") \
            .ilike("chat_number", chat_number) \
            .limit(5).execute()
        found = res.data or []
        return jsonify({
            "exists": len(found) > 0,
            "count": len(found),
            "first_seen": found[0].get("inserted_date") if found else None,
            "inserted_by": found[0].get("name") if found else None,
        })
    except Exception as e:
        return jsonify({"exists": False, "error": str(e)})
    
@app.route("/scrapping-summary-data", methods=["GET"])
@login_required
def scrapping_summary_data():
    """Fetch scrapping data for summary generation (filtered by date)"""
    try:
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to",   "").strip()
        date_on   = request.args.get("date_on",   "").strip()   # single date shortcut
        CHUNK = 1000
        all_rows = []
        offset = 0
        while True:
            q = supabase.table("scrapping_data") \
                .select("name,platform,chat_number,group_name,scam_type,inserted_date")
            if date_on:
                q = q.eq("inserted_date", date_on)
            else:
                if date_from:
                    q = q.gte("inserted_date", date_from)
                if date_to:
                    q = q.lte("inserted_date", date_to)
            resp = q.order("id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        return jsonify({"success": True, "rows": all_rows, "total": len(all_rows)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/delete-social-record", methods=["POST"])
@login_required
def delete_social_record():
    try:
        data = request.get_json()
        record_id = data.get("id")
        if not record_id:
            return jsonify({"success": False, "error": "No ID provided"})
        # Fetch info before deleting for logging
        try:
            old_resp = social_supabase.table("social_media_accounts") \
                .select("platform,login_user,number").eq("id", record_id).limit(1).execute()
            old_info = old_resp.data[0] if old_resp.data else {}
        except Exception:
            old_info = {}
        social_supabase.table("social_media_accounts") \
            .delete().eq("id", record_id).execute()
        log_activity(
            action_type="field_update",
            target_table="social_media_accounts",
            target_record_id=record_id,
            field_name="DELETE",
            old_value=f"platform={old_info.get('platform','?')}, number={old_info.get('number','?')}",
            new_value="DELETED",
            extra_info={"platform": old_info.get("platform")}
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
@app.route("/get-scraping-record/<int:record_id>", methods=["GET"])
@login_required
def get_scraping_record(record_id):
    try:
        resp = supabase.table("scrapping_data") \
            .select("id,name,platform,post_url,chat_number,group_name,scam_type,share_status,inserted_date,screenshot") \
            .eq("id", record_id).limit(1).execute()
        if resp.data:
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Record not found"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
@app.route("/update-scraping-record", methods=["POST"])
@login_required
def update_scraping_record():
    try:
        data = request.get_json()
        record_id = data.get("id")
        if not record_id:
            return jsonify({"success": False, "error": "No ID"})
        ALLOWED = ["platform", "post_url", "chat_number", "group_name", "scam_type", "share_status"]
        updates = {k: v for k, v in data.items() if k in ALLOWED}
        if not updates:
            return jsonify({"success": False, "error": "No valid fields to update"})
        resp = supabase.table("scrapping_data").update(updates).eq("id", record_id).execute()
        log_activity(
            action_type="field_update",
            target_table="scrapping_data",
            target_record_id=record_id,
            field_name=",".join(updates.keys()),
            new_value=str(updates),
        )
        return jsonify({"success": True, "record": resp.data[0] if resp.data else {}})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
@app.route("/delete-scraping-record", methods=["POST"])
@login_required
def delete_scraping_record():
    try:
        data = request.get_json()
        record_id = data.get("id")
        if not record_id:
            return jsonify({"success": False, "error": "No ID"})
        supabase.table("scrapping_data").delete().eq("id", record_id).execute()
        log_activity(
            action_type="field_update",
            target_table="scrapping_data",
            target_record_id=record_id,
            field_name="DELETE",
            old_value="EXISTS",
            new_value="DELETED"
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
@app.route("/my-scraping-count", methods=["GET"])
@login_required
def my_scraping_count():
    try:
        display_name = session.get("display_name", "")
        clean_name = re.sub(r'\s*\(.*?\)\s*', '', display_name).strip()
        if not clean_name:
            return jsonify({"success": True, "count": 0})
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to",   "").strip()
        q = supabase.table("scrapping_data") \
            .select("id", count='exact') \
            .ilike("name", f"%{clean_name}%")
        if date_from:
            q = q.gte("inserted_date", date_from)
        if date_to:
            q = q.lte("inserted_date", date_to)
        resp = q.execute()
        return jsonify({"success": True, "count": resp.count or 0})
    except Exception as e:
        return jsonify({"success": False, "count": 0, "error": str(e)})
# ============================================================
# INSIGHTS — BS Investment Scam Analytics
# ============================================================
@app.route("/investment-insights-data", methods=["GET"])
@login_required
def investment_insights_data():
    if "insights" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        date_from   = request.args.get("date_from",   "").strip()
        date_to     = request.args.get("date_to",     "").strip()
        search_for  = request.args.get("search_for",  "").strip()
        scam_type   = request.args.get("scam_type",   "").strip()
        wallet      = request.args.get("wallet",      "").strip()
        input_user  = request.args.get("input_user",  "").strip()
        is_sm_search = search_for in ("__SM__", "SM Counts")
        CHUNK = 1000
        all_rows, offset = [], 0
        while True:
            q = supabase.table("BS_Investment_Scam").select(
                "Inserted_date,Input_user,Search_for,Scam_type,"
                "Upi_bank_account_wallet,Upi_vpa,Bank_account_number,Web_contact_no,Bank_name"
            )
            if date_from:  q = q.gte("Inserted_date", date_from)
            if date_to:    q = q.lte("Inserted_date", date_to)
            if is_sm_search:
                q = q.in_("Search_for", BS_INVESTMENT_SM_SEARCH_FOR_VALUES)
            elif search_for:
                q = q.eq("Search_for", search_for)
            if scam_type:  q = q.eq("Scam_type",  scam_type)
            if wallet:     q = q.eq("Upi_bank_account_wallet", wallet)
            if input_user: q = q.eq("Input_user", input_user)
            resp  = q.order("Id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        rows = [{k.lower(): v for k, v in r.items()} for r in all_rows]
        # ---------- unique UPI per date ----------
        upi_by_date = {}
        upi_set = set()
        bank_set = set()
        dated_rows = []
        for r in rows:
            d      = (r.get("inserted_date") or "")[:10]
            wallet_val = (r.get("upi_bank_account_wallet") or "").strip()
            upi    = (r.get("upi_vpa") or "").strip()
            bank_acc = (r.get("bank_account_number") or "").strip()
            try:
                parsed_date = datetime.strptime(d, "%Y-%m-%d").date() if d else None
            except ValueError:
                parsed_date = None
            if parsed_date:
                dated_rows.append((r, parsed_date))
            if not d: continue
            if wallet_val == "UPI" and upi and upi.upper() not in ("NA", "N/A", ""):
                if d not in upi_by_date:
                    upi_by_date[d] = set()
                upi_by_date[d].add(upi)
                upi_set.add(upi)
            if wallet_val == "Bank Account" and bank_acc and bank_acc.upper() not in ("NA", "N/A", ""):
                bank_set.add(bank_acc)
        # upi_series: cumulative unique UPIs seen up to each date
        # This ensures chart total matches the card total
        seen_upis = set()
        upi_series = {}
        for d in sorted(upi_by_date.keys()):
            seen_upis.update(upi_by_date[d])
            upi_series[d] = len(upi_by_date[d])
        def _trend_bucket(start_date, end_date):
            case_count = 0
            upis = set()
            banks = set()
            for row, parsed_date in dated_rows:
                if not (start_date <= parsed_date <= end_date):
                    continue
                case_count += 1
                wallet_val = (row.get("upi_bank_account_wallet") or "").strip()
                upi_val = (row.get("upi_vpa") or "").strip()
                bank_val = (row.get("bank_account_number") or "").strip()
                if wallet_val == "UPI" and upi_val and upi_val.upper() not in ("NA", "N/A", ""):
                    upis.add(upi_val)
                if wallet_val == "Bank Account" and bank_val and bank_val.upper() not in ("NA", "N/A", ""):
                    banks.add(bank_val)
            return {
                "cases": case_count,
                "upi": len(upis),
                "bank": len(banks),
            }
        def _trend_series(start_date, end_date):
            buckets = {}
            current_date = start_date
            while current_date <= end_date:
                buckets[current_date] = {"cases": 0, "upis": set(), "banks": set()}
                current_date += timedelta(days=1)
            for row, parsed_date in dated_rows:
                if parsed_date not in buckets:
                    continue
                buckets[parsed_date]["cases"] += 1
                wallet_val = (row.get("upi_bank_account_wallet") or "").strip()
                upi_val = (row.get("upi_vpa") or "").strip()
                bank_val = (row.get("bank_account_number") or "").strip()
                if wallet_val == "UPI" and upi_val and upi_val.upper() not in ("NA", "N/A", ""):
                    buckets[parsed_date]["upis"].add(upi_val)
                if wallet_val == "Bank Account" and bank_val and bank_val.upper() not in ("NA", "N/A", ""):
                    buckets[parsed_date]["banks"].add(bank_val)
            ordered_dates = sorted(buckets)
            return {
                "labels": [d.isoformat() for d in ordered_dates],
                "cases": [buckets[d]["cases"] for d in ordered_dates],
                "upi": [len(buckets[d]["upis"]) for d in ordered_dates],
                "bank": [len(buckets[d]["banks"]) for d in ordered_dates],
            }
        def _trend_metric(current_value, previous_value):
            delta = current_value - previous_value
            pct = None if previous_value == 0 else round((delta / previous_value) * 100, 1)
            return {"current": current_value, "previous": previous_value, "delta": delta, "percent": pct}
        if dated_rows:
            trend_end = max(parsed_date for _, parsed_date in dated_rows)
            trend_start = trend_end - timedelta(days=29)
            prev_end = trend_start - timedelta(days=1)
            prev_start = prev_end - timedelta(days=29)
            current_trend = _trend_bucket(trend_start, trend_end)
            previous_trend = _trend_bucket(prev_start, prev_end)
        else:
            trend_end = datetime.utcnow().date()
            trend_start = trend_end - timedelta(days=29)
            prev_end = trend_start - timedelta(days=1)
            prev_start = prev_end - timedelta(days=29)
            current_trend = {"cases": 0, "upi": 0, "bank": 0}
            previous_trend = {"cases": 0, "upi": 0, "bank": 0}
        trend_30d = {
            "period_start": trend_start.isoformat(),
            "period_end": trend_end.isoformat(),
            "previous_start": prev_start.isoformat(),
            "previous_end": prev_end.isoformat(),
            "cases": _trend_metric(current_trend["cases"], previous_trend["cases"]),
            "upi": _trend_metric(current_trend["upi"], previous_trend["upi"]),
            "bank": _trend_metric(current_trend["bank"], previous_trend["bank"]),
            "series": _trend_series(trend_start, trend_end),
        }
        # ---------- user counts per date ----------
        user_by_date = {}
        for r in rows:
            d    = (r.get("inserted_date") or "")[:10]
            user = (r.get("input_user") or "Unknown").strip()
            if not d: continue
            if d not in user_by_date:
                user_by_date[d] = {}
            user_by_date[d][user] = user_by_date[d].get(user, 0) + 1
        all_users = sorted({u for dmap in user_by_date.values() for u in dmap})
        # ---------- scam type counts ----------
        scam_counts = {}
        for r in rows:
            st = (r.get("scam_type") or "Unknown").strip() or "Unknown"
            scam_counts[st] = scam_counts.get(st, 0) + 1
        # ---------- search_for counts ----------
        sf_counts = {}
        for r in rows:
            sf = (r.get("search_for") or "Unknown").strip() or "Unknown"
            wc = (r.get("web_contact_no") or "").strip()
            has_contact = wc and wc.upper() not in ("NA", "N/A", "", "NONE", "NULL")
            if has_contact:
                sf_counts["WhatsApp"] = sf_counts.get("WhatsApp", 0) + 1
            else:
                # No contact number → original sf mein count karo
                sf_counts[sf] = sf_counts.get(sf, 0) + 1
        # Normalize WhatsApp variants (case fix)
        for key in list(sf_counts.keys()):
            if key.lower() == "whatsapp" and key != "WhatsApp":
                sf_counts["WhatsApp"] = sf_counts.get("WhatsApp", 0) + sf_counts.pop(key)
        # ---------- bank name counts (merged in from /investment-bank-data) ----------
        bank_counts = {}
        monthly_counts = {}
        for r in rows:
            bn = (r.get("bank_name") or "").strip()
            if bn and bn.upper() not in ("NA", "N/A", "") and bn.lower() != "unknown":
                bank_counts[bn] = bank_counts.get(bn, 0) + 1
            d7 = (r.get("inserted_date") or "")[:7]
            if d7:
                monthly_counts[d7] = monthly_counts.get(d7, 0) + 1
        sorted_banks = dict(sorted(bank_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        monthly_counts = {k: monthly_counts[k] for k in sorted(monthly_counts)}

        # ── Average daily cases ──────────────────────────────
        active_dates = [d for d in user_by_date.keys() if d]
        if active_dates:
            if input_user:
                user_daily = {
                    d: user_by_date[d].get(input_user, 0)
                    for d in active_dates
                    if user_by_date[d].get(input_user, 0) > 0
                }
                avg_cases = round(sum(user_daily.values()) / len(user_daily), 1) if user_daily else 0
            else:
                total_daily = {d: sum(user_by_date[d].values()) for d in active_dates}
                avg_cases = round(sum(total_daily.values()) / len(total_daily), 1) if total_daily else 0
        else:
            avg_cases = 0
        # ── Per-user stats for comparison ───────────────────
        user_stats = {}
        for u in all_users:
            u_upi_set  = set()
            u_bank_set = set()
            u_dates    = set()
            u_total    = 0
            for r in rows:
                ru = (r.get("input_user") or "Unknown").strip()
                if ru != u:
                    continue
                u_total += 1
                d = (r.get("inserted_date") or "")[:10]
                if d:
                    u_dates.add(d)
                wallet_val = (r.get("upi_bank_account_wallet") or "").strip()
                upi_val    = (r.get("upi_vpa") or "").strip()
                bank_val   = (r.get("bank_account_number") or "").strip()
                if wallet_val == "UPI" and upi_val and upi_val.upper() not in ("NA", "N/A", ""):
                    u_upi_set.add(upi_val)
                if wallet_val == "Bank Account" and bank_val and bank_val.upper() not in ("NA", "N/A", ""):
                    u_bank_set.add(bank_val)
            u_avg = round(u_total / len(u_dates), 1) if u_dates else 0
            user_stats[u] = {
                "total":       u_total,
                "avg":         u_avg,
                "unique_upi":  len(u_upi_set),
                "unique_bank": len(u_bank_set),
            }
        return jsonify({
            "success": True,
            "total_rows": len(rows),
            "unique_upi_count": len(upi_set),
            "unique_bank_count": len(bank_set),
            "avg_cases": avg_cases,
            "trend_30d": trend_30d,
            "upi_series":   upi_series,
            "user_by_date": {d: user_by_date[d] for d in sorted(user_by_date)},
            "all_users":    all_users,
            "scam_counts":  scam_counts,
            "sf_counts":    sf_counts,
            "user_stats":   user_stats,
            "all_input_users": sorted(list({(r.get("input_user") or "Unknown").strip() for r in rows if r.get("input_user")})),
            "bank_counts": sorted_banks,
            "monthly_counts": monthly_counts,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# NOTE: /investment-bank-data ab bhi maujood hai backward-compat ke liye,
# lekin dashboard isko alag se call nahi karta — bank_counts/monthly_counts
# ab /investment-insights-data ke response mein hi bundled aate hain, jisse
# har dashboard load pe full-table Supabase scan 2 se ghatkar 1 ho gaya hai.
@app.route("/investment-bank-data", methods=["GET"])
@login_required
def investment_bank_data():
    try:
        date_from  = request.args.get("date_from",  "").strip()
        date_to    = request.args.get("date_to",    "").strip()
        search_for = request.args.get("search_for", "").strip()
        scam_type  = request.args.get("scam_type",  "").strip()
        wallet     = request.args.get("wallet",     "").strip()
        input_user = request.args.get("input_user", "").strip()
        is_sm_search = search_for in ("__SM__", "SM Counts")
        CHUNK = 1000
        all_rows, offset = [], 0
        while True:
            q = supabase.table("BS_Investment_Scam").select(
                "Bank_name,Inserted_date,Search_for,Input_user"
            )
            if date_from:  q = q.gte("Inserted_date", date_from)
            if date_to:    q = q.lte("Inserted_date", date_to)
            if is_sm_search:
                q = q.in_("Search_for", BS_INVESTMENT_SM_SEARCH_FOR_VALUES)
            elif search_for:
                q = q.eq("Search_for", search_for)
            if scam_type:  q = q.eq("Scam_type", scam_type)
            if wallet:     q = q.eq("Upi_bank_account_wallet", wallet)
            if input_user: q = q.eq("Input_user", input_user)
            resp  = q.order("Id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        rows = [{k.lower(): v for k, v in r.items()} for r in all_rows]
        bank_counts = {}
        for r in rows:
            bn = (r.get("bank_name") or "").strip()
            if not bn or bn.upper() in ("NA", "N/A", "") or bn.lower() == "unknown":
                continue
            bank_counts[bn] = bank_counts.get(bn, 0) + 1
        sorted_banks = sorted(bank_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        monthly_counts = {}
        for r in rows:
            d = (r.get("inserted_date") or "")[:7]
            if not d:
                continue
            monthly_counts[d] = monthly_counts.get(d, 0) + 1
        return jsonify({
            "success": True,
            "bank_counts": dict(sorted_banks),
            "monthly_counts": {k: monthly_counts[k] for k in sorted(monthly_counts)},
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# WEBSITE DIRECTORY — LIST / FILTER
# ============================================================
@app.route("/website-directory", methods=["GET"])
@login_required
def website_directory():
    user = get_current_user()
    allowed_pages = session.get("allowed_pages", [])
    if "website_directory" not in allowed_pages:
        flash("You don't have access to Website Directory.", "error")
        return redirect_to_allowed_page(allowed_pages)
    wd_search    = request.args.get("wd_search",   "").strip()
    wd_remark    = request.args.get("wd_remark",   "").strip()
    wd_category  = request.args.get("wd_category", "").strip()
    wd_search_for= request.args.get("wd_search_for","").strip()
    wd_date_from = request.args.get("wd_date_from","").strip()
    wd_date_to   = request.args.get("wd_date_to",  "").strip()
    page = int(request.args.get("page_num", 1))
    items = []
    total_rows = 0
    total_pages = 1
    try:
        query = supabase.table("website_directory").select("*", count="exact")
        query = query.or_("remark.is.null,remark.eq.NA,remark.eq.,remark.eq.IPG")
        if wd_search:
            lt = f"%{wd_search}%"
            query = query.or_(
                f"url.ilike.{lt},"
                f"final_url.ilike.{lt},"
                f"name.ilike.{lt},"
                f"group_app_name.ilike.{lt},"
                f"login_id.ilike.{lt},"
                f"number.ilike.{lt},"
                f"email.ilike.{lt},"
                f"invitation_code.ilike.{lt},"
                f"password.ilike.{lt},"
                f"remark.ilike.{lt},"
                f"origin.ilike.{lt},"
                f"category.ilike.{lt},"
                f"search_for.ilike.{lt},"
                f"automated_website.ilike.{lt},"
                f"payment_gateway.ilike.{lt}"
            )         
        if wd_remark:
            query = query.eq("remark", wd_remark)
        if wd_category:
            query = query.eq("category", wd_category)
        if wd_search_for:
            query = query.eq("search_for", wd_search_for)
        if wd_date_from:
            query = query.gte("date", wd_date_from)
        if wd_date_to:
            query = query.lte("date", wd_date_to)
        query = query.order("id", desc=True)
        offset = (page - 1) * PER_PAGE
        query = query.range(offset, offset + PER_PAGE - 1)
        resp = query.execute()
        items = resp.data or []
        total_rows = resp.count or 0
        total_pages = max(1, math.ceil(total_rows / PER_PAGE))
    except Exception as e:
        flash(f"Error fetching website directory: {e}", "error")
    clean_display_name = get_clean_display_name(session.get("display_name", "User"))
    return render_template(
        "website_directory.html",
        items=items,
        wd_search=wd_search,
        wd_remark=wd_remark,
        wd_category=wd_category,
        wd_search_for=wd_search_for,
        wd_date_from=wd_date_from,
        wd_date_to=wd_date_to,
        page_num=page,
        total_pages=total_pages,
        total_rows=total_rows,
        wd_category_options=WEBSITE_DIRECTORY_CATEGORY_OPTIONS,
        wd_search_for_options=WEBSITE_DIRECTORY_SEARCH_FOR_OPTIONS,
        current_user=user,
        allowed_pages=allowed_pages,
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        can_view_activity_log=session.get("can_view_activity_log", False),
        is_admin=session.get("is_admin", False),
        role=session.get("role", "user"),
    )
# ============================================================
# WEBSITE DIRECTORY — IMPORT
# ============================================================
@app.route("/website-directory-import", methods=["POST"])
@login_required
def website_directory_import():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("No file selected", "error")
        return redirect("/website-directory")
    if not is_allowed_file(file.filename):
        flash("Unsupported file type.", "error")
        return redirect("/website-directory")
    try:
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "csv"
        df = read_data_file(temp_path, file_ext)
        df.columns = df.columns.astype(str).str.strip()
        df = df.fillna("")
        COL_MAP = {
            "date":             ["date", "Date", "DATE"],
            "name":             ["name", "Name", "NAME"],
            "url":              ["url", "URL", "Url", "website_url", "Website URL"],
            "final_url":        ["final_url", "Final URL", "FinalURL", "final url"],
            "invitation_code":  ["invitation_code", "Invitation Code", "InvitationCode", "invite_code"],
            "search_for":       ["search_for", "Search For", "SearchFor", "search for"],
            "group_app_name":   ["group_app_name", "Group/App Name", "Group App Name", "GroupAppName", "group_name", "app_name"],
            "number":           ["number", "Number", "Phone", "phone", "Mobile", "mobile"],
            "email":            ["email", "Email", "EMAIL", "mail"],
            "login_id":         ["login_id", "Login ID", "LoginID", "login id", "Login Id"],
            "password":         ["password", "Password", "PASSWORD", "pass"],
            "remark":           ["remark", "Remark", "REMARK", "remarks", "Remarks"],
            "origin":           ["origin", "Origin", "ORIGIN"],
            "category":         ["category", "Category", "CATEGORY", "scam_type", "Scam Type"],
            "automated_website":["automated_website", "Automated Website", "AutomatedWebsite", "automated"],
            "payment_gateway":  ["payment_gateway", "Payment Gateway", "PaymentGateway", "gateway"],
        }
        def resolve_col(target_cols):
            for c in target_cols:
                if c in df.columns:
                    return c
            for c in target_cols:
                for col in df.columns:
                    if c.lower() == col.lower():
                        return col
            return None
        DATE_FIELDS = {"date"}
        records = []
        for _, row in df.iterrows():
            rec = {}
            for db_col, candidates in COL_MAP.items():
                src = resolve_col(candidates)
                val = str(row[src]).strip() if src else ""
                if db_col in DATE_FIELDS:
                    rec[db_col] = val if val and val.upper() not in ("NA","N/A","","NONE","NULL") else None
                else:
                    rec[db_col] = val if val else "NA"
            records.append(rec)
        supabase.table("website_directory").insert(records).execute()
        log_activity(
            action_type="import",
            target_table="website_directory",
            extra_info={"file_name": filename, "records_count": len(records)}
        )
        flash(f"Imported successfully! {len(records)} records added.", "success")
        os.remove(temp_path)
    except Exception as e:
        flash(f"Import Error: {e}", "error")
    return redirect("/website-directory")
# ============================================================
# WEBSITE DIRECTORY — EXPORT
# ============================================================
@app.route("/website-directory-export", methods=["GET"])
@login_required
def website_directory_export():
    try:
        wd_search     = request.args.get("wd_search",    "").strip()
        wd_remark     = request.args.get("wd_remark",    "").strip()
        wd_category   = request.args.get("wd_category",  "").strip()
        wd_search_for = request.args.get("wd_search_for","").strip()
        wd_date_from  = request.args.get("wd_date_from", "").strip()
        wd_date_to    = request.args.get("wd_date_to",   "").strip()

        def _build_wd_query():
            q = supabase.table("website_directory").select("*")
            q = q.or_("remark.is.null,remark.eq.NA,remark.eq.,remark.eq.IPG")
            if wd_search:
                lt = f"%{wd_search}%"
                q = q.or_(
                    f"url.ilike.{lt},"
                    f"final_url.ilike.{lt},"
                    f"name.ilike.{lt},"
                    f"group_app_name.ilike.{lt},"
                    f"login_id.ilike.{lt},"
                    f"number.ilike.{lt},"
                    f"email.ilike.{lt},"
                    f"invitation_code.ilike.{lt},"
                    f"password.ilike.{lt},"
                    f"remark.ilike.{lt},"
                    f"origin.ilike.{lt},"
                    f"category.ilike.{lt},"
                    f"search_for.ilike.{lt},"
                    f"automated_website.ilike.{lt},"
                    f"payment_gateway.ilike.{lt}"
                )
            if wd_remark:    q = q.eq("remark", wd_remark)
            if wd_category:  q = q.eq("category",      wd_category)
            if wd_search_for:q = q.eq("search_for",    wd_search_for)
            if wd_date_from: q = q.gte("date",          wd_date_from)
            if wd_date_to:   q = q.lte("date",          wd_date_to)
            return q

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return _stream_supabase_csv(_build_wd_query, "id", f"website_directory_{ts}.csv")
    except Exception as e:
        flash(f"Export Error: {e}", "error")
        return redirect("/website-directory")
# ============================================================
# WEBSITE DIRECTORY — TRACKER STATS
# ============================================================
@app.route("/website-directory-tracker-stats", methods=["GET"])
@login_required
def website_directory_tracker_stats():
    try:
        CHUNK = 1000
        all_rows, offset = [], 0
        # ── Fetch ALL rows first (inoperable records excluded — same filter
        # as the main directory list view) ──
        while True:
            resp = supabase.table("website_directory") \
                .select("category,search_for,remark,date,name") \
                .or_("remark.is.null,remark.eq.NA,remark.eq.,remark.eq.IPG") \
                .order("id", desc=False) \
                .range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        # ── Aggregation AFTER loop ──
        cat_counts        = {}
        sf_counts         = {}
        remark_counts     = {}
        daily_counts      = {}
        user_total_counts = {}
        user_cat_counts   = {}
        for row in all_rows:
            cat  = (row.get("category")   or "Unknown").strip() or "Unknown"
            sf   = (row.get("search_for") or "Unknown").strip() or "Unknown"
            rem  = (row.get("remark")     or "").strip()
            dt   = (row.get("date")       or "")[:10]
            user = (row.get("name")       or "Unknown").strip() or "Unknown"
            if user.upper() in ("NA", "N/A", ""):
                user = "Unknown"
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            sf_counts[sf]   = sf_counts.get(sf,  0) + 1
            if rem and rem.upper() not in ("NA", "N/A", ""):
                remark_counts[rem] = remark_counts.get(rem, 0) + 1
            if dt:
                daily_counts[dt] = daily_counts.get(dt, 0) + 1
            user_total_counts[user] = user_total_counts.get(user, 0) + 1
            if user not in user_cat_counts:
                user_cat_counts[user] = {}
            user_cat_counts[user][cat] = user_cat_counts[user].get(cat, 0) + 1
        return jsonify({
            "success": True,
            "total": len(all_rows),
            "cat_counts":   cat_counts,
            "sf_counts":    sf_counts,
            "remark_counts": dict(
                sorted(remark_counts.items(), key=lambda x: x[1], reverse=True)[:20]
            ),
            "daily_counts": {k: daily_counts[k] for k in sorted(daily_counts)},
            "user_total_counts": dict(
                sorted(user_total_counts.items(), key=lambda x: x[1], reverse=True)
            ),
            "user_cat_counts": user_cat_counts,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# WEBSITE DIRECTORY — INSERT SINGLE RECORD
# ============================================================
@app.route("/website-directory-insert", methods=["POST"])
@login_required
def website_directory_insert():
    try:
        if "website_directory" not in session.get("allowed_pages", []):
            return jsonify({"success": False, "error": "Didnt have Website directory access"}), 403
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data"})
        ALLOWED = [
            "date","name","url","final_url","invitation_code","search_for",
            "group_app_name","number","email","login_id","password",
            "remark","origin","category","automated_website","payment_gateway"
        ]
        record = {}
        for f in ALLOWED:
            val = str(data.get(f, "")).strip()
            if f == "date":
                record[f] = val if val and val.upper() not in ("NA","N/A","") else None
            else:
                record[f] = val if val else "NA"
        resp = supabase.table("website_directory").insert(record).execute()
        if resp.data:
            log_activity(
                action_type="import",
                target_table="website_directory",
                extra_info={"file_name": "manual_insert", "records_count": 1}
            )
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Insert failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/website-directory-update", methods=["POST"])
@login_required
def website_directory_update():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data"})
        rid = data.get("id")
        if rid is None:
            return jsonify({"success": False, "error": "No ID"})

        ALLOWED = [
            "date","name","url","final_url","invitation_code","search_for",
            "group_app_name","number","email","login_id","password",
            "remark","origin","category","automated_website","payment_gateway"
        ]
        record = {}
        for f in ALLOWED:
            if f not in data:
                continue  # Only update fields that are sent
            val = str(data.get(f, "")).strip()
            if f == "date":
                record[f] = val if val and val.upper() not in ("NA","N/A","") else None
            elif f == "remark":
                record[f] = val  # Allow empty string for remark
            else:
                record[f] = val if val else "NA"
        resp = supabase.table("website_directory").update(record).eq("id", rid).execute()
        if resp.data:
            log_activity(
                action_type="field_update",
                target_table="website_directory",
                target_record_id=rid,
                field_name="edit",
                new_value=str(record)
            )
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Update failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/website-directory-get-record", methods=["GET"])
@login_required
def website_directory_get_record():
    try:
        rid = request.args.get("id")
        if not rid:
            return jsonify({"success": False, "error": "No ID"})
        resp = supabase.table("website_directory").select("*").eq("id", rid).limit(1).execute()
        if resp.data:
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Record not found"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# WEBSITE DIRECTORY — DELETE SINGLE RECORD
# ============================================================
@app.route("/website-directory-delete", methods=["POST"])
@login_required
def website_directory_delete():
    try:
        data = request.get_json()
        rid = data.get("id")
        if not rid:
            return jsonify({"success": False, "error": "No ID"})
        supabase.table("website_directory").delete().eq("id", rid).execute()
        log_activity(
            action_type="field_update",
            target_table="website_directory",
            target_record_id=rid,
            field_name="DELETE",
            old_value="EXISTS",
            new_value="DELETED"
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
@app.route("/website-directory-delete-bulk", methods=["POST"])
@login_required
def website_directory_delete_bulk():
    try:
        data = request.get_json()
        ids = data.get("ids", [])
        if not ids:
            return jsonify({"success": False, "error": "No IDs provided"})
        clean_ids = []
        for item in ids:
            try:
                clean_ids.append(int(str(item).strip()))
            except (ValueError, TypeError):
                continue
        if not clean_ids:
            return jsonify({"success": False, "error": "No valid IDs"})
        supabase.table("website_directory").delete().in_("id", clean_ids).execute()
        log_activity(
            action_type="field_update",
            target_table="website_directory",
            field_name="DELETE_BULK",
            old_value="EXISTS",
            new_value="DELETED",
            extra_info={"ids": clean_ids, "count": len(clean_ids)}
        )
        return jsonify({"success": True, "deleted": len(clean_ids)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/website-directory-search-api", methods=["GET"])
@login_required
def website_directory_search_api():
    try:
        q = request.args.get("q", "").strip()
        if not q:
            return jsonify({"success": True, "items": []})
        like_term = f"%{q}%"
        resp = supabase.table("website_directory") \
            .select("id,url,final_url,search_for,login_id,password,origin,category,invitation_code") \
            .or_(f"url.ilike.{like_term},final_url.ilike.{like_term}") \
            .order("id", desc=True) \
            .limit(5) \
            .execute()
        items = resp.data or []
        return jsonify({"success": True, "items": items})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# WEBSITE DIRECTORY — DOWNLOAD TEMPLATE
# ============================================================
@app.route("/website-directory-template", methods=["GET"])
@login_required
def website_directory_template():
    headers = [
        "date","name","url","final_url","invitation_code","search_for",
        "group_app_name","number","email","login_id","password",
        "remark","origin","category","automated_website","payment_gateway"
    ]
    out = io.StringIO()
    csv.writer(out).writerow(headers)
    out.seek(0)
    return send_file(
        io.BytesIO(out.getvalue().encode("utf-8-sig")),
        download_name="Website_Directory_Template.csv",
        as_attachment=True,
        mimetype="text/csv"
    )

@app.route("/website-directory-user-summary", methods=["GET"])
@login_required
def website_directory_user_summary():
    try:
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to",   "").strip()
        date_on   = request.args.get("date_on",   "").strip()
        CHUNK = 1000
        all_rows, offset = [], 0
        while True:
            q = supabase.table("website_directory").select("name")
            if date_on:
                q = q.eq("date", date_on)
            else:
                if date_from:
                    q = q.gte("date", date_from)
                if date_to:
                    q = q.lte("date", date_to)
            resp = q.order("id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        user_counts = {}
        for row in all_rows:
            name = (row.get("name") or "Unknown").strip() or "Unknown"
            if name.upper() in ("NA", "N/A", ""):
                name = "Unknown"
            user_counts[name] = user_counts.get(name, 0) + 1

        return jsonify({"success": True, "user_counts": user_counts, "total": len(all_rows)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/website-directory-summary-stats", methods=["GET"])
@login_required
def website_directory_summary_stats():
    try:
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to",   "").strip()
        date_on   = request.args.get("date_on",   "").strip()
        CHUNK = 1000
        all_rows, offset = [], 0
        while True:
            q = supabase.table("website_directory").select("category,url,final_url")
            if date_on:
                q = q.eq("date", date_on)
            else:
                if date_from:
                    q = q.gte("date", date_from)
                if date_to:
                    q = q.lte("date", date_to)
            resp = q.order("id", desc=False).range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK

        def _clean_category(value):
            cat = (value or "Unknown").strip() or "Unknown"
            return "Unknown" if cat.upper() in ("NA", "N/A", "NULL", "UNKNOWN", "") else cat

        def _url_keys(value):
            raw = (value or "").strip()
            if not raw or raw.upper() in ("NA", "N/A", "NULL"):
                return []
            keys = {raw.lower().rstrip("/")}
            parsed = urlparse(raw if "://" in raw else f"https://{raw}")
            host = (parsed.netloc or "").lower()
            path = (parsed.path or "").strip("/")
            if host.startswith("www."):
                host = host[4:]
            if host:
                keys.add(host)
                if path:
                    keys.add(f"{host}/{path}".rstrip("/"))
            return [k for k in keys if k]

        known_category_by_url = {}
        known_rows, known_offset = [], 0
        while True:
            known_resp = supabase.table("website_directory") \
                .select("category,url,final_url") \
                .order("id", desc=False) \
                .range(known_offset, known_offset + CHUNK - 1).execute()
            known_chunk = known_resp.data or []
            known_rows.extend(known_chunk)
            if len(known_chunk) < CHUNK:
                break
            known_offset += CHUNK

        for row in known_rows:
            cat = _clean_category(row.get("category"))
            if cat == "Unknown":
                continue
            for key in _url_keys(row.get("url")) + _url_keys(row.get("final_url")):
                known_category_by_url.setdefault(key, cat)

        cat_counts = {}
        for row in all_rows:
            cat = _clean_category(row.get("category"))
            if cat == "Unknown":
                for key in _url_keys(row.get("url")) + _url_keys(row.get("final_url")):
                    if key in known_category_by_url:
                        cat = known_category_by_url[key]
                        break
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
        return jsonify({"success": True, "cat_counts": cat_counts, "total": len(all_rows)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/website-directory-inoperable", methods=["GET"])
@login_required
def website_directory_inoperable():
    try:
        offset = int(request.args.get("offset", 0))
        limit  = int(request.args.get("limit", 1000))
        resp = supabase.table("website_directory") \
            .select("id,date,name,url,final_url,search_for,login_id,password,remark,origin,category,group_app_name") \
            .not_.eq("remark", "IPG") \
            .not_.ilike("remark", "Allotted to %") \
            .not_.is_("remark", "null") \
            .order("id", desc=True) \
            .range(offset, offset + limit - 1) \
            .execute()
        return jsonify({"success": True, "items": resp.data or []})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/scam-website-allotment-bulk-match", methods=["POST"])
@login_required
def scam_website_allotment_bulk_match():
    """Bulk-pasted URLs ko website_directory mein match karke unke IDs +
    details return karta hai — Website Allot modal ke 'Paste URLs' tab ke liye.
    Ek hi request mein saari URLs ke against DB search karta hai (chunked)."""
    if not is_allotment_admin(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json(force=True) or {}
        raw_urls = data.get("urls", [])
        if not raw_urls:
            return jsonify({"success": False, "error": "No URLs provided"})

        cleaned_urls = []
        for u in raw_urls:
            u = str(u).strip()
            if u:
                cleaned_urls.append(u)
        if not cleaned_urls:
            return jsonify({"success": False, "error": "No valid URLs found"})

        def _domain_of(u):
            try:
                netloc = urlparse(u if u.startswith("http") else "https://" + u).netloc
                return netloc[4:] if netloc.startswith("www.") else netloc
            except Exception:
                return ""

        results = []
        seen_ids = set()
        for u in cleaned_urls:
            like_term = f"%{u}%"
            found_row = None
            try:
                resp = supabase.table("website_directory") \
                    .select("id,date,url,final_url,search_for,login_id,password,remark,origin,category,group_app_name") \
                    .or_(f"url.ilike.{like_term},final_url.ilike.{like_term}") \
                    .order("id", desc=True) \
                    .limit(1).execute()
                if resp.data:
                    found_row = resp.data[0]
            except Exception:
                found_row = None

            if not found_row:
                domain = _domain_of(u)
                if domain:
                    try:
                        like_domain = f"%{domain}%"
                        resp2 = supabase.table("website_directory") \
                            .select("id,date,url,final_url,search_for,login_id,password,remark,origin,category,group_app_name") \
                            .or_(f"url.ilike.{like_domain},final_url.ilike.{like_domain}") \
                            .order("id", desc=True) \
                            .limit(1).execute()
                        if resp2.data:
                            found_row = resp2.data[0]
                    except Exception:
                        found_row = None

            if found_row and found_row.get("id") not in seen_ids:
                seen_ids.add(found_row.get("id"))
                results.append({
                    "pasted_url": u,
                    "found": True,
                    "id": found_row.get("id"),
                    "date": found_row.get("date"),
                    "url": found_row.get("url"),
                    "final_url": found_row.get("final_url"),
                    "search_for": found_row.get("search_for"),
                    "category": found_row.get("category"),
                    "remark": found_row.get("remark"),
                })
            else:
                results.append({"pasted_url": u, "found": False})

        # ── Allotment status bhi bata do (sirf AAJ ke allotment count hote hain —
        # purane din allot hui website aaj phir se available maani jaati hai) ──
        matched_ids = [r["id"] for r in results if r.get("found")]
        allotment_status_by_id = {}
        if matched_ids:
            try:
                matched_rows = [r for r in results if r.get("found")]
                url_keys = set()
                for r in matched_rows:
                    category = (r.get("category") or "").strip()
                    display_url = r.get("final_url") if category == "Investment Scam" else (r.get("url") or r.get("final_url"))
                    url_keys.add(((display_url or "").strip().lower()))
                today_str = datetime.now().strftime("%Y-%m-%d")
                allot_resp = supabase.table(WEBSITE_ALLOTMENT_TABLE) \
                    .select("final_url,category,search_for,alloted_user_name") \
                    .gte("allotted_at", today_str + " 00:00:00") \
                    .lte("allotted_at", today_str + " 23:59:59") \
                    .execute()
                allot_map = {}
                for row in allot_resp.data or []:
                    key = (row.get("final_url") or "").strip().lower()
                    if key:
                        allot_map[key] = row.get("alloted_user_name")
                for r in matched_rows:
                    category = (r.get("category") or "").strip()
                    display_url = r.get("final_url") if category == "Investment Scam" else (r.get("url") or r.get("final_url"))
                    key = (display_url or "").strip().lower()
                    if key in allot_map and allot_map[key]:
                        allotment_status_by_id[r["id"]] = f"Allotted to {allot_map[key]}"
            except Exception as e:
                print(f"[BULK MATCH] allotment status lookup error: {e}")

        for r in results:
            if r.get("found"):
                status = allotment_status_by_id.get(r["id"])
                r["is_already_allotted"] = bool(status)
                r["allotment_status"] = status or "Available"

        found_count = sum(1 for r in results if r.get("found"))
        return jsonify({
            "success": True,
            "results": results,
            "total_pasted": len(cleaned_urls),
            "found_count": found_count,
            "not_found_count": len(cleaned_urls) - found_count,
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/website-directory-operable", methods=["GET"])
@login_required
def website_directory_operable():
    """Website Allot picker ke liye website list.
    Already-allotted websites bhi bhejte hain taaki UI status dikha sake,
    lekin frontend un rows ko selectable nahi rakhta."""
    try:
        wd_category   = request.args.get("wd_category", "").strip()
        wd_search_for = request.args.get("wd_search_for", "").strip()
        wd_date_from  = request.args.get("wd_date_from", "").strip()
        wd_date_to    = request.args.get("wd_date_to", "").strip()
        if not wd_date_from and not wd_date_to:
            today_str = datetime.now().strftime("%Y-%m-%d")
            wd_date_from = today_str
            wd_date_to = today_str
        offset = int(request.args.get("offset", 0))
        limit  = int(request.args.get("limit", 1000))
        query = supabase.table("website_directory") \
            .select("id,date,name,url,final_url,search_for,login_id,password,remark,origin,category,group_app_name")
        if wd_category:
            query = query.eq("category", wd_category)
        if wd_search_for:
            query = query.eq("search_for", wd_search_for)
        query = query.order("id", desc=True).range(offset, offset + limit - 1)
        resp = query.execute()
        allotment_status_by_key = {}
        try:
            allot_query = supabase.table(WEBSITE_ALLOTMENT_TABLE) \
                .select("final_url,category,search_for,alloted_user_name")
            if wd_category:
                allot_query = allot_query.eq("category", wd_category)
            if wd_search_for:
                allot_query = allot_query.eq("search_for", wd_search_for)
            if wd_date_from:
                allot_query = allot_query.gte("allotted_at", wd_date_from + " 00:00:00")
            if wd_date_to:
                allot_query = allot_query.lte("allotted_at", wd_date_to + " 23:59:59")
            allot_resp = allot_query.execute()
            for allot_row in allot_resp.data or []:
                allot_url = (allot_row.get("final_url") or "").strip().lower()
                allot_category = (allot_row.get("category") or "").strip()
                allot_search_for = (allot_row.get("search_for") or "").strip()
                allot_user = (allot_row.get("alloted_user_name") or "").strip()
                if allot_url and allot_user:
                    allotment_status_by_key[(allot_category, allot_search_for, allot_url)] = f"Allotted to {allot_user}"
        except Exception as status_lookup_exc:
            print(f"[ALLOT LIST] Could not load allotment status map: {status_lookup_exc}")

        items = []
        for row in resp.data or []:
            remark = (row.get("remark") or "").strip()
            remark_normalized = remark.upper()
            if (
                remark
                and remark_normalized not in ("NA", "N/A", "IPG")
                and not remark.lower().startswith("allotted to ")
            ):
                continue
            category = (row.get("category") or "").strip()
            search_for = (row.get("search_for") or "").strip()
            display_url = row.get("final_url") if category == "Investment Scam" else (row.get("url") or row.get("final_url"))
            status_from_allotment = allotment_status_by_key.get((category, search_for, (display_url or "").strip().lower()))
            is_allotted = bool(status_from_allotment)
            row["allotment_status"] = status_from_allotment or "Available"
            row["is_already_allotted"] = is_allotted
            items.append(row)
        return jsonify({"success": True, "items": items})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

# ============================================================
# SCAM++ WEBSITE ALLOTMENT
# ============================================================
@app.route("/scam-website-allotment", methods=["GET"])
@login_required
def scam_website_allotment():
    user = get_current_user()
    allowed_pages = session.get("allowed_pages", [])
    if not can_access_allotment(session):
        flash("You don't have access to Website Allotment.", "error")
        return redirect_to_allowed_page(allowed_pages)

    is_admin_allot = is_allotment_admin(session)
    current_display_name = session.get("display_name", "")
    clean_display_name_filter = get_clean_display_name(current_display_name)

    sa_search     = request.args.get("sa_search", "").strip()
    sa_category   = request.args.get("sa_category", "").strip()
    sa_search_for = request.args.get("sa_search_for", "").strip()
    sa_remark     = request.args.get("sa_remark", "").strip()
    sa_date_from  = request.args.get("sa_date_from", "").strip()
    sa_date_to    = request.args.get("sa_date_to", "").strip()
    sa_alloted_user = request.args.get("sa_alloted_user", "").strip()
    # Default view — koi bhi date filter na diya ho toh aaj ka allotment dikhao
    if not sa_date_from and not sa_date_to:
        today_str = datetime.now().strftime("%Y-%m-%d")
        sa_date_from = today_str
        sa_date_to = today_str
    page = int(request.args.get("page_num", 1))

    items = []
    total_rows = 0
    total_pages = 1
    try:
        # Fetch only the page rows first. We deliberately do NOT pass
        # `count="exact"` here — on a large `website_allotment` table the
        # combined `select("*", count="exact")` query forces Postgres to
        # materialize a full COUNT, which can take >60s and gets killed
        # upstream by Cloudflare with a 522 "Connection timed out" page
        # that breaks downstream JSON parsing.
        query = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("*")

        # Data isolation — plain "allotment" users only see rows allotted to their own name.
        # ilike wildcard so old records saved with "(Role)" suffix and new clean-name
        # records both match against the logged-in user's clean display name.
        if not is_admin_allot:
            query = query.ilike("alloted_user_name", f"{clean_display_name_filter}%")

        if sa_alloted_user:
            query = query.eq("alloted_user_name", sa_alloted_user)

        if sa_search:
            lt = f"%{sa_search}%"
            query = query.or_(
                f"final_url.ilike.{lt},"
                f"login_id.ilike.{lt},"
                f"password.ilike.{lt},"
                f"remark.ilike.{lt},"
                f"category.ilike.{lt},"
                f"search_for.ilike.{lt},"
                f"alloted_user_name.ilike.{lt}"
            )
        if sa_category:
            query = query.eq("category", sa_category)
        if sa_search_for:
            query = query.eq("search_for", sa_search_for)
        if sa_remark == "PENDING":
            query = query.or_("remark.is.null,remark.eq.,remark.eq.NA")
        elif sa_remark:
            query = query.eq("remark", sa_remark)
        if sa_date_from:
            query = query.gte("allotted_at", sa_date_from + " 00:00:00")
        if sa_date_to:
            query = query.lte("allotted_at", sa_date_to + " 23:59:59")

        query = query.order("id", desc=True)
        offset = (page - 1) * PER_PAGE
        query = query.range(offset, offset + PER_PAGE - 1)
        resp = query.execute()
        items = resp.data or []

        # Separate, lightweight COUNT. We use `select("id")` (cheapest
        # projection), restrict to a single row with `range(0, 0)`, and
        # ask PostgREST for an exact count via the `count` kwarg. This
        # query plan only needs an index on `id` (which is the primary
        # key) and times out orders of magnitude less often than the old
        # combined `count="exact"` on `select("*")`. We isolate it in
        # its own try/except so a slow COUNT never blocks the rows we
        # already have — pagination falls back to a "1+ pages" estimate.
        try:
            count_query = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("id", count="exact")
            if not is_admin_allot:
                count_query = count_query.ilike("alloted_user_name", f"{clean_display_name_filter}%")
            if sa_alloted_user:
                count_query = count_query.eq("alloted_user_name", sa_alloted_user)
            if sa_search:
                lt = f"%{sa_search}%"
                count_query = count_query.or_(
                    f"final_url.ilike.{lt},"
                    f"login_id.ilike.{lt},"
                    f"password.ilike.{lt},"
                    f"remark.ilike.{lt},"
                    f"category.ilike.{lt},"
                    f"search_for.ilike.{lt},"
                    f"alloted_user_name.ilike.{lt}"
                )
            if sa_category:
                count_query = count_query.eq("category", sa_category)
            if sa_search_for:
                count_query = count_query.eq("search_for", sa_search_for)
            if sa_remark == "PENDING":
                count_query = count_query.or_("remark.is.null,remark.eq.,remark.eq.NA")
            elif sa_remark:
                count_query = count_query.eq("remark", sa_remark)
            if sa_date_from:
                count_query = count_query.gte("allotted_at", sa_date_from + " 00:00:00")
            if sa_date_to:
                count_query = count_query.lte("allotted_at", sa_date_to + " 23:59:59")
            count_resp = count_query.range(0, 0).execute()
            total_rows = count_resp.count or 0
        except Exception as count_exc:
            # COUNT timed out / 522 / any DB hiccup — show rows anyway,
            # and report "1 page" so the pagination control is at worst
            # disabled on the first page load instead of crashing the page.
            print(f"[ALLOTMENT] count query failed: {count_exc}")
            total_rows = len(items)
            if total_rows >= PER_PAGE:
                total_rows = (offset + PER_PAGE) + 1  # at least one more page exists

        total_pages = max(1, math.ceil(total_rows / PER_PAGE)) if total_rows else 1
    except Exception as e:
        flash(f"Error fetching website allotment: {e}", "error")

    # Admin-only: list of users who have websites allotted, with their counts,
    # so the "Allotted User Name" filter dropdown can show how many each user has.
    allotment_user_counts = []
    if is_admin_allot:
        try:
            # Cap the loop at 10k rows so a large `website_allotment` table
            # cannot push this past the 30s Supabase client timeout (and
            # trigger a Cloudflare 522). If a user's allotment date filter
            # is set we narrow the scan to that window via the date fields
            # already requested in `sa_date_from` / `sa_date_to`.
            _uc_counter = {}
            _uc_start = 0
            _UC_PAGE = 1000
            _UC_MAX_ROWS = 10000  # hard cap to protect request latency
            _uc_rows_seen = 0
            while _uc_rows_seen < _UC_MAX_ROWS:
                _uc_q = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("alloted_user_name")
                # Narrow the scan with the same date window the table is
                # showing, so on a big table the loop only walks today's
                # allotment instead of all history.
                if sa_date_from:
                    _uc_q = _uc_q.gte("allotted_at", sa_date_from + " 00:00:00")
                if sa_date_to:
                    _uc_q = _uc_q.lte("allotted_at", sa_date_to + " 23:59:59")
                _uc_resp = _uc_q.range(_uc_start, _uc_start + _UC_PAGE - 1).execute()
                _uc_rows = _uc_resp.data or []
                for _uc_r in _uc_rows:
                    _uc_n = (_uc_r.get("alloted_user_name") or "").strip()
                    if _uc_n and _uc_n not in ("NA", "N/A"):
                        _uc_counter[_uc_n] = _uc_counter.get(_uc_n, 0) + 1
                _uc_rows_seen += len(_uc_rows)
                if len(_uc_rows) < _UC_PAGE:
                    break
                _uc_start += _UC_PAGE
            allotment_user_counts = sorted(_uc_counter.items(), key=lambda kv: (-kv[1], kv[0]))
        except Exception:
            allotment_user_counts = []

    target_completed = False
    if not is_admin_allot:
        try:
            target_query = supabase.table(WEBSITE_ALLOTMENT_TABLE) \
                .select("remark") \
                .ilike("alloted_user_name", f"{clean_display_name_filter}%")
            # Congrats banner ab wahi date-range respect karta hai jo table mein dikhaya ja raha hai
            if sa_date_from:
                target_query = target_query.gte("allotted_at", sa_date_from + " 00:00:00")
            if sa_date_to:
                target_query = target_query.lte("allotted_at", sa_date_to + " 23:59:59")
            target_rows_resp = target_query.execute()
            target_rows = target_rows_resp.data or []
            total_target = len(target_rows)
            pending_target = sum(
                1 for r in target_rows
                if r.get("remark") is None
                or not str(r.get("remark")).strip()
                or str(r.get("remark")).strip().upper() in ("NA", "N/A")
            )
            target_completed = total_target > 0 and pending_target == 0
        except Exception:
            target_completed = False

    clean_display_name = get_clean_display_name(session.get("display_name", "User"))
    return render_template(
        "scam_website_allotment.html",
        target_completed=target_completed,
        items=items,
        sa_search=sa_search,
        sa_category=sa_category,
        sa_search_for=sa_search_for,
        sa_remark=sa_remark,
        sa_date_from=sa_date_from,
        sa_date_to=sa_date_to,
        sa_alloted_user=sa_alloted_user,
        allotment_user_counts=allotment_user_counts,
        page_num=page,
        total_pages=total_pages,
        total_rows=total_rows,
        wd_category_options=WEBSITE_DIRECTORY_CATEGORY_OPTIONS,
        wd_search_for_options=WEBSITE_DIRECTORY_SEARCH_FOR_OPTIONS,
        allotment_remark_options=ALLOTMENT_REMARK_OPTIONS,
        is_admin_allot=is_admin_allot,
        current_user=user,
        allowed_pages=allowed_pages,
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        can_view_activity_log=session.get("can_view_activity_log", False),
        is_admin=session.get("is_admin", False),
        role=session.get("role", "user"),
    )


@app.route("/scam-website-allotment-users", methods=["GET"])
@login_required
def scam_website_allotment_users():
    """List of users eligible to receive allotments (for the admin's dropdown)."""
    if not is_allotment_admin(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        client = get_auth_supabase()
        res = client.table("dashboard_users") \
            .select("id,display_name,allowed_pages") \
            .eq("is_active", True) \
            .execute()
        rows = res.data or []
        names = sorted({
            get_clean_display_name(r.get("display_name")) for r in rows
            if r.get("display_name") and (
                "allotment" in (r.get("allowed_pages") or []) or
                "allotment_admin" in (r.get("allowed_pages") or [])
            )
        })
        return jsonify({"success": True, "users": [{"display_name": n} for n in names]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/scam-website-allotment-allot", methods=["POST"])
@login_required
def scam_website_allotment_allot():
    if not is_allotment_admin(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json(force=True) or {}
        target_name = (data.get("alloted_user_name") or "").strip()
        website_ids = data.get("website_ids") or []
        if not target_name or not website_ids:
            return jsonify({"success": False, "error": "User and websites are required."})

        dir_resp = supabase.table("website_directory") \
            .select("id,url,final_url,login_id,password,remark,category,search_for") \
            .in_("id", website_ids) \
            .execute()
        dir_rows = dir_resp.data or []
        prepared_rows = []
        display_urls = []
        for row in dir_rows:
            category = (row.get("category") or "").strip()
            display_url = row.get("final_url") if category == "Investment Scam" else (row.get("url") or row.get("final_url"))
            prepared_rows.append({
                "dir_row": row,
                "category": category,
                "display_url": display_url,
            })
            if display_url:
                display_urls.append(display_url)

        existing_by_url = {}
        if display_urls:
            try:
                existing_resp = supabase.table(WEBSITE_ALLOTMENT_TABLE) \
                    .select("id,final_url,category,search_for") \
                    .in_("final_url", display_urls) \
                    .execute()
                for erow in existing_resp.data or []:
                    key = (erow.get("final_url") or "").strip().lower()
                    if key:
                        # Agar same URL ke multiple purane allotment records hain
                        # (different users ko), sabse latest wale ko reuse karo.
                        existing_by_url.setdefault(key, []).append(erow)
            except Exception as e:
                print(f"[ALLOT] existing lookup error: {e}")

        insert_rows = []
        update_count = 0
        now_iso = datetime.now().isoformat()

        for prepped in prepared_rows:
            row = prepped["dir_row"]
            category = prepped["category"]
            display_url = prepped["display_url"]
            url_key = (display_url or "").strip().lower()
            existing_matches = existing_by_url.get(url_key, []) if url_key else []

            if existing_matches:
                # Reassign the most recent existing allotment row for this URL to
                # today's date + new user, instead of creating a duplicate.
                target_row = sorted(existing_matches, key=lambda r: r.get("id", 0), reverse=True)[0]
                update_payload = {
                    "alloted_user_name": target_name,
                    "login_id":          row.get("login_id"),
                    "password":          row.get("password"),
                    "remark":            None,
                    "category":          category,
                    "search_for":        row.get("search_for"),
                    "allotted_at":       now_iso,
                }
                supabase.table(WEBSITE_ALLOTMENT_TABLE) \
                    .update(update_payload) \
                    .eq("id", target_row["id"]) \
                    .execute()
                update_count += 1
            else:
                insert_rows.append({
                    "alloted_user_name": target_name,
                    "final_url":         display_url,
                    "login_id":          row.get("login_id"),
                    "password":          row.get("password"),
                    # Website Directory ka purana remark (e.g. "IPG") copy NAHI karna hai —
                    # naya allotment hamesha "pending" state se start hona chahiye taaki
                    # target-completion tabhi true ho jab user khud remark select kare.
                    "remark":            None,
                    "category":          category,
                    "search_for":        row.get("search_for"),
                    "allotted_at":       now_iso,
                })

        if insert_rows:
            supabase.table(WEBSITE_ALLOTMENT_TABLE).insert(insert_rows).execute()

        total_allotted = len(insert_rows) + update_count
        if total_allotted:
            log_activity("allotment_create", target_table=WEBSITE_ALLOTMENT_TABLE,
                         extra_info=f"Allotted {len(insert_rows)} new + reassigned {update_count} existing website(s) to {target_name}")

        return jsonify({"success": True, "allotted": total_allotted, "new": len(insert_rows), "reassigned": update_count})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/scam-website-allotment-reassign", methods=["POST"])
@login_required
def scam_website_allotment_reassign():
    """Admin-only: ek allotment row ko naye user ko reassign karta hai —
    purane user se allotment hat jaata hai, naye user ko mil jaata hai."""
    if not is_allotment_admin(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json(force=True) or {}
        rid = data.get("id")
        new_user = (data.get("new_user_name") or "").strip()
        if not rid or not new_user:
            return jsonify({"success": False, "error": "Record ID and new user are required."})

        row_resp = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("*").eq("id", rid).limit(1).execute()
        if not row_resp.data:
            return jsonify({"success": False, "error": "Allotment record not found."})
        old_row = row_resp.data[0]
        old_user = old_row.get("alloted_user_name")

        update_payload = {
            "alloted_user_name": new_user,
            "remark": None,  # naye user ke liye pending state se restart
            "allotted_at": datetime.now().isoformat(),
        }
        resp = supabase.table(WEBSITE_ALLOTMENT_TABLE).update(update_payload).eq("id", rid).execute()
        if resp.data:
            log_activity("allotment_reassign", target_table=WEBSITE_ALLOTMENT_TABLE,
                         target_record_id=rid,
                         old_value=old_user, new_value=new_user)
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Reassign failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/scam-website-allotment-update-remark", methods=["POST"])
@login_required
def scam_website_allotment_update_remark():
    """Sirf website_allotment table ka remark update karta hai —
    website_directory table par koi asar nahi padta."""
    if not can_access_allotment(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json(force=True) or {}
        rid = data.get("id")
        remark = (data.get("remark") or "").strip()
        if not rid:
            return jsonify({"success": False, "error": "No ID provided."})
        resp = supabase.table(WEBSITE_ALLOTMENT_TABLE).update({"remark": remark}).eq("id", rid).execute()
        if resp.data:
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Update failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    
@app.route("/scam-website-allotment-counts", methods=["GET"])
@login_required
def scam_website_allotment_counts():
    """Filter bar ke Total/Remaining Allotment Count ke liye — remark ke
    basis pe 'remaining' (pending) vs 'total' (sab allotted) calculate karta hai."""
    try:
        is_admin_allot = is_allotment_admin(session)
        date_from = request.args.get("sa_date_from", "").strip()
        date_to = request.args.get("sa_date_to", "").strip()
        if not date_from and not date_to:
            today_str = datetime.now().strftime("%Y-%m-%d")
            date_from = today_str
            date_to = today_str
        query = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("remark")
        if not is_admin_allot:
            clean_name = get_clean_display_name(session.get("display_name", ""))
            query = query.ilike("alloted_user_name", f"{clean_name}%")
        if date_from:
            query = query.gte("allotted_at", date_from + " 00:00:00")
        if date_to:
            query = query.lte("allotted_at", date_to + " 23:59:59")
        resp = query.execute()
        rows = resp.data or []
        total = len(rows)
        remaining = sum(
            1 for r in rows
            if r.get("remark") is None
            or not str(r.get("remark")).strip()
            or str(r.get("remark")).strip().upper() in ("NA", "N/A")
        )
        return jsonify({"success": True, "total": total, "remaining": remaining, "is_admin": is_admin_allot})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/scam-website-allotment-check-target", methods=["GET"])
@login_required
def scam_website_allotment_check_target():
    """Remark update ke turant baad, bina page refresh kiye, congrats banner
    ka status recalculate karne ke liye lightweight endpoint. Currently applied
    date filter (agar hai) usi ke hisaab se target check karta hai."""
    if is_allotment_admin(session):
        return jsonify({"success": True, "target_completed": False})
    try:
        date_from = request.args.get("sa_date_from", "").strip()
        date_to   = request.args.get("sa_date_to", "").strip()
        # Filter na ho toh bhi "today" default maan kar check karo — jaisa
        # index route mein hota hai — taaki default view par bhi banner sahi
        # se dikhe.
        if not date_from and not date_to:
            today_str = datetime.now().strftime("%Y-%m-%d")
            date_from = today_str
            date_to = today_str
        clean_name = get_clean_display_name(session.get("display_name", ""))

        target_query = supabase.table(WEBSITE_ALLOTMENT_TABLE) \
            .select("remark") \
            .ilike("alloted_user_name", f"{clean_name}%")
        if date_from:
            target_query = target_query.gte("allotted_at", date_from + " 00:00:00")
        if date_to:
            target_query = target_query.lte("allotted_at", date_to + " 23:59:59")
        target_rows_resp = target_query.execute()
        target_rows = target_rows_resp.data or []
        total_target = len(target_rows)
        pending_target = sum(
            1 for r in target_rows
            if r.get("remark") is None
            or not str(r.get("remark")).strip()
            or str(r.get("remark")).strip().upper() in ("NA", "N/A")
        )
        target_completed = total_target > 0 and pending_target == 0
        return jsonify({"success": True, "target_completed": target_completed})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/scam-website-allotment-delete-row", methods=["POST"])
@login_required
def scam_website_allotment_delete_row():
    """Admin-only: website_allotment table se ek row permanently delete karta hai."""
    if not is_allotment_admin(session):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json(force=True) or {}
        rid = data.get("id")
        if not rid:
            return jsonify({"success": False, "error": "No ID provided."})
        supabase.table(WEBSITE_ALLOTMENT_TABLE).delete().eq("id", rid).execute()
        log_activity("allotment_delete", target_table=WEBSITE_ALLOTMENT_TABLE, target_record_id=rid)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/scam-website-allotment-export", methods=["GET"])
@login_required
def scam_website_allotment_export():
    if not can_access_allotment(session):
        flash("You don't have access to Website Allotment.", "error")
        return redirect_to_allowed_page(session.get("allowed_pages", []))
    try:
        is_admin_allot = is_allotment_admin(session)
        query = supabase.table(WEBSITE_ALLOTMENT_TABLE).select("*")
        if not is_admin_allot:
            clean_export_name = get_clean_display_name(session.get("display_name", ""))
            query = query.ilike("alloted_user_name", f"{clean_export_name}%")

        sa_search     = request.args.get("sa_search", "").strip()
        sa_category   = request.args.get("sa_category", "").strip()
        sa_search_for = request.args.get("sa_search_for", "").strip()
        sa_remark     = request.args.get("sa_remark", "").strip()
        sa_alloted_user = request.args.get("sa_alloted_user", "").strip()
        sa_date_from  = request.args.get("sa_date_from", "").strip()
        sa_date_to    = request.args.get("sa_date_to", "").strip()
        if sa_search:
            lt = f"%{sa_search}%"
            query = query.or_(
                f"final_url.ilike.{lt},"
                f"login_id.ilike.{lt},password.ilike.{lt},remark.ilike.{lt},"
                f"category.ilike.{lt},search_for.ilike.{lt},alloted_user_name.ilike.{lt}"
            )
        if sa_category:
            query = query.eq("category", sa_category)
        if sa_search_for:
            query = query.eq("search_for", sa_search_for)
        if sa_remark == "PENDING":
            query = query.or_("remark.is.null,remark.eq.,remark.eq.NA")
        elif sa_remark:
            query = query.eq("remark", sa_remark)
        if sa_alloted_user:
            query = query.eq("alloted_user_name", sa_alloted_user)
        if sa_date_from:
            query = query.gte("allotted_at", sa_date_from + " 00:00:00")
        if sa_date_to:
            query = query.lte("allotted_at", sa_date_to + " 23:59:59")

        resp = query.order("id", desc=True).execute()
        rows = resp.data or []

        df = pd.DataFrame(rows)
        cols = ["alloted_user_name", "final_url", "login_id",
                "password", "remark", "search_for", "category", "allotted_at"]
        cols = [c for c in cols if c in df.columns]
        df = df[cols] if cols else df

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Website Allotment")
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"website_allotment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        flash(f"Error exporting website allotment: {e}", "error")
        return redirect("/scam-website-allotment")

@app.route("/social-search-ajax", methods=["GET"])
@login_required
def social_search_ajax():
    try:
        query = request.args.get("q", "").strip()
        platform = request.args.get("platform", "").strip()
        status = request.args.get("status", "").strip()
        department = request.args.get("department", "").strip()
        permanent_block = request.args.get("permanent_block", "").strip()
        q = social_supabase.table("social_media_accounts").select("*", count='exact')
        allowed_depts = session.get("allowed_departments")
        if allowed_depts:
            if len(allowed_depts) == 1:
                q = q.eq("department", allowed_depts[0])
            else:
                q = q.in_("department", allowed_depts)
        if query:
            like_term = f"%{query}%"
            q = q.or_(
                f"login_user.ilike.{like_term},"
                f"number.ilike.{like_term},"
                f"full_name.ilike.{like_term},"
                f"page_name.ilike.{like_term},"
                f"platform.ilike.{like_term},"
                f"account_status.ilike.{like_term},"
                f"owned_by.ilike.{like_term},"
                f"department.ilike.{like_term},"
                f"sim_operator.ilike.{like_term},"
                f"mail_id.ilike.{like_term},"
                f"account_id.ilike.{like_term},"
                f"number_type.ilike.{like_term},"
                f"login_device.ilike.{like_term},"
                f"sim_inserted_device.ilike.{like_term}"
            )

        if platform:
            q = q.eq("platform", platform)
        if department:
            q = q.eq("department", department)

        if permanent_block == "true":
            q = q.eq("account_status", "Permanent Block")
        else:
            if status:
                q = q.eq("account_status", status)
            else:
                q = q.neq("account_status", "Permanent Block")
        q = q.order("id", desc=False).range(0, 999)
        resp = q.execute()
        items = [dict(row) for row in (resp.data or [])]
        total = resp.count or len(items)
        return jsonify({"success": True, "items": items, "total": total})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ============================================================
# TOTAL NUMBERS — Public Read-Only API
# ============================================================
TN_COLUMNS = "id,department,owned_by,number,sim_inserted_device,account_status,number_type,sim_operator,recharge_date"

# ============================================================
# CASE REPORT GENERATOR ROUTES
# ============================================================

CASE_REPORT_UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
CASE_REPORT_REPORTS_FOLDER = os.path.join(os.path.dirname(__file__), "generated_reports")
os.makedirs(CASE_REPORT_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CASE_REPORT_REPORTS_FOLDER, exist_ok=True)
CASE_REPORT_ALLOWED_EXT = {"png", "jpg", "jpeg", "webp"}
CASE_REPORT_SCREENSHOT_SEARCH_DIRS = [
    CASE_REPORT_UPLOAD_FOLDER,
    os.environ.get("CASE_REPORT_SCREENSHOT_DIR", "").strip(),
    os.path.join(os.path.dirname(os.path.dirname(__file__)), "IS_Websites_Automate", "api_gateway_screenshot"),
    r"C:\Users\Acer\OneDrive - Pixeltruth\Code_Hub\IS_Websites_Automate\api_gateway_screenshot",
]

# ============================================================
# AML GUI — BULK REGENERATE CASES (headless Selenium + OCR captcha)
# ============================================================
def configure_tesseract_cmd():
    """Configure Tesseract portably.

    Priority:
    1. TESSERACT_CMD env var
    2. tesseract available in system PATH
    3. Common Windows install paths
    4. Common Linux path for Render/Docker
    """
    candidates = []

    env_cmd = os.environ.get("TESSERACT_CMD")
    if env_cmd:
        candidates.append(env_cmd)

    path_cmd = shutil.which("tesseract")
    if path_cmd:
        candidates.append(path_cmd)

    candidates.extend([
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
    ])

    for cmd in candidates:
        if cmd and os.path.exists(cmd):
            pytesseract.pytesseract.tesseract_cmd = cmd
            return cmd

    # Last fallback: let pytesseract try PATH and produce a clear error if missing.
    pytesseract.pytesseract.tesseract_cmd = "tesseract"
    return "tesseract"

TESSERACT_CMD_RESOLVED = configure_tesseract_cmd()
print(f"[TESSERACT] command={TESSERACT_CMD_RESOLVED}", flush=True)

AML_LOGIN_URL  = "https://aml-gui.chargebackzero.com/index.php"
AML_REPORT_URL = "https://aml-gui.chargebackzero.com/report_generation/index_mfilter.php"
AML_USERNAME = os.environ.get("AML_USERNAME", "").strip()
AML_PASSWORD = os.environ.get("AML_PASSWORD", "").strip()
def get_aml_credentials_for_user(email=None, display_name=None):
    def _pick(row, *keys):
        for key in keys:
            val = row.get(key)
            if val is not None and str(val).strip():
                return str(val).strip()
        return None

    current_email = (email or "").strip().lower()
    current_name = get_clean_display_name(display_name or "").strip().lower()
    lookup_keys = [k for k in [current_email, current_name] if k]

    try:
        client = get_auth_supabase()
        user_row = None
        for key in lookup_keys:
            res = client.table("dashboard_users").select("*").eq("email", key).eq("is_active", True).limit(1).execute()
            if res.data:
                user_row = res.data[0]
                break

        if user_row:
            aml_blob = user_row.get("aml_credentials")
            if isinstance(aml_blob, str):
                try:
                    aml_blob = json.loads(aml_blob)
                except Exception:
                    aml_blob = {}
            if not isinstance(aml_blob, dict):
                aml_blob = {}

            username = _pick(user_row, "aml_username", "aml_user", "aml_email", "aml_login_username", "username", "user")
            password = _pick(user_row, "aml_password", "aml_pass", "aml_login_password", "password", "pass")
            customer = _pick(user_row, "aml_customer", "aml_input_customer", "aml_report_customer", "customer")
            platform = _pick(user_row, "aml_platform", "aml_input_platform", "aml_report_platform", "platform")

            username = username or _pick(aml_blob, "username", "user", "email")
            password = password or _pick(aml_blob, "password", "pass")
            customer = customer or _pick(aml_blob, "customer")
            platform = platform or _pick(aml_blob, "platform")

            if username and password:
                return (
                    username,
                    password,
                    customer or os.environ.get("AML_INPUT_CUSTOMER", "Mystery Shopping"),
                    platform or os.environ.get("AML_INPUT_PLATFORM", "v3_pre_scraper_merchantlaundering_data_table"),
                )
    except Exception as e:
        print(f"[AML CREDS] dashboard_users lookup error: {e}")

    if AML_USERNAME and AML_PASSWORD:
        return (
            AML_USERNAME,
            AML_PASSWORD,
            os.environ.get("AML_INPUT_CUSTOMER", "Mystery Shopping"),
            os.environ.get("AML_INPUT_PLATFORM", "v3_pre_scraper_merchantlaundering_data_table"),
        )
    raise RuntimeError("AML credentials not configured for this dashboard user")

def get_aml_credentials():
    return get_aml_credentials_for_user(session.get("email"), session.get("display_name"))

# TODO: DevTools se AML login page pe jaakar in XPaths ko verify/update karo
AML_USERNAME_INPUT_XPATH = "/html/body/div/div/div/div/div/form/div[1]/input"
AML_PASSWORD_INPUT_XPATH = "/html/body/div/div/div/div/div/form/div[2]/input"
AML_CAPTCHA_IMG_XPATH    = "//img[contains(@id,'captcha') or contains(@src,'captcha')]"
AML_CAPTCHA_INPUT_XPATH  = "//input[contains(@name,'captcha') or contains(@id,'captcha')]"
AML_LOGIN_BUTTON_XPATH   = "/html/body/div/div/div/div/div/form/button"

AML_TITLE1_XPATH       = "/html/body/div[1]/div/div/div/div/div/form/div[1]/div[1]/div/input[1]"
AML_INPUT1_XPATH        = "/html/body/div[1]/div/div/div/div/div/form/div[1]/div[1]/div/input[2]"
AML_DESCRIPTION1_XPATH  = "/html/body/div[1]/div/div/div/div/div/form/div[1]/div[3]/textarea"
AML_ADD_MORE_BTN_XPATH   = "/html/body/div[1]/div/div/div/div/div/form/div[2]/div[2]/button"
AML_TITLE2_XPATH        = "/html/body/div[1]/div/div/div/div/div/form/div[1]/div[4]/div/input"
AML_DESCRIPTION2_XPATH  = "/html/body/div[1]/div/div/div/div/div/form/div[1]/div[6]/textarea"
AML_GENERATE_BTN_XPATH   = "/html/body/div[1]/div/div/div/div/div/form/div[2]/div[1]/button"
AML_RESULT_LINK_XPATH      = "/html/body/a[2]"
AML_RESULT_LINKS_ALL_XPATH = "//a[contains(@href,'.pdf')]"
MAX_BULK_REGENERATE_REPORTS = 70
MAX_IMAGES_PER_REGENERATED_PDF = 10

def case_report_allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in CASE_REPORT_ALLOWED_EXT

def case_report_clean_up(*paths):
    for path in paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except OSError as exc:
            print(f"[CASE_REPORT] Could not remove temp file {path}: {exc}")


def resolve_bulk_screenshot_path(raw_path):
    raw = str(raw_path or "").strip().strip('"').strip("'")
    if not raw:
        return ""
    expanded = os.path.expandvars(os.path.expanduser(raw))
    if os.path.isfile(expanded):
        return expanded

    filename = re.split(r"[\\/]+", raw)[-1].strip()
    if not filename:
        return expanded

    for base_dir in CASE_REPORT_SCREENSHOT_SEARCH_DIRS:
        if not base_dir:
            continue
        candidate = os.path.join(base_dir, filename)
        if os.path.isfile(candidate):
            return candidate
    return expanded

# ============================================================
# AML GUI BULK REGENERATE — HELPERS
# ============================================================
def extract_screenshots_from_pdf(pdf_path, output_dir, prefix=None):
    saved_paths = []
    doc = fitz.open(pdf_path)
    counter = 1
    for page_index in range(len(doc)):
        page = doc[page_index]
        images = page.get_images(full=True)
        if not images:
            continue

        placement_entries = []  # (y0, x0, xref) — ek entry per placement
        for img in images:
            xref = img[0]
            try:
                rects = page.get_image_rects(xref)
            except Exception:
                rects = []
            if not rects:
                placement_entries.append((0, 0, xref))
                continue
            for rect in rects:
                placement_entries.append((rect.y0, rect.x0, xref))

        placement_entries.sort(key=lambda e: (e[0], e[1]))

        for _, _, xref in placement_entries:
            base_img = doc.extract_image(xref)
            img_bytes = base_img.get("image")
            filename = f"{prefix}_{counter}.png" if prefix else f"{counter}.png"
            out_path = os.path.join(output_dir, filename)
            try:
                with Image.open(io.BytesIO(img_bytes)) as img:
                    if img.mode not in ("RGB", "RGBA"):
                        img = img.convert("RGB")
                    img.save(out_path, format="PNG")
            except Exception:
                with open(out_path, "wb") as f:
                    f.write(img_bytes)
            saved_paths.append(out_path)
            counter += 1
    doc.close()
    return saved_paths

def download_and_extract_report_images(report, output_dir, prefix=None):
    pdf_url = report.get("pdf_url")
    temp_pdf_path = os.path.join(tempfile.gettempdir(), f"report_{report.get('id')}_{int(time.time())}.pdf")

    resp = requests.get(pdf_url, timeout=30, stream=True)
    resp.raise_for_status()
    with open(temp_pdf_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=1024 * 256):
            if chunk:
                f.write(chunk)

    image_paths = extract_screenshots_from_pdf(temp_pdf_path, output_dir, prefix=prefix)
    case_report_clean_up(temp_pdf_path)
    return image_paths
_DDDDOCR_ENGINE = None
_DDDDOCR_LOCK = threading.Lock()


def aml_clean_captcha_text(text):
    return re.sub(r"[^a-zA-Z0-9]", "", text or "").lower()


def aml_get_hex_captcha_text_with_easyocr(captcha_bytes):
    global _EASYOCR_READER
    with _EASYOCR_LOCK:
        if _EASYOCR_READER is None:
            import easyocr
            _EASYOCR_READER = easyocr.Reader(["en"], gpu=False, verbose=False)
    text = "".join(_EASYOCR_READER.readtext(captcha_bytes, detail=0, allowlist="0123456789abcdef"))
    return re.sub(r"[^0-9a-f]", "", text.lower())


def aml_solve_gui_download_captcha(session_obj):
    img = session_obj.get(urllib.parse.urljoin(AML_GUI_BASE_URL, "captcha.php"), timeout=30).content
    return aml_get_captcha_text_from_bytes(img)


def aml_get_captcha_text_with_tesseract(captcha_bytes):
    img = Image.open(io.BytesIO(captcha_bytes)).convert("L")
    if img.width and img.height:
        img = img.resize((img.width * 5, img.height * 5))
    text = pytesseract.image_to_string(
        img,
        config="--psm 7 -c tessedit_char_whitelist=abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    ).strip()
    return aml_clean_captcha_text(text)


def aml_get_captcha_text_with_ddddocr(captcha_bytes):
    """Tesseract-free captcha OCR fallback using the pure Python package ddddocr.

    ddddocr is loaded lazily so normal app startup does not fail if the package is
    absent. Add it to requirements.txt on Render Python runtime.
    """
    global _DDDDOCR_ENGINE
    with _DDDDOCR_LOCK:
        if _DDDDOCR_ENGINE is None:
            import ddddocr
            _DDDDOCR_ENGINE = ddddocr.DdddOcr(show_ad=False)
    text = _DDDDOCR_ENGINE.classification(captcha_bytes)
    return aml_clean_captcha_text(text)


def aml_get_captcha_text_from_bytes(captcha_bytes):
    """Captcha OCR helper for direct HTTP captcha."""
    engine = os.environ.get("AML_OCR_ENGINE", "auto").strip().lower()
    tesseract_available = bool(shutil.which("tesseract")) or (
        TESSERACT_CMD_RESOLVED not in ("tesseract", None) and os.path.exists(TESSERACT_CMD_RESOLVED)
    )

    engines = []
    if engine == "easyocr":
        engines = ["easyocr"]
    elif engine == "tesseract":
        engines = ["tesseract"]
    elif engine in ("ddddocr", "dddocr"):
        engines = ["ddddocr"]
    elif tesseract_available:
        engines = ["tesseract"]
    else:
        engines = []

    if not engines:
        raise RuntimeError("Tesseract OCR is not available. Set AML_OCR_ENGINE=ddddocr or easyocr only if Render has enough memory.")

    last_error = None
    for candidate in engines:
        try:
            if candidate == "easyocr":
                captcha_text = aml_get_hex_captcha_text_with_easyocr(captcha_bytes)
            elif candidate == "tesseract":
                captcha_text = aml_get_captcha_text_with_tesseract(captcha_bytes)
            else:
                captcha_text = aml_get_captcha_text_with_ddddocr(captcha_bytes)
            if captcha_text:
                print(f"[AML OCR] engine={candidate} text_len={len(captcha_text)}", flush=True)
                return captcha_text
        except Exception as exc:
            last_error = exc
            print(f"[AML OCR] engine={candidate} failed: {exc}", flush=True)

    if last_error:
        raise last_error
    return ""


def aml_get_captcha_text(captcha_element):
    # Backward-compatible helper for the old Selenium path.
    return aml_get_captcha_text_from_bytes(captcha_element.screenshot_as_png)


def aml_extract_pdf_links_from_html(html_text, base_url):
    """Extract generated PDF links from AML response HTML/JSON/JS/plain text.

    AML's upload endpoint may return links as anchors, plain URLs, escaped JSON
    strings, or JavaScript snippets. Keep extraction broad but still limited to
    PDF-like URLs/paths.
    """
    raw = html_text or ""
    candidates = []

    def _add(value):
        if not value:
            return
        value = value.strip().strip('"\'`<> )(')
        if not value:
            return
        value = value.replace("\\/", "/")
        value = value.replace("&amp;", "&")
        value = urllib.parse.unquote(value)
        # Trim common trailing punctuation that can appear in JS/JSON snippets.
        value = re.sub(r"[;,'\")\]}]+$", "", value)
        if ".pdf" not in value.lower():
            return
        # If a broad relative regex catches a domain-style value without scheme,
        # normalize it as absolute instead of urljoining it under /report_generation/.
        if not re.match(r"^[a-z][a-z0-9+.-]*://", value, flags=re.IGNORECASE):
            if re.match(r"^[A-Za-z0-9.-]+\.[A-Za-z]{2,}/", value):
                value = "https://" + value
        basename = value.rsplit("/", 1)[-1]
        if "/" not in value and any(c.rsplit("/", 1)[-1] == basename for c in candidates):
            return
        candidates.append(value)

    # 1) Standard anchors / attributes.
    for match in re.findall(r"(?:href|src|data-url|url)\s*=\s*[\"']([^\"']+\.pdf[^\"']*)[\"']", raw, flags=re.IGNORECASE):
        _add(match)

    # 2) Absolute URLs anywhere in the response, including JSON/JS strings.
    for match in re.findall(r"https?:\\?/\\?/[^\s\"'<>]+?\.pdf[^\s\"'<>]*", raw, flags=re.IGNORECASE):
        _add(match)

    # 3) Relative PDF paths anywhere in the response.
    for match in re.findall(r"(?:\.\./|\./|/)?[A-Za-z0-9_./%+-]+\.pdf[^\s\"'<>]*", raw, flags=re.IGNORECASE):
        _add(match)

    hrefs = []
    seen = set()
    seen_basenames = set()
    for candidate in candidates:
        absolute = urllib.parse.urljoin(base_url, candidate)
        basename = absolute.rsplit("/", 1)[-1].lower()
        if absolute not in seen and basename not in seen_basenames:
            seen.add(absolute)
            seen_basenames.add(basename)
            hrefs.append(absolute)

    def _link_sort_key(href):
        h = href.lower()
        if "mfilterit" in h:
            return 0
        if "npci" in h:
            return 1
        if "without_header" in h:
            return 2
        return 3

    hrefs.sort(key=_link_sort_key)
    return hrefs


def aml_login_requests(max_captcha_attempts=7, creds=None, require_report_form=True):
    """Login to AML GUI without launching Chrome. Returns authenticated requests.Session."""
    session_obj = requests.Session()
    session_obj.headers.update({
        "User-Agent": AML_GUI_UA,
        "Referer": AML_LOGIN_URL,
    })

    for attempt in range(1, max_captcha_attempts + 1):
        username_value, password_value, customer_value, platform_value = creds or get_aml_credentials()
        session_obj.cookies.clear()
        print(f"[AML HTTP LOGIN] Loading login page attempt={attempt}", flush=True)
        login_page = session_obj.get(AML_GUI_BASE_URL, timeout=30)
        login_page.raise_for_status()
        print(f"[AML HTTP LOGIN] Solving captcha attempt={attempt}", flush=True)
        captcha_text = aml_solve_gui_download_captcha(session_obj)
        print(f"[AML HTTP LOGIN] Captcha attempt {attempt}: '{captcha_text}'")

        if len(captcha_text or "") != 6:
            time.sleep(1)
            continue

        payload = {
            "usernamee": username_value,
            "passwordd": password_value,
            "inputcustomer": customer_value,
            "inputplatform": platform_value,
            "captcha": captcha_text,
            "sub": "",
        }
        login_resp = session_obj.post(
            AML_GUI_BASE_URL,
            data=payload,
            headers={"Referer": AML_GUI_BASE_URL, "Origin": AML_GUI_BASE_URL.rstrip("/")},
            timeout=60,
            allow_redirects=False,
        )
        login_resp.raise_for_status()

        location = login_resp.headers.get("Location", "")
        login_ok = login_resp.status_code == 302 and "datalist_npci" in location
        if login_ok and not require_report_form:
            print("[AML HTTP LOGIN] Success")
            return session_obj

        if login_ok:
            report_page = session_obj.get(AML_REPORT_URL, timeout=30, allow_redirects=True)
            report_page.raise_for_status()
            if re.search(r"name=[\"']left_image\[\][\"']", report_page.text, flags=re.IGNORECASE):
                print("[AML HTTP LOGIN] Success")
                return session_obj
            print(f"[AML HTTP LOGIN] Attempt {attempt} failed - report form not available. Retrying.")
        else:
            print(f"[AML HTTP LOGIN] Attempt {attempt} failed - status={login_resp.status_code} location={location!r}. Retrying.")
        time.sleep(1)

    raise RuntimeError("AML GUI HTTP login failed after captcha retries")
def aml_build_driver(headless=False):
    options = ChromeOptions()
    # Default headed for GUI-style AML flows; headless can still be requested.
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--remote-debugging-port=0")
    # Unique profile dir per run — avoids "user data directory already in use" crashes
    unique_profile = os.path.join(tempfile.gettempdir(), f"chrome_profile_{uuid.uuid4().hex}")
    options.add_argument(f"--user-data-dir={unique_profile}")

    try:
        # Auto-downloads/matches the chromedriver version to the installed Chrome browser
        service = ChromeService(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=options)
    except Exception as exc:
        print(f"[AML DRIVER] webdriver-manager failed, falling back to system chromedriver: {exc}")
        return webdriver.Chrome(options=options)


def aml_login(driver, max_captcha_attempts=5):
    wait = WebDriverWait(driver, 20)
    driver.get(AML_LOGIN_URL)

    for attempt in range(1, max_captcha_attempts + 1):
        try:
            username_value, password_value, _, _ = get_aml_credentials()
            username_field = wait.until(EC.presence_of_element_located((By.XPATH, AML_USERNAME_INPUT_XPATH)))
            username_field.clear()
            username_field.send_keys(username_value)

            password_field = driver.find_element(By.XPATH, AML_PASSWORD_INPUT_XPATH)
            password_field.clear()
            password_field.send_keys(password_value)

            captcha_img = wait.until(EC.presence_of_element_located((By.XPATH, AML_CAPTCHA_IMG_XPATH)))
            captcha_text = aml_get_captcha_text(captcha_img)
            print(f"[AML LOGIN] Captcha attempt {attempt}: '{captcha_text}'")

            if not captcha_text:
                print("[AML LOGIN] Captcha OCR returned empty text — retrying with a fresh page.")
                driver.get(AML_LOGIN_URL)
                time.sleep(0.5)
                continue

            captcha_field = driver.find_element(By.XPATH, AML_CAPTCHA_INPUT_XPATH)
            captcha_field.clear()
            captcha_field.send_keys(captcha_text)

            login_btn = driver.find_element(By.XPATH, AML_LOGIN_BUTTON_XPATH)
            login_btn.click()
            time.sleep(3)

            # Login form (username field) abhi bhi page pe hai ya nahi — yehi
            # asli success/fail signal hai. URL-based check bharosemand nahi tha
            # (login page ka URL khud hi "index.php" hai, "login" word nahi hai,
            # isliye purana check galat-positive "Success" de raha tha).
            try:
                driver.find_element(By.XPATH, AML_USERNAME_INPUT_XPATH)
                login_form_still_present = True
            except NoSuchElementException:
                login_form_still_present = False

            if not login_form_still_present:
                print("[AML LOGIN] Success:", driver.current_url)
                return True

            print(f"[AML LOGIN] Attempt {attempt} failed — login form still visible (wrong captcha?). Retrying.")
            driver.get(AML_LOGIN_URL)
            time.sleep(1)

        except TimeoutException:
            print(f"[AML LOGIN] Attempt {attempt} timed out waiting for login elements — retrying.")
            driver.get(AML_LOGIN_URL)
            time.sleep(1)
            continue

    print("[AML LOGIN] Failed after retries.")
    return False
def aml_submit_chunk(driver, title_text, input_text, description_text, image_chunk):
    """image_chunk: list of 1-4 image paths (sequence maintained). Returns extracted new PDF URL/text."""
    wait = WebDriverWait(driver, 20)
    if len(image_chunk) < 1:
        return None

    try:
        _ = driver.current_url  # session health check — crashes here if browser already dead
    except Exception as exc:
        raise RuntimeError(f"Chrome session is dead before submission: {exc}")

    driver.get(AML_REPORT_URL)

    try:
        left_img = wait.until(EC.presence_of_element_located((By.NAME, "left_image[]")))
    except TimeoutException:
        debug_id = uuid.uuid4().hex[:8]
        debug_dir = os.path.join(tempfile.gettempdir(), "aml_debug")
        os.makedirs(debug_dir, exist_ok=True)
        screenshot_path = os.path.join(debug_dir, f"debug_leftimg_{debug_id}.png")
        html_path = os.path.join(debug_dir, f"debug_leftimg_{debug_id}.html")
        try:
            driver.save_screenshot(screenshot_path)
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
        except Exception as dbg_exc:
            print(f"[AML SUBMIT] Could not save debug artifacts: {dbg_exc}")
        raise RuntimeError(
            f"'left_image[]' field not found on report generation page "
            f"(current URL: {driver.current_url}). Possible causes: AML GUI layout changed, "
            f"captcha not solved, or session expired. "
            f"screenshot/HTML id={debug_id} saved in {debug_dir}"
        )
    left_img.send_keys(image_chunk[0])

    if len(image_chunk) >= 2:
        right_img = driver.find_element(By.NAME, "right_image[]")
        right_img.send_keys(image_chunk[1])

    title1 = driver.find_element(By.XPATH, AML_TITLE1_XPATH)
    title1.clear()
    title1.send_keys(title_text)

    input_field = driver.find_element(By.XPATH, AML_INPUT1_XPATH)
    input_field.clear()
    input_field.send_keys(input_text)

    description1 = driver.find_element(By.XPATH, AML_DESCRIPTION1_XPATH)
    description1.clear()
    description1.send_keys(description_text)

    if len(image_chunk) > 2:
        add_btn = wait.until(EC.element_to_be_clickable((By.XPATH, AML_ADD_MORE_BTN_XPATH)))
        driver.execute_script("arguments[0].click();", add_btn)
        time.sleep(2)

        title2 = wait.until(EC.presence_of_element_located((By.XPATH, AML_TITLE2_XPATH)))
        title2.clear()
        title2.send_keys(title_text)

        description2 = driver.find_element(By.XPATH, AML_DESCRIPTION2_XPATH)
        description2.clear()
        description2.send_keys(description_text)

        left_img2 = driver.find_elements(By.NAME, "left_image[]")[1]
        left_img2.send_keys(image_chunk[2])

        if len(image_chunk) == 4:
            right_img2 = driver.find_elements(By.NAME, "right_image[]")[1]
            right_img2.send_keys(image_chunk[3])

    generate_btn = wait.until(EC.element_to_be_clickable((By.XPATH, AML_GENERATE_BTN_XPATH)))
    generate_btn.click()
    long_wait = WebDriverWait(driver, 40)  # PDF generation can take longer than 20s
    npci_href = None
    all_links = None

    # Generate ke baad portal teen links deta hai (mfilterit / npci / without_header
    # style variants — generate_screenshot_urls() ki tarah). "Regenerate Cases"
    # (basic) results sheet mein sirf NPCI link chahiye, lekin "Regenerate with
    # Final Sheet" wali Investment Scam sheet mein teeno links (comma-separated)
    # chahiye — isliye dono yahin collect kar rahe hain.
    try:
        long_wait.until(EC.presence_of_element_located((By.XPATH, AML_RESULT_LINKS_ALL_XPATH)))
        anchors = driver.find_elements(By.XPATH, AML_RESULT_LINKS_ALL_XPATH)
        seen = set()
        hrefs = []
        for a in anchors:
            href = (a.get_attribute("href") or "").strip()
            if href and href not in seen:
                seen.add(href)
                hrefs.append(href)

        # Investment Scam Final Sheet ke liye fixed sequence chahiye:
        # mfilterit -> npci -> without_header
        def _link_sort_key(href):
            h = href.lower()
            if "mfilterit" in h:
                return 0
            if "npci" in h:
                return 1
            if "without_header" in h:
                return 2
            return 3
        hrefs.sort(key=_link_sort_key)

        for href in hrefs:
            if "npci" in href.lower():
                npci_href = href
                break
        if hrefs:
            all_links = ",".join(hrefs)
    except TimeoutException:
        pass  # try fallbacks below before giving up

    result_text = npci_href or (all_links.split(",")[0] if all_links else None)

    # ── Fallback: old single-anchor XPath, agar link markup badal jaaye ──
    if not result_text:
        try:
            result_text = driver.find_element(By.XPATH, AML_RESULT_LINK_XPATH).text
            all_links = all_links or result_text
        except Exception:
            pass

    # ── Fallback 2: save screenshot + HTML so we can see what actually rendered ──
    if not result_text:
        debug_id = uuid.uuid4().hex[:8]
        debug_dir = os.path.join(tempfile.gettempdir(), "aml_debug")
        os.makedirs(debug_dir, exist_ok=True)
        screenshot_path = os.path.join(debug_dir, f"debug_{debug_id}.png")
        html_path = os.path.join(debug_dir, f"debug_{debug_id}.html")
        try:
            driver.save_screenshot(screenshot_path)
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(f"[AML SUBMIT] Result link not found. Debug saved:\n  {screenshot_path}\n  {html_path}\n  Current URL: {driver.current_url}")
        except Exception as dbg_exc:
            print(f"[AML SUBMIT] Could not save debug artifacts: {dbg_exc}")
        raise RuntimeError(
            f"Result link not found after Generate click (current URL: {driver.current_url}). "
            f"Debug screenshot/HTML saved with id={debug_id} in {debug_dir}"
        )

    return {"npci": result_text, "all": all_links or result_text}


def aml_submit_chunk_requests(session_obj, title_text, input_text, description_text, image_chunk):
    """Submit one AML regenerate chunk without Selenium/Chromium.

    image_chunk: list of 1-4 image paths. Returns the same shape as aml_submit_chunk():
    {"npci": <npci-link>, "all": <comma-separated-links>}.
    """
    if len(image_chunk) < 1:
        return None

    report_page = session_obj.get(AML_REPORT_URL, timeout=30, allow_redirects=True)
    report_page.raise_for_status()
    if not re.search(r"name=[\"']left_image\[\][\"']", report_page.text, flags=re.IGNORECASE):
        raise RuntimeError("AML report form not available. Session may have expired or login failed.")

    upload_url = urllib.parse.urljoin(AML_REPORT_URL, "ajaxUpload_mfilter.php")
    data = [
        ("title[]", title_text),
        ("input", input_text),
        ("description[]", description_text),
        ("generate_pdf", ""),
    ]
    files = []
    opened_files = []
    try:
        upload_debug = []

        def _attach(field_name, file_path):
            # Use the same extracted PNG path/file for regenerate upload.
            # The extractor above guarantees .png filenames.
            fh = open(file_path, "rb")
            opened_files.append(fh)
            filename = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            try:
                with Image.open(file_path) as img:
                    image_format = img.format
                    image_size = f"{img.width}x{img.height}"
            except Exception:
                image_format = "unknown"
                image_size = "unknown"
            upload_debug.append(f"{field_name}={filename}:{file_size}B:{image_format}:{image_size}")
            files.append((field_name, (filename, fh, "image/png")))

        def _attach_empty(field_name):
            # Browser submits an empty multipart part for file inputs left blank.
            # This matters for 1/3-image cases where right_image[] exists in the form
            # but the user has not selected a file.
            upload_debug.append(f"{field_name}=EMPTY")
            files.append((field_name, ("", io.BytesIO(b""), "application/octet-stream")))

        # AML form has two image columns per row: left_image[] + right_image[].
        # The portal's Add More UI creates more rows. In direct HTTP we mimic that
        # by repeating title[]/description[] and file fields for every image pair.
        for row_index in range(0, len(image_chunk), 2):
            if row_index > 0:
                data.append(("title[]", title_text))
                data.append(("description[]", description_text))

            _attach("left_image[]", image_chunk[row_index])
            if row_index + 1 < len(image_chunk):
                _attach("right_image[]", image_chunk[row_index + 1])
            else:
                _attach_empty("right_image[]")

        print(f"[AML HTTP SUBMIT] Uploading files: {'; '.join(upload_debug)}", flush=True)
        resp = session_obj.post(
            upload_url,
            data=data,
            files=files,
            timeout=90,
            allow_redirects=True,
            headers={"Referer": AML_REPORT_URL},
        )
        resp.raise_for_status()
        hrefs = aml_extract_pdf_links_from_html(resp.text, upload_url)
        if not hrefs:
            debug_id = uuid.uuid4().hex[:8]
            debug_dir = os.path.join(tempfile.gettempdir(), "aml_debug")
            os.makedirs(debug_dir, exist_ok=True)
            html_path = os.path.join(debug_dir, f"http_debug_{debug_id}.html")
            try:
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(resp.text)
            except Exception as dbg_exc:
                print(f"[AML HTTP SUBMIT] Could not save debug HTML: {dbg_exc}")
            response_tail = (resp.text or "")[-1200:].replace("\n", " ").replace("\r", " ")
            print(
                f"[AML HTTP SUBMIT] No PDF link parsed. status={resp.status_code} "
                f"url={resp.url} content_type={resp.headers.get('content-type')} "
                f"debug_id={debug_id} response_tail={response_tail}",
                flush=True,
            )
            raise RuntimeError(f"Result PDF link not found in AML HTTP response. Debug HTML id={debug_id} saved in {debug_dir}")

        npci_href = None
        for href in hrefs:
            if "npci" in href.lower():
                npci_href = href
                break
        all_links = ",".join(hrefs)
        return {"npci": npci_href or hrefs[0], "all": all_links}
    finally:
        for fh in opened_files:
            try:
                fh.close()
            except Exception:
                pass
def aml_strip_scheme(url):
    """Scheme (https://, http://) hata ke sirf domain/handle deta hai —
    Input field mein koi bhi '/' (path ya trailing slash) na jaaye."""
    if not url:
        return url
    stripped = url.replace("https://", "").replace("http://", "").strip()
    # Agar path/query hai (domain ke baad koi '/'), to sirf domain rakho
    if "/" in stripped:
        stripped = stripped.split("/", 1)[0]
    return stripped.rstrip("/")
def aml_chunk_list(lst, size):
    for i in range(0, len(lst), size):
        yield lst[i:i + size]

# ============================================================
# BULK REGENERATE — background job tracking (Stop button support)
# ============================================================
REGEN_JOBS = {}
REGEN_JOBS_LOCK = threading.Lock()

def _regen_job_response(job):
    return {
        "status": job["status"],
        "phase": job.get("phase"),
        "total": job["total"],
        "completed": job["completed"],
        "results": job["results"],
        "folder": job["folder"],
        "final_excel": job["final_excel"],
        "message": job["message"],
    }


def _write_regen_job_snapshot(job_id):
    with REGEN_JOBS_LOCK:
        job = REGEN_JOBS.get(job_id)
        if not job or not job.get("folder"):
            return
        snapshot = _regen_job_response(job)

    try:
        folder = Path(snapshot["folder"])
        folder.mkdir(parents=True, exist_ok=True)
        with open(folder / "job_status.json", "w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        print(f"[BULK REGENERATE] Could not save job snapshot {job_id}: {exc}", flush=True)


def _load_regen_job_snapshot(job_id):
    try:
        regen_jobs_base = Path(__file__).parent / "regenerate_jobs"
        candidates = list(regen_jobs_base.glob(f"{job_id}_*/job_status.json"))
        if not candidates:
            return None
        latest = max(candidates, key=lambda p: p.stat().st_mtime)
        with open(latest, "r", encoding="utf-8") as f:
            snapshot = json.load(f)
        if snapshot.get("status") in ("running", "stopping"):
            snapshot["status"] = "error"
            snapshot["message"] = snapshot.get("message") or "Server restarted while this job was running. Completed progress was restored from disk."
        return snapshot
    except Exception as exc:
        print(f"[BULK REGENERATE] Could not load job snapshot {job_id}: {exc}", flush=True)
        return None


def _cleanup_paths(paths):
    for path in paths or []:
        try:
            if path and os.path.isfile(path):
                os.remove(path)
        except Exception as exc:
            print(f"[CLEANUP] Could not remove {path}: {exc}", flush=True)


def _run_bulk_regenerate_job(job_id, report_ids, mode, input_user, aml_creds):
    """Background thread — actual regenerate ka kaam yahi karta hai,
    stop_event check karte hue taaki beech mein rok sakein."""
    with REGEN_JOBS_LOCK:
        job = REGEN_JOBS.get(job_id)
    if not job:
        return
    stop_event = job["stop_event"]

    try:
        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["phase"] = "extracting"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        regen_jobs_base = os.path.join(os.path.dirname(__file__), "regenerate_jobs")
        os.makedirs(regen_jobs_base, exist_ok=True)
        session_folder = os.path.join(regen_jobs_base, f"{job_id}_{timestamp}")
        images_folder  = os.path.join(session_folder, "Screenshots")
        os.makedirs(images_folder, exist_ok=True)
        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["folder"] = session_folder
        _write_regen_job_snapshot(job_id)

        final_rows = []
        final_excel_path = os.path.join(session_folder, f"final_regenerated_results_{timestamp}.xlsx")

        aml_session = aml_login_requests(creds=aml_creds)
        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["phase"] = "regenerating"
            REGEN_JOBS[job_id]["driver"] = None  # Requests-based flow does not launch Chrome.
        _write_regen_job_snapshot(job_id)

        for report_id in report_ids:
            if stop_event.is_set():
                with REGEN_JOBS_LOCK:
                    REGEN_JOBS[job_id]["results"].append({
                        "id": report_id, "old_pdf_url": None, "new_pdf_url": None,
                        "source_url": None, "status": "Stopped by user"
                    })
                    REGEN_JOBS[job_id]["completed"] += 1
                _write_regen_job_snapshot(job_id)
                break

            paths_for_report = []
            resp = supabase.table("reports").select("*").eq("id", report_id).execute()
            if not resp.data:
                row_result = {"id": report_id, "old_pdf_url": None, "new_pdf_url": None, "source_url": None, "status": "Report not found"}
                with REGEN_JOBS_LOCK:
                    REGEN_JOBS[job_id]["results"].append(row_result)
                    REGEN_JOBS[job_id]["completed"] += 1
                final_rows.append(row_result)
                _write_regen_job_snapshot(job_id)
                continue

            report = resp.data[0]
            try:
                try:
                    paths_for_report = download_and_extract_report_images(report, images_folder, prefix=str(report_id))
                except Exception as exc:
                    paths_for_report = []

                if not paths_for_report:
                    row_result = {"id": report.get("id"), "old_pdf_url": report.get("pdf_url"), "new_pdf_url": None, "source_url": report.get("source_url"), "status": "No screenshots extracted"}
                    with REGEN_JOBS_LOCK:
                        REGEN_JOBS[job_id]["results"].append(row_result)
                        REGEN_JOBS[job_id]["completed"] += 1
                    final_rows.append(row_result)
                    _write_regen_job_snapshot(job_id)
                    continue

                source_url = report.get("source_url") or "NA"
                title_text = source_url
                description_text = source_url
                input_text = aml_strip_scheme(source_url)
                new_pdf_url = None        # NPCI-only link — "Regenerate Cases" (basic) results sheet ke liye
                new_pdf_links_all = None  # Teeno links (comma-separated) — Investment Scam Final Sheet ke liye
                status = "Success"
                try:
                    # Submit up to MAX_IMAGES_PER_REGENERATED_PDF screenshots in one AML request
                    # so one old PDF maps to one new PDF whenever the old PDF has up to 10 images.
                    chunk_results = []
                    for chunk in aml_chunk_list(paths_for_report, MAX_IMAGES_PER_REGENERATED_PDF):
                        if stop_event.is_set():
                            status = "Stopped by user"
                            break
                        chunk_result = aml_submit_chunk_requests(aml_session, title_text, input_text, description_text, chunk)
                        if chunk_result:
                            chunk_results.append(chunk_result)
                    if status != "Stopped by user":
                        if chunk_results:
                            # Har chunk ek alag AML report submission hai (portal max 4/submission
                            # allow karta hai) — saare chunks ke links combine karke rakhte hain.
                            npci_links = [r.get("npci") for r in chunk_results if r.get("npci")]
                            all_links_list = [r.get("all") for r in chunk_results if r.get("all")]
                            new_pdf_url = ",".join(npci_links) if npci_links else "Failed"
                            new_pdf_links_all = ",".join(all_links_list) if all_links_list else new_pdf_url
                        else:
                            new_pdf_url = "Failed"
                            status = "No result returned"
                except Exception as exc:
                    tb_text = traceback.format_exc()
                    print(f"[BULK REGENERATE] report_id={report_id} failed:\n{tb_text}")
                    err_msg = str(exc).strip() or type(exc).__name__
                    status = f"Error: {err_msg}"
                    # Session/captcha can expire; re-login once so remaining reports can continue.
                    try:
                        aml_session = aml_login_requests(creds=aml_creds)
                    except Exception as login_exc:
                        status = f"{status} | HTTP re-login failed: {str(login_exc).strip() or type(login_exc).__name__}"

                row_result = {
                    "id": report.get("id"),
                    "old_pdf_url": report.get("pdf_url"),
                    "new_pdf_url": new_pdf_url,
                    "source_url": report.get("source_url"),
                    "status": status,
                }
                with REGEN_JOBS_LOCK:
                    REGEN_JOBS[job_id]["results"].append(row_result)
                    REGEN_JOBS[job_id]["completed"] += 1
                final_rows.append(row_result)
                _write_regen_job_snapshot(job_id)
            finally:
                _cleanup_paths(paths_for_report)

        FINAL_RESULT_COLUMNS = ["id", "old_pdf_url", "new_pdf_url", "source_url", "status"]
        pd.DataFrame(final_rows, columns=FINAL_RESULT_COLUMNS).to_excel(final_excel_path, index=False)
        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["final_excel"] = final_excel_path
        _write_regen_job_snapshot(job_id)

        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["status"] = "stopped" if stop_event.is_set() else "done"
            REGEN_JOBS[job_id]["driver"] = None
        _write_regen_job_snapshot(job_id)
    except Exception as exc:
        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id]["status"]  = "error"
            REGEN_JOBS[job_id]["message"] = str(exc)
            REGEN_JOBS[job_id]["driver"]  = None
        _write_regen_job_snapshot(job_id)

@app.route("/case-report", methods=["GET"])
@login_required
def case_report_page():
    allowed_pages = session.get("allowed_pages", [])
    if "case_report" not in allowed_pages:
        flash("You don't have access to Case Report Generator.", "error")
        return redirect_to_allowed_page(allowed_pages)
    clean_display_name = get_clean_display_name(session.get("display_name", "User"))
    return render_template(
        "case_report.html",
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        allowed_pages=allowed_pages,
        is_admin=session.get("is_admin", False),
    )

@app.route("/generate-case-report", methods=["POST"])
@login_required
def generate_case_report():
    source_url = request.form.get("source_url", "").strip()
    if not source_url:
        return jsonify({"status": "error", "message": "Website URL is required."}), 400

    transaction_type = request.form.get("transaction_type", "").strip().lower()
    if transaction_type not in ("upi", "bank"):
        return jsonify({"status": "error", "message": "Please select UPI or Bank Account."}), 400

    # ── Common flat columns — same set used for both UPI and Bank Account ──
    upi_vpa              = "NA"
    payment_gateway_url   = "NA"
    bank_account_number   = "NA"
    ifsc_code             = "NA"
    ac_holder_name        = "NA"
    search_for            = "NA"
    scam_type             = "NA"
    chat_number           = "NA"

    if transaction_type == "upi":
        upi_vpa             = request.form.get("upi_vpa", "").strip()
        payment_gateway_url = request.form.get("payment_gateway_url", "").strip()
        scam_type           = request.form.get("scam_type", "").strip()
        search_for           = request.form.get("search_for", "").strip() or "NA"
        chat_number          = request.form.get("chat_number", "").strip() or "NA"
        if not (upi_vpa and source_url and payment_gateway_url and scam_type):
            return jsonify({"status": "error", "message": "Please fill all required UPI fields."}), 400
    else:  # bank
        bank_account_number = request.form.get("bank_account_number", "").strip()
        ifsc_code            = request.form.get("ifsc_code", "").strip()
        ac_holder_name        = request.form.get("ac_holder_name", "").strip()
        search_for            = request.form.get("search_for", "").strip()
        scam_type             = request.form.get("scam_type", "").strip()
        chat_number           = request.form.get("chat_number", "").strip() or "NA"
        if not (bank_account_number and ifsc_code and ac_holder_name and source_url and search_for and scam_type):
            return jsonify({"status": "error", "message": "Please fill all required Bank Account fields."}), 400

    screenshot_files = request.files.getlist("screenshots[]")
    if not screenshot_files or all(f.filename == "" for f in screenshot_files):
        return jsonify({"status": "error", "message": "At least one screenshot is required."}), 400

    saved_image_paths = []
    for f in screenshot_files:
        if f.filename == "":
            continue
        if not case_report_allowed_file(f.filename):
            return jsonify({
                "status": "error",
                "message": f"'{f.filename}' is not a supported image type. Use PNG, JPG, JPEG, or WEBP."
            }), 400
        safe_name = secure_filename(f.filename)
        dest = os.path.join(CASE_REPORT_UPLOAD_FOLDER, safe_name)
        counter = 1
        base, ext = os.path.splitext(dest)
        while os.path.exists(dest):
            dest = f"{base}_{counter}{ext}"
            counter += 1
        f.save(dest)
        saved_image_paths.append(dest)

    if not saved_image_paths:
        return jsonify({"status": "error", "message": "No valid screenshots were uploaded."}), 400

    pdf_filename = generate_filename(source_url)
    local_pdf_path = os.path.join(CASE_REPORT_REPORTS_FOLDER, pdf_filename)

    try:
        total_pages = generate_pdf(
            source_url=source_url,
            image_paths=saved_image_paths,
            output_path=local_pdf_path,
        )
        pdf_url = upload_pdf(local_pdf_path)

        file_size = os.path.getsize(local_pdf_path)
        input_user = get_clean_display_name(session.get("display_name", "User"))

        supabase.table("reports").insert({
            "filename": pdf_filename,
            "source_url": source_url,
            "pdf_url": pdf_url,
            "total_pages": total_pages,
            "file_size": file_size,
            "upi_vpa": upi_vpa,
            "payment_gateway_url": payment_gateway_url,
            "bank_account_number": bank_account_number,
            "ifsc_code": ifsc_code,
            "ac_holder_name": ac_holder_name,
            "search_for": search_for,
            "scam_type": scam_type,
            "chat_number": chat_number,
            "input_user": input_user,
        }).execute()

        return jsonify({
            "status": "success",
            "pdf_url": pdf_url,
            "filename": pdf_filename,
            "pages": total_pages,
        })
    except Exception as exc:
        print(f"[CASE_REPORT] Generation failed: {exc}")
        return jsonify({"status": "error", "message": str(exc)}), 500
    finally:
        case_report_clean_up(*saved_image_paths, local_pdf_path)

@app.route("/case-reports-list", methods=["GET"])
@login_required
def case_reports_list():
    try:
        current_user = get_clean_display_name(session.get("display_name", "User"))
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to", "").strip()

        # ── Total count (all-time, unfiltered) for this user ──
        total_resp = supabase.table("reports") \
            .select("id", count="exact") \
            .eq("input_user", current_user) \
            .execute()
        total_count = total_resp.count or 0

        query = supabase.table("reports").select(
            "id,source_url,pdf_url,upi_vpa,payment_gateway_url,bank_account_number,"
            "ifsc_code,ac_holder_name,search_for,scam_type,chat_number,input_user,created_at",
            count="exact"
        ).eq("input_user", current_user)

        if date_from:
            query = query.gte("created_at", f"{date_from}T00:00:00")
        if date_to:
            try:
                next_day = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
                query = query.lt("created_at", f"{next_day}T00:00:00")
            except ValueError:
                query = query.lte("created_at", f"{date_to}T23:59:59")

        response = query.order("created_at", desc=True).execute()
        filtered_rows = response.data or []
        filtered_count = response.count if response.count is not None else len(filtered_rows)

        return jsonify({
            "status": "success",
            "reports": filtered_rows,
            "total_count": total_count,
            "filtered_count": filtered_count,
        })
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

@app.route("/delete-case-report/<report_id>", methods=["DELETE"])
@login_required
def delete_case_report(report_id):
    try:
        resp = supabase.table("reports").select("filename,input_user").eq("id", report_id).execute()
        if not resp.data:
            return jsonify({"status": "error", "message": "Report not found."}), 404
        record = resp.data[0]
        current_user = get_clean_display_name(session.get("display_name", "User"))
        # User apna hi data delete kar sake — kisi aur ka report delete na ho
        if not is_admin_or_above(session) and record.get("input_user") != current_user:
            return jsonify({"status": "error", "message": "You can only delete your own reports."}), 403
        filename = record["filename"]
        try:
            delete_from_s3(filename)
        except Exception as s3_exc:
            print(f"[CASE_REPORT] S3 delete warning: {s3_exc}")
        supabase.table("reports").delete().eq("id", report_id).execute()
        return jsonify({"status": "success", "message": f"Report {report_id} deleted."})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

# ============================================================
# AML GUI — BULK REGENERATE CASES (Steps 1-4 workflow)
# ============================================================
@app.route("/bulk-regenerate-cases", methods=["POST"])
@login_required
def bulk_regenerate_cases():
    
    try:
        data = request.get_json(silent=True) or {}
        report_ids = data.get("report_ids", [])
        mode = "basic"
        if not report_ids:
            return jsonify({"status": "error", "message": "No report_ids provided"}), 400
        if len(report_ids) > MAX_BULK_REGENERATE_REPORTS:
            return jsonify({
                "status": "error",
                "message": f"Maximum {MAX_BULK_REGENERATE_REPORTS} reports can be regenerated in one batch."
            }), 400

        job_id = uuid.uuid4().hex
        current_email = session.get("email", "")
        current_display_name = session.get("display_name", "User")
        input_user = get_clean_display_name(current_display_name)
        aml_creds = get_aml_credentials_for_user(current_email, current_display_name)

        with REGEN_JOBS_LOCK:
            REGEN_JOBS[job_id] = {
                "status": "running",
                "phase": "queued",
                "total": len(report_ids),
                "completed": 0,
                "results": [],
                "folder": None,
                "final_excel": None,
                "message": None,
                "stop_event": threading.Event(),
                "driver": None,
            }

        thread = threading.Thread(
            target=_run_bulk_regenerate_job,
            args=(job_id, list(report_ids), mode, input_user, aml_creds),
            daemon=True
        )
        thread.start()

        return jsonify({"status": "success", "job_id": job_id})
    except Exception as exc:
        return jsonify({"status": "error", "message": str(exc)}), 500

@app.route("/download-regenerate-file/<job_id>/<file_type>", methods=["GET"])
@login_required
def download_regenerate_file(job_id, file_type):
    with REGEN_JOBS_LOCK:
        job = REGEN_JOBS.get(job_id)
        job_data = _regen_job_response(job) if job else None
    if not job_data:
        job_data = _load_regen_job_snapshot(job_id)
    if not job_data:
        return jsonify({"status": "error", "message": "Job not found"}), 404

    allowed_types = {
        "final_excel": job_data.get("final_excel"),
    }
    file_path = allowed_types.get(file_type)
    if not file_path or not os.path.isfile(file_path):
        return jsonify({"status": "error", "message": "File not available"}), 404

    return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))


@app.route("/regenerate-job-status/<job_id>", methods=["GET"])
@login_required
def regenerate_job_status(job_id):
    with REGEN_JOBS_LOCK:
        job = REGEN_JOBS.get(job_id)
        job_data = _regen_job_response(job) if job else None
    if not job_data:
        job_data = _load_regen_job_snapshot(job_id)
    if not job_data:
        return jsonify({"status": "error", "message": "Job not found"}), 404
    return jsonify(job_data)


@app.route("/stop-regenerate-job/<job_id>", methods=["POST"])
@login_required
def stop_regenerate_job(job_id):
    with REGEN_JOBS_LOCK:
        job = REGEN_JOBS.get(job_id)
        if not job:
            return jsonify({"status": "error", "message": "Job not found"}), 404
        job["stop_event"].set()
        if job["status"] == "running":
            job["status"] = "stopping"
        drv = job.get("driver")
    if drv:
        try:
            drv.quit()
        except Exception:
            pass
    return jsonify({"status": "success", "message": "Stop signal sent"})

@app.route("/api/total-numbers", methods=["GET"])
def api_total_numbers_list():
    try:
        department     = request.args.get("department",     "").strip()
        account_status = request.args.get("account_status", "").strip()
        number_type    = request.args.get("number_type",    "").strip()
        sim_operator   = request.args.get("sim_operator",   "").strip()
        search         = request.args.get("search",         "").strip()
        page           = max(1, int(request.args.get("page",     1)))
        per_page       = min(500, max(1, int(request.args.get("per_page", 100))))
        query = social_supabase.table("social_media_accounts") \
            .select(TN_COLUMNS, count="exact") \
            .eq("platform", "Total Numbers")
        if department:
            query = query.eq("department", department)
        if account_status:
            query = query.eq("account_status", account_status)
        if number_type:
            query = query.eq("number_type", number_type)
        if sim_operator:
            query = query.ilike("sim_operator", f"%{sim_operator}%")
        if search:
            lt = f"%{search}%"
            query = query.or_(
                f"owned_by.ilike.{lt},"
                f"number.ilike.{lt},"
                f"sim_inserted_device.ilike.{lt},"
                f"account_status.ilike.{lt},"
                f"number_type.ilike.{lt},"
                f"sim_operator.ilike.{lt},"
                f"department.ilike.{lt}"
            )
        offset = (page - 1) * per_page
        resp   = query.order("id", desc=False) \
                      .range(offset, offset + per_page - 1) \
                      .execute()
        items      = resp.data or []
        total_rows = resp.count or 0
        return jsonify({
            "success":     True,
            "total":       total_rows,
            "page":        page,
            "per_page":    per_page,
            "total_pages": max(1, math.ceil(total_rows / per_page)),
            "items":       items
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/total-numbers/<int:record_id>", methods=["GET"])
def api_total_numbers_get(record_id):
    try:
        resp = social_supabase.table("social_media_accounts") \
            .select(TN_COLUMNS) \
            .eq("id", record_id) \
            .eq("platform", "Total Numbers") \
            .limit(1) \
            .execute()
        if not resp.data:
            return jsonify({"success": False, "error": "Record not found"}), 404
        return jsonify({"success": True, "record": resp.data[0]})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/total-numbers/stats", methods=["GET"])
def api_total_numbers_stats():
    try:
        resp = social_supabase.table("social_media_accounts") \
            .select(TN_COLUMNS) \
            .eq("platform", "Total Numbers") \
            .execute()
        rows = resp.data or []
        status_counts   = {}
        num_type_counts = {}
        dept_counts     = {}
        sim_op_counts   = {}
        for r in rows:
            s  = (r.get("account_status") or "Unknown").strip()
            nt = (r.get("number_type")    or "Unknown").strip()
            d  = (r.get("department")     or "Unknown").strip()
            so = (r.get("sim_operator")   or "Unknown").strip()
            status_counts[s]   = status_counts.get(s,   0) + 1
            num_type_counts[nt] = num_type_counts.get(nt, 0) + 1
            dept_counts[d]     = dept_counts.get(d,     0) + 1
            sim_op_counts[so]  = sim_op_counts.get(so,  0) + 1
        return jsonify({
            "success":        True,
            "total":          len(rows),
            "by_status":      status_counts,
            "by_number_type": num_type_counts,
            "by_department":  dept_counts,
            "by_sim_operator": dict(
                sorted(sim_op_counts.items(), key=lambda x: x[1], reverse=True)
            )
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
@app.route("/social-download-template", methods=["GET"])
@login_required
def social_download_template():
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation

    platform = request.args.get("platform", "").strip()

    PLATFORM_COLUMNS_MAP = {
        'Facebook':       ['login_user','number','login_device','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','account_type','mail_id','account_id','password','page_name','owned_by','platform','department'],
        'Instagram':      ['login_user','number','login_device','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','mail_id','account_id','password','owned_by','platform','department'],
        'Telegram':       ['login_user','number','login_device','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','owned_by','platform','department'],
        'WhatsApp':       ['login_user','number','login_device','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','account_type','owned_by','platform','department'],
        'Amazon':         ['login_user','number','login_device','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','account_type','mail_id','account_id','password','owned_by','platform','department'],
        'Gmail Accounts': ['login_user','number','account_status','review_status','blocked_date','unblock_date','account_create_date','full_name','mail_id','password','owned_by','platform','department'],
        'Total Numbers':  ['owned_by','number','sim_inserted_device','account_status','review_status','number_type','blocked_date','unblock_date','account_create_date','sim_operator','full_name','recharge_date','sim_buy_date','platform','department'],
    }

    ALL_HEADERS = [
        'owned_by','login_user','number','login_device','sim_inserted_device',
        'account_status','review_status','number_type','blocked_date','unblock_date',
        'account_create_date','sim_operator','full_name','recharge_date','sim_buy_date',
        'account_type','mail_id','account_id','password','page_name','platform','department',
    ]

    PLATFORM_STATUS_OPTIONS = {
        'Facebook':       ['Active','Temporarily Block','Restricted','Permanent Block'],
        'Instagram':      ['Active','Temporarily Block','Permanent Block'],
        'Telegram':       ['Active','Frozen','Permanent Block'],
        'WhatsApp':       ['Active','Temporarily Block','Permanent Block','Restricted'],
        'Amazon':         ['Active','Temporarily Block','Permanent Block'],
        'Gmail Accounts': ['Active','Temporarily Block','Permanent Block'],
        'Total Numbers':  ['Active','Temporarily Block','Permanent Block'],
        '':               ['Active','Temporarily Block','Restricted','Frozen','Permanent Block'],
    }

    REVIEW_STATUS_OPTIONS = ['NA','Send','Appeal Submit','Video Verification Done']

    headers = PLATFORM_COLUMNS_MAP.get(platform, ALL_HEADERS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Social Media Accounts"
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 4)

    # Dropdowns — find column index for account_status and review_status
    status_options = PLATFORM_STATUS_OPTIONS.get(platform, PLATFORM_STATUS_OPTIONS[''])
    status_formula = '"' + ','.join(status_options) + '"'
    review_formula = '"' + ','.join(REVIEW_STATUS_OPTIONS) + '"'

    if 'account_status' in headers:
        col_letter = openpyxl.utils.get_column_letter(headers.index('account_status') + 1)
        dv_status = DataValidation(
            type="list",
            formula1=status_formula,
            allow_blank=True,
            showDropDown=False
        )
        dv_status.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv_status)

    if 'review_status' in headers:
        col_letter = openpyxl.utils.get_column_letter(headers.index('review_status') + 1)
        dv_review = DataValidation(
            type="list",
            formula1=review_formula,
            allow_blank=True,
            showDropDown=False
        )
        dv_review.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv_review)

    # If platform is fixed, pre-fill platform column
    if platform and 'platform' in headers:
        col_letter = openpyxl.utils.get_column_letter(headers.index('platform') + 1)
        plat_dv = DataValidation(
            type="list",
            formula1=f'"{platform}"',
            allow_blank=False,
            showDropDown=False
        )
        plat_dv.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(plat_dv)
        for row in range(2, 6):
            ws.cell(row=row, column=headers.index('platform') + 1, value=platform)

    # number_type dropdown
    if 'number_type' in headers:
        col_letter = openpyxl.utils.get_column_letter(headers.index('number_type') + 1)
        dv_nt = DataValidation(
            type="list",
            formula1='"Prepaid,Postpaid,Disposable Number,NA"',
            allow_blank=True,
            showDropDown=False
        )
        dv_nt.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv_nt)

    # platform dropdown (agar platform fix nahi hai toh saare options do)
    if 'platform' in headers and not platform:
        col_letter = openpyxl.utils.get_column_letter(headers.index('platform') + 1)
        all_platforms = 'Facebook,Instagram,Telegram,WhatsApp,Amazon,Gmail Accounts,Total Numbers'
        dv_plat = DataValidation(
            type="list",
            formula1=f'"{all_platforms}"',
            allow_blank=True,
            showDropDown=False
        )
        dv_plat.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv_plat)

    # department dropdown
    if 'department' in headers:
        col_letter = openpyxl.utils.get_column_letter(headers.index('department') + 1)
        dv_dept = DataValidation(
            type="list",
            formula1='"AML,Investment Scam,ITC,Infringement,Chargeback"',
            allow_blank=True,
            showDropDown=False
        )
        dv_dept.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv_dept)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    platform_suffix = f"_{platform.replace(' ', '_')}" if platform else ""
    filename = f"Social_Media_Accounts{platform_suffix}_Template.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================
# LUNCH BREAK TRACKER
# ============================================================
@app.route("/lunch-break", methods=["GET"])
@login_required
def lunch_break():
    user = get_current_user()
    is_admin = session.get("is_admin", False)
    allowed_pages = session.get("allowed_pages", [])
    if not can_access_lunch(session):
        flash("Access denied.", "error")
        return redirect_to_allowed_page(allowed_pages)

    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to",   "").strip()
    emp_filter  = request.args.get("emp_filter",  "").strip()

    # Lunch page access means the user can see and fill lunch for everyone.
    visible_employees = ALL_EMPLOYEES if "lunch" in allowed_pages else []

    try:
        query = get_auth_supabase().table("lunch_breaks").select("*")
        if date_from:
            query = query.gte("date", date_from)
        if date_to:
            query = query.lte("date", date_to)
        if emp_filter:
            query = query.eq("employee_name", emp_filter)

        resp = query.order("date", desc=True).order("id", desc=True).limit(500).execute()
        raw_items = resp.data or []
        # A logged-in user must NOT see their OWN lunch-break entries. We identify
        # "own" records by matching the record's employee_name against the currently
        # authenticated user's clean display name (the existing identity mapping used
        # throughout the project). Other users' records stay fully visible.
        items = [it for it in raw_items if not is_own_lunch_record(it, user)]
        total = len(items)
    except Exception as e:
        items = []
        total = 0
        flash(f"Error: {e}", "error")

    clean_display_name = get_clean_display_name(session.get("display_name", "User"))

    return render_template(
        "lunch_break.html",
        items=items,
        total=total,
        is_admin=is_admin_or_above(session),
        role=session.get("role", "user"),
        visible_employees=visible_employees,
        all_employees=ALL_EMPLOYEES,
        date_from=date_from,
        date_to=date_to,
        emp_filter=emp_filter,
        current_user=user,
        allowed_pages=allowed_pages,
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        can_view_activity_log=session.get("can_view_activity_log", False),
        current_user_email=session.get("email", ""),
    )


@app.route("/lunch-break/insert", methods=["POST"])
@login_required
def lunch_break_insert():
    try:
        data = request.get_json()
        email    = session.get("email", "")
        allowed_pages = session.get("allowed_pages", [])

        employee_name = (data.get("employee_name") or "").strip()
        date_val      = (data.get("date")          or "").strip()
        start_time    = (data.get("lunch_start")   or "").strip()
        end_time      = (data.get("lunch_end")     or "").strip()
        remark        = (data.get("remark")        or "NA").strip()

        if not employee_name or not date_val or not start_time:
            return jsonify({"success": False, "error": "Date, employee name and start time are required"})

        if "lunch" not in allowed_pages:
            return jsonify({"success": False, "error": "Lunch tracker did not have access"}), 403
        if employee_name not in ALL_EMPLOYEES:
            return jsonify({"success": False, "error": "Invalid employee name"})

        duration_str = "NA"
        if end_time:
            try:
                from datetime import datetime as dt
                fmt = "%H:%M"
                s = dt.strptime(start_time, fmt)
                e = dt.strptime(end_time, fmt)
                diff = e - s
                total_mins = int(diff.total_seconds() / 60)
                if total_mins < 0:
                    return jsonify({"success": False, "error": "End time cannot be before start time"})
                hours   = total_mins // 60
                minutes = total_mins % 60
                duration_str = f"{hours}h {minutes}m" if hours else f"{minutes}m"
            except Exception:
                duration_str = "NA"

        record = {
            "date":                 date_val,
            "employee_name":        employee_name,
            "lunch_start":          start_time,
            "lunch_end":            end_time or None,
            "total_break_duration": duration_str,
            "remark":               remark if remark else "NA",
            "filled_by":            get_clean_display_name(session.get("display_name", "")),
            "filled_by_email":      email,
        }

        resp = get_auth_supabase().table("lunch_breaks").insert(record).execute()
        if resp.data:
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Insert failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/lunch-break/update", methods=["POST"])
@login_required
def lunch_break_update():
    try:
        data = request.get_json()
        rid      = data.get("id")
        allowed_pages = session.get("allowed_pages", [])
        if not rid:
            return jsonify({"success": False, "error": "No ID"})

        if "lunch" not in allowed_pages:
            return jsonify({"success": False, "error": "Lunch tracker did not have access"}), 403

        from datetime import datetime as dt
        start = data.get("lunch_start", "")
        end   = data.get("lunch_end", "")
        duration_str = "NA"
        try:
            fmt = "%H:%M"
            s = dt.strptime(start, fmt)
            e = dt.strptime(end, fmt)
            diff_mins = int((e - s).total_seconds() / 60)
            if diff_mins >= 0:
                h, m = diff_mins // 60, diff_mins % 60
                duration_str = f"{h}h {m}m" if h else f"{m}m"
        except Exception:
            pass

        record = {
            "date":                 data.get("date", ""),
            "lunch_start":          start,
            "lunch_end":            end,
            "total_break_duration": duration_str,
            "remark":               data.get("remark", "NA") or "NA",
        }
        resp = get_auth_supabase().table("lunch_breaks").update(record).eq("id", rid).execute()
        if resp.data:
            return jsonify({"success": True})
        return jsonify({"success": False, "error": "Update failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/lunch-break/delete", methods=["POST"])
@login_required
def lunch_break_delete():
    try:
        data = request.get_json()
        rid  = data.get("id")
        allowed_pages = session.get("allowed_pages", [])
        if not rid:
            return jsonify({"success": False, "error": "No ID"})

        if "lunch" not in allowed_pages:
            return jsonify({"success": False, "error": "Lunch tracker did not have access"}), 403

        get_auth_supabase().table("lunch_breaks").delete().eq("id", rid).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/lunch-break/export", methods=["GET"])
@login_required
def lunch_break_export():
    try:
        allowed_pages = session.get("allowed_pages", [])
        date_from = request.args.get("date_from", "").strip()
        date_to   = request.args.get("date_to",   "").strip()
        emp_filter  = request.args.get("emp_filter",  "").strip()

        if "lunch" not in allowed_pages:
            flash("Access denied.", "error")
            return redirect_to_allowed_page(allowed_pages)

        query = get_auth_supabase().table("lunch_breaks").select("*")
        if date_from:
            query = query.gte("date", date_from)
        if date_to:
            query = query.lte("date", date_to)
        if emp_filter:
            query = query.eq("employee_name", emp_filter)

        resp = query.order("date", desc=False).execute()
        rows_all = resp.data or []
        current_user = get_current_user()
        rows = [r for r in rows_all if not is_own_lunch_record(r, current_user)]
        df   = pd.DataFrame(rows) if rows else pd.DataFrame()
        out  = io.StringIO()
        df.to_csv(out, index=False, encoding="utf-8-sig")
        out.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            io.BytesIO(out.getvalue().encode("utf-8-sig")),
            download_name=f"lunch_breaks_{ts}.csv",
            as_attachment=True,
            mimetype="text/csv"
        )
    except Exception as e:
        flash(f"Export error: {e}", "error")
        return redirect("/lunch-break")

# ============================================================
# DASHBOARD MANAGEMENT (dashboard_users)
# ============================================================
DASHBOARD_MANAGEMENT_PAGE = "dashboard_management"

# Full set of page keys a user can be granted (for the Allowed Pages picker).
ALLOWED_PAGES_OPTIONS = [
    ("scraping",          "Data Scraping"),
    ("sheet",             "Summary & Sheet Generator"),
    ("social",            "Social Media Accounts"),
    ("investment",        "Investment Scam Data"),
    ("qc",                "QC Review"),
    ("website_directory", "Investment Website Directory"),
    ("allotment",         "Website Case Assignment"),
    ("allotment_admin",   "Website Case Assignment Admin"),
    ("insights",          "User Performance Dashboard"),
    ("case_report",       "Case & Report Generator"),
    ("lunch",             "Break Time Management"),
    (DASHBOARD_MANAGEMENT_PAGE, "User & Access Management"),
]

INVESTMENT_SCAM_USERS_TABLE = "investment_scam_users"

GUI_STATUS_TABLES = [
    {
        "page_key": "scraping",
        "page_name": "Data Scraping",
        "table": "scrapping_data",
        "date_columns": ["created_at", "updated_at", "Inserted_date", "inserted_date", "id"],
    },
    {
        "page_key": "social",
        "page_name": "Social Media Accounts",
        "table": "social_media_accounts",
        "date_columns": ["updated_at", "created_at", "blocked_date", "account_create_date", "id"],
    },
    {
        "page_key": "investment",
        "page_name": "Investment Scam Data",
        "table": "BS_Investment_Scam",
        "date_columns": ["Inserted_date", "created_at", "updated_at", "Id"],
    },
    {
        "page_key": "website_directory",
        "page_name": "Investment Website Directory",
        "table": "website_directory",
        "date_columns": ["updated_at", "created_at", "date", "id"],
    },
    {
        "page_key": "allotment",
        "page_name": "Website Case Assignment",
        "table": "website_allotment",
        "date_columns": ["updated_at", "created_at", "allotted_date", "id"],
    },
    {
        "page_key": "case_report",
        "page_name": "Case & Report Generator",
        "table": "reports",
        "date_columns": ["created_at", "updated_at", "id"],
    },
    {
        "page_key": "lunch",
        "page_name": "Break Time Management",
        "table": "lunch_breaks",
        "date_columns": ["date", "updated_at", "created_at", "id"],
    },
    {
        "page_key": "qc",
        "page_name": "QC Review",
        "table": "qc_table",
        "date_columns": ["updated_at", "created_at", "inserted_datetime", "inserted_date", "id"],
    },
    {
        "page_key": "dashboard_management",
        "page_name": "Dashboard Users",
        "table": "dashboard_users",
        "date_columns": ["updated_at", "created_at", "id"],
    },
    {
        "page_key": "investment_users",
        "page_name": "Investment Scam Users",
        "table": INVESTMENT_SCAM_USERS_TABLE,
        "date_columns": ["updated_at", "created_at", "joining_date", "id"],
    },
]


def can_manage_dashboard(user_session=None):
    """Only a superadmin who has 'dashboard_management' in their allowed_pages can
    manage users. This is the single source of truth used by every
    Dashboard-Management route (page + JSON), so it cannot be bypassed via URL."""
    user_session = session if user_session is None else user_session
    allowed_pages = user_session.get("allowed_pages") or []
    return is_superadmin(user_session) and DASHBOARD_MANAGEMENT_PAGE in allowed_pages


def dashboard_management_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not can_manage_dashboard():
            flash("Access denied.", "error")
            return redirect_to_allowed_page(session.get("allowed_pages", []))
        return f(*args, **kwargs)
    return decorated


def dashboard_management_required_json(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not can_manage_dashboard():
            return jsonify({"success": False, "error": "Access denied."}), 403
        return f(*args, **kwargs)
    return decorated


def parse_string_list(value):
    """Normalize an allowed_pages / allowed_departments value from the client
    (list, comma separated string, or JSON string) into a clean Python list."""
    if value is None:
        return []
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                value = parsed
            else:
                return [v.strip() for v in value.split(",") if v.strip()]
        except Exception:
            return [v.strip() for v in value.split(",") if v.strip()]
    return [str(v).strip() for v in value if str(v).strip()]


def mask_password(password):
    """Return a non-reversible placeholder so plaintext passwords are never
    exposed in the UI. The real value is never sent to the browser."""
    return "********" if password else ""


def is_own_lunch_record(record, user_ctx):
    """Return True when a lunch_breaks record belongs to the currently logged-in user.

    The lunch tracker stores the EMPLOYEE name on each record (employee_name), and the
    authenticated dashboard user has a display_name. We consider a record the user's
    OWN when its employee_name matches the user's clean display name. This keeps every
    OTHER user's records fully visible (including records the current user may have
    filled for others) while hiding only the current user's own entries."""
    current_clean = get_clean_display_name((user_ctx or {}).get("display_name", ""))
    if not current_clean or current_clean == "User":
        return False
    emp_name = (record.get("employee_name") or "").strip()
    if not emp_name:
        return False
    return emp_name.strip().lower() == current_clean.strip().lower()


@app.route("/dashboard-management", methods=["GET"])
@login_required
@dashboard_management_required
def dashboard_management():
    return render_template(
        "dashboard_management.html",
        allowed_pages=session.get("allowed_pages", []),
        is_admin=is_admin_or_above(session),
        can_manage_dashboard=True,
        display_name=session.get("display_name", "User"),
        clean_display_name=get_clean_display_name(session.get("display_name", "User")),
        can_view_activity_log=session.get("can_view_activity_log", False),
        current_user=get_current_user(),
        page_options=ALLOWED_PAGES_OPTIONS,
        department_options=DEPARTMENT_OPTIONS,
    )


def _serialize_dashboard_user(row):
    """Serialize dashboard users for the admin-only User & Access Management UI."""
    out = dict(row)
    raw_password = out.get("password") or ""
    raw_aml_password = out.get("aml_password") or ""
    out["password"] = raw_password
    out["has_aml_password"] = bool(raw_aml_password)
    out["aml_password"] = raw_aml_password
    return out


def _estimate_json_bytes(rows):
    try:
        return len(json.dumps(rows or [], default=str).encode("utf-8"))
    except Exception:
        return 0


def _format_status_level(total_rows, query_ms, storage_bytes):
    if storage_bytes and storage_bytes >= 50 * 1024 * 1024:
        return "Heavy"
    if total_rows >= 50000 or query_ms >= 1500:
        return "Heavy"
    if storage_bytes and storage_bytes >= 10 * 1024 * 1024:
        return "Medium"
    if total_rows >= 10000 or query_ms >= 600:
        return "Medium"
    return "Normal"


def _load_table_storage_map(client):
    try:
        resp = client.rpc("get_table_storage_stats").execute()
        rows = resp.data or []
        return {
            (r.get("table_name") or r.get("table") or r.get("relname") or "").lower(): r
            for r in rows
            if r.get("table_name") or r.get("table") or r.get("relname")
        }, None
    except Exception as e:
        return {}, str(e)


def _latest_value_for_table(client, table_name, date_columns):
    for col in date_columns:
        try:
            resp = client.table(table_name).select(col).order(col, desc=True).limit(1).execute()
            if resp.data:
                return resp.data[0].get(col), col
        except Exception:
            continue
    return None, None


def _table_status_row(client, cfg, storage_map):
    started = time.perf_counter()
    row = {
        "page_key": cfg["page_key"],
        "page_name": cfg["page_name"],
        "table": cfg["table"],
        "total_rows": 0,
        "latest_value": None,
        "latest_column": None,
        "query_ms": 0,
        "sample_rows": 0,
        "sample_bytes": 0,
        "estimated_payload_bytes": 0,
        "storage_bytes": None,
        "table_bytes": None,
        "index_bytes": None,
        "storage_source": "RPC not configured",
        "status": "Unknown",
        "error": None,
    }
    try:
        count_resp = client.table(cfg["table"]).select("*", count="exact").limit(1).execute()
        row["total_rows"] = count_resp.count if count_resp.count is not None else 0

        latest_value, latest_column = _latest_value_for_table(client, cfg["table"], cfg["date_columns"])
        row["latest_value"] = latest_value
        row["latest_column"] = latest_column

        sample_resp = client.table(cfg["table"]).select("*").limit(25).execute()
        sample_rows = sample_resp.data or []
        row["sample_rows"] = len(sample_rows)
        row["sample_bytes"] = _estimate_json_bytes(sample_rows)
        if row["sample_rows"] and row["total_rows"]:
            row["estimated_payload_bytes"] = int((row["sample_bytes"] / row["sample_rows"]) * row["total_rows"])

        storage = storage_map.get(cfg["table"].lower()) if storage_map else None
        if storage:
            row["storage_bytes"] = storage.get("total_bytes") or storage.get("total_size")
            row["table_bytes"] = storage.get("table_bytes") or storage.get("table_size")
            row["index_bytes"] = storage.get("index_bytes") or storage.get("index_size")
            row["storage_source"] = "RPC"

        row["query_ms"] = int((time.perf_counter() - started) * 1000)
        row["status"] = _format_status_level(
            int(row["total_rows"] or 0),
            row["query_ms"],
            int(row["storage_bytes"] or 0),
        )
    except Exception as e:
        row["query_ms"] = int((time.perf_counter() - started) * 1000)
        row["status"] = "Error"
        row["error"] = str(e)
    return row


@app.route("/dashboard-management/api/users", methods=["GET"])
@login_required
@dashboard_management_required_json
def dashboard_management_users():
    try:
        resp = get_auth_supabase().table("dashboard_users") \
            .select("*").order("id", desc=False).execute()
        users = [_serialize_dashboard_user(u) for u in (resp.data or [])]
        return jsonify({"success": True, "users": users})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/gui-status", methods=["GET"])
@login_required
@dashboard_management_required_json
def dashboard_management_gui_status():
    try:
        client = get_auth_supabase()
        started = time.perf_counter()
        storage_map, rpc_error = _load_table_storage_map(client)
        rows = [_table_status_row(client, cfg, storage_map) for cfg in GUI_STATUS_TABLES]
        totals = {
            "pages": len(rows),
            "tables_ok": sum(1 for r in rows if not r.get("error")),
            "tables_error": sum(1 for r in rows if r.get("error")),
            "total_rows": sum(int(r.get("total_rows") or 0) for r in rows),
            "estimated_payload_bytes": sum(int(r.get("estimated_payload_bytes") or 0) for r in rows),
            "storage_bytes": sum(int(r.get("storage_bytes") or 0) for r in rows if r.get("storage_bytes")),
            "slowest_page": None,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total_query_ms": 0,
            "rpc_error": rpc_error,
        }
        if rows:
            totals["slowest_page"] = max(rows, key=lambda r: int(r.get("query_ms") or 0)).get("page_name")
        totals["total_query_ms"] = int((time.perf_counter() - started) * 1000)
        return jsonify({"success": True, "rows": rows, "totals": totals})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/investment-scam-users", methods=["GET"])
@login_required
@dashboard_management_required_json
def dashboard_management_investment_scam_users():
    try:
        resp = get_auth_supabase().table(INVESTMENT_SCAM_USERS_TABLE) \
            .select("*").order("id", desc=False).execute()
        return jsonify({"success": True, "users": resp.data or []})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


def _investment_user_duration_payload(employee_type, duration_months):
    employee_type = (employee_type or "employee").strip().lower()
    if employee_type not in ("employee", "intern"):
        employee_type = "employee"
    if employee_type != "intern":
        return "Employee", None, None, None

    try:
        months = int(duration_months or 0)
    except (TypeError, ValueError):
        months = 0
    if months <= 0:
        return "Intern", None, None, "Internship duration (months) is required for interns."
    return "Intern", f"{months} Months", months, None


def _calculate_internship_end_date(joining_date, months):
    if not joining_date or not months:
        return None
    try:
        start = datetime.strptime(str(joining_date)[:10], "%Y-%m-%d").date()
        month_index = start.month - 1 + int(months)
        year = start.year + month_index // 12
        month = month_index % 12 + 1
        last_day = (datetime(year + (month // 12), (month % 12) + 1, 1) - timedelta(days=1)).day
        day = min(start.day, last_day)
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except Exception:
        return None


@app.route("/dashboard-management/api/investment-scam-users", methods=["POST"])
@login_required
@dashboard_management_required_json
def dashboard_management_create_investment_scam_user():
    try:
        data = request.get_json() or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"success": False, "error": "Name is required."}), 400
        joining_date = (data.get("joining_date") or "").strip() or None
        if not joining_date:
            return jsonify({"success": False, "error": "Joining date is required."}), 400
        employee_type, duration, duration_months, duration_error = _investment_user_duration_payload(
            data.get("employee_type") or data.get("user_type"),
            data.get("duration_months") or data.get("internship_duration_months"),
        )
        if duration_error:
            return jsonify({"success": False, "error": duration_error}), 400
        password = data.get("paasword") or ""
        if not password:
            return jsonify({"success": False, "error": "Password is required."}), 400

        record = {
            "name": name,
            "employee_type": employee_type,
            "contact_no": (data.get("contact_no") or "").strip() or None,
            "mail":      (data.get("mail") or "").strip() or None,
            "joining_date": joining_date,
            "dob":       (data.get("dob") or "").strip() or None,
            "duration":   duration,
            "internship_end_date": _calculate_internship_end_date(joining_date, duration_months),
            "gui_cred":  (data.get("gui_cred") or "").strip() or None,
            "paasword":  password,
        }
        client = get_auth_supabase()
        resp = client.table(INVESTMENT_SCAM_USERS_TABLE).insert(record).execute()
        if resp.data:
            created = resp.data[0]
            log_activity("CREATE", target_table=INVESTMENT_SCAM_USERS_TABLE,
                         target_record_id=created.get("id"),
                         extra_info=f"Created investment scam user {name}")
            return jsonify({"success": True, "user": created}), 201
        return jsonify({"success": False, "error": "Create failed."}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/investment-scam-users/<int:user_id>", methods=["PUT"])
@login_required
@dashboard_management_required_json
def dashboard_management_update_investment_scam_user(user_id):
    try:
        data = request.get_json() or {}
        client = get_auth_supabase()
        existing = client.table(INVESTMENT_SCAM_USERS_TABLE) \
            .select("id").eq("id", user_id).limit(1).execute()
        if not existing.data:
            return jsonify({"success": False, "error": "User not found."}), 404

        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"success": False, "error": "Name is required."}), 400
        joining_date = (data.get("joining_date") or "").strip() or None
        if not joining_date:
            return jsonify({"success": False, "error": "Joining date is required."}), 400

        employee_type, duration, duration_months, duration_error = _investment_user_duration_payload(
            data.get("employee_type") or data.get("user_type"),
            data.get("duration_months") or data.get("internship_duration_months"),
        )
        if duration_error:
            return jsonify({"success": False, "error": duration_error}), 400

        record = {
            "name": name,
            "employee_type": employee_type,
            "contact_no": (data.get("contact_no") or "").strip() or None,
            "mail":      (data.get("mail") or "").strip() or None,
            "joining_date": joining_date,
            "dob":       (data.get("dob") or "").strip() or None,
            "duration":   duration,
            "internship_end_date": _calculate_internship_end_date(joining_date, duration_months),
            "gui_cred":  (data.get("gui_cred") or "").strip() or None,
        }
        new_pw = (data.get("paasword") or "").strip()
        if new_pw:
            record["paasword"] = new_pw

        upd = client.table(INVESTMENT_SCAM_USERS_TABLE).update(record).eq("id", user_id).execute()
        if upd.data:
            log_activity("UPDATE", target_table=INVESTMENT_SCAM_USERS_TABLE,
                         target_record_id=user_id,
                         extra_info=f"Updated investment scam user {name}")
            return jsonify({"success": True, "user": upd.data[0]})
        return jsonify({"success": False, "error": "Update failed."}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/investment-scam-users/<int:user_id>", methods=["DELETE"])
@login_required
@dashboard_management_required_json
def dashboard_management_delete_investment_scam_user(user_id):
    try:
        client = get_auth_supabase()
        existing = client.table(INVESTMENT_SCAM_USERS_TABLE) \
            .select("id,name").eq("id", user_id).limit(1).execute()
        if not existing.data:
            return jsonify({"success": False, "error": "User not found."}), 404
        name = existing.data[0].get("name") or f"id={user_id}"
        d = client.table(INVESTMENT_SCAM_USERS_TABLE).delete().eq("id", user_id).execute()
        log_activity("DELETE", target_table=INVESTMENT_SCAM_USERS_TABLE,
                     target_record_id=user_id,
                     extra_info=f"Deleted investment scam user {name}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/users", methods=["POST"])
@login_required
@dashboard_management_required_json
def dashboard_management_create_user():
    try:
        data = request.get_json() or {}
        email        = (data.get("email") or "").strip().lower()
        password     = data.get("password") or ""
        display_name = (data.get("display_name") or "").strip()

        if not email or "@" not in email:
            return jsonify({"success": False, "error": "A valid email address is required."}), 400
        if not password:
            return jsonify({"success": False, "error": "Password is required."}), 400
        if not display_name:
            return jsonify({"success": False, "error": "Display name is required."}), 400

        client = get_auth_supabase()
        existing = client.table("dashboard_users").select("id").eq("email", email).limit(1).execute()
        if existing.data:
            return jsonify({"success": False, "error": "A user with this email already exists."}), 409

        record = {
            "email":                 email,
            "password":              password,
            "display_name":          display_name,
            "allowed_pages":         parse_string_list(data.get("allowed_pages")),
            "is_admin":              parse_bool(data.get("is_admin")),
            "role":                  (data.get("role")
                                      or ("admin" if parse_bool(data.get("is_admin")) else "user")),
            "is_active":             parse_bool(data.get("is_active", True)),
            "can_view_activity_log": parse_bool(data.get("can_view_activity_log")),
            "allowed_departments":   parse_string_list(data.get("allowed_departments")) or None,
            "aml_username":          (data.get("aml_username") or "").strip() or None,
            "aml_password":          (data.get("aml_password") or "").strip() or None,
        }
        resp = client.table("dashboard_users").insert(record).execute()
        if resp.data:
            created = resp.data[0]
            log_activity("CREATE", target_table="dashboard_users",
                         target_record_id=created.get("id"),
                         extra_info=f"Created dashboard user {email}")
            return jsonify({"success": True, "user": _serialize_dashboard_user(created)}), 201
        return jsonify({"success": False, "error": "Create failed."}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/users/<int:user_id>", methods=["PUT"])
@login_required
@dashboard_management_required_json
def dashboard_management_update_user(user_id):
    try:
        data = request.get_json() or {}
        client = get_auth_supabase()
        existing_resp = client.table("dashboard_users").select("*").eq("id", user_id).limit(1).execute()
        if not existing_resp.data:
            return jsonify({"success": False, "error": "User not found."}), 404

        email        = (data.get("email") or "").strip().lower()
        display_name = (data.get("display_name") or "").strip()
        if not email or "@" not in email:
            return jsonify({"success": False, "error": "A valid email address is required."}), 400
        if not display_name:
            return jsonify({"success": False, "error": "Display name is required."}), 400

        dup = client.table("dashboard_users").select("id") \
            .eq("email", email).neq("id", user_id).limit(1).execute()
        if dup.data:
            return jsonify({"success": False, "error": "Another user already has this email."}), 409

        record = {
            "email":                 email,
            "display_name":          display_name,
            "allowed_pages":         parse_string_list(data.get("allowed_pages")),
            "is_admin":              parse_bool(data.get("is_admin")),
            "role":                  (data.get("role")
                                      or ("admin" if parse_bool(data.get("is_admin")) else "user")),
            "is_active":             parse_bool(data.get("is_active")),
            "can_view_activity_log": parse_bool(data.get("can_view_activity_log")),
            "allowed_departments":   parse_string_list(data.get("allowed_departments")) or None,
            "aml_username":          (data.get("aml_username") or "").strip() or None,
        }
        # Password is only replaced when the admin supplies a new one; otherwise the
        # existing credential (required by the current auth system) is preserved.
        new_password = (data.get("password") or "").strip()
        if new_password and new_password != "********":
            record["password"] = new_password
        # Same for the AML password: a blank / masked value keeps the existing one.
        new_aml_pass = (data.get("aml_password") or "").strip()
        if new_aml_pass and new_aml_pass != "********":
            record["aml_password"] = new_aml_pass

        upd = client.table("dashboard_users").update(record).eq("id", user_id).execute()
        if upd.data:
            log_activity("UPDATE", target_table="dashboard_users", target_record_id=user_id,
                         extra_info=f"Updated dashboard user {email}")
            return jsonify({"success": True, "user": _serialize_dashboard_user(upd.data[0])})
        return jsonify({"success": False, "error": "Update failed."}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/users/<int:user_id>", methods=["DELETE"])
@login_required
@dashboard_management_required_json
def dashboard_management_delete_user(user_id):
    try:
        client = get_auth_supabase()
        found = client.table("dashboard_users").select("id,email").eq("id", user_id).limit(1).execute()
        if not found.data:
            return jsonify({"success": False, "error": "User not found."}), 404
        email = found.data[0].get("email") or f"id={user_id}"
        client.table("dashboard_users").delete().eq("id", user_id).execute()
        log_activity("DELETE", target_table="dashboard_users", target_record_id=user_id,
                     extra_info=f"Deleted dashboard user {email}")
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/dashboard-management/api/users/<int:user_id>/toggle-active", methods=["POST"])
@login_required
@dashboard_management_required_json
def dashboard_management_toggle_active(user_id):
    try:
        data = request.get_json() or {}
        is_active = parse_bool(data.get("is_active"))
        client = get_auth_supabase()
        found = client.table("dashboard_users").select("id,email").eq("id", user_id).limit(1).execute()
        if not found.data:
            return jsonify({"success": False, "error": "User not found."}), 404
        upd = client.table("dashboard_users").update({"is_active": is_active}).eq("id", user_id).execute()
        if upd.data:
            log_activity("UPDATE", target_table="dashboard_users", target_record_id=user_id,
                         extra_info=f"Set is_active={is_active} for {found.data[0].get('email')}")
            return jsonify({"success": True, "user": _serialize_dashboard_user(upd.data[0])})
        return jsonify({"success": False, "error": "Update failed."}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================================
# GUI QC — LIST / FILTER / PAGINATION
# ============================================================
QC_COLUMNS = [
    "id", "gui_id", "date", "input_user", "approved_by", "scam_type",
    "search_for", "upi_bank_account_wallet", "bank_name",
    "bank_account_number", "upi_vpa", "ifsc_code", "ac_holder_name",
    "web_contact_no", "payment_gateway_url", "website_url", "screenshot",
    "screenshot_case_report_link", "qc_remarks",
    "new_update_remark", "qc_status", "feature_type",
    "qc_assigned_date", "qc_approved_date", "total_allotment_date",
    "inserted_date", "inserted_datetime"
]

# Search For options shown in the GUI QC Update Report search section.
# This is a QC-specific list so it does not affect other pages; "App" is added
# as a new option per the QC page requirement.
QC_SEARCH_FOR_OPTIONS = list(BS_INVESTMENT_SEARCH_FOR_OPTIONS) + ["App"]

QC_TABLE = "qc_table"

# Columns in qc_table whose Postgres type is integer — need numeric conversion so
# empty / "NA" cells become NULL instead of an unparseable string at insert time.
QC_INT_COLUMNS = {"gui_id", "bank_account_number", "web_contact_no"}

# QC Remark options for the dropdown in the edit modal
QC_REMARKS_OPTIONS = [
    "Approved",
    "Major Mistake - Wrong Website/Group/URL Case",
    "Major Mistake - Wrong/Non-Existent Payment Details Filled",
    "Major Mistake - Wrong Payment Gateway URL",
    "Major Mistake - Wrong Video/Non-Linear Recording/Background Sound",
    "Major Mistake - Wrong First/Second Screenshot/ Same Screenshots",
    "Major Mistake - Payment Details/Web URL/PG URL/Mobile Screenshot Not Showing",
    "Major Mistake - Extension/Profile Picture/Taskbar & Tab Showing",
    "Major Mistake - Post URL and WhatsApp Number are Mismatched",
    "Minor Mistake - First/Second Screenshot Is Missing",
    "Minor Mistake - Wrong Transaction/Transfer Mode/Search For Option",
    "Minor Mistake - Worked on Duplicate/Extra/Others Website Or Timeline Not Followed",
    "GUI Issue - First/Second Screenshot Is Crossed",
    "GUI Issue",
    "Rejected",
]

QC_SEARCH_FIELDS = [
    "gui_id", "input_user", "approved_by", "scam_type", "search_for",
    "upi_bank_account_wallet", "bank_name", "bank_account_number",
    "upi_vpa", "ifsc_code", "ac_holder_name", "web_contact_no",
    "payment_gateway_url", "website_url", "qc_remarks",
    "new_update_remark", "feature_type"
]

QC_EXPORT_COLUMNS = [
    "gui_id", "input_user", "approved_by", "scam_type",
    "search_for", "upi_bank_account_wallet", "bank_name",
    "bank_account_number", "upi_vpa", "ifsc_code", "ac_holder_name",
    "web_contact_no", "payment_gateway_url", "website_url", "screenshot",
    "screenshot_case_report_link", "qc_remarks", "new_update_remark",
    "qc_status", "feature_type", "inserted_date",
    "qc_assigned_date", "qc_approved_date", "total_allotment_date"
]

QC_IMPORT_ALIASES = {
    "gui_id": ["gui_id", "gui id", "gui", "id"],
    "input_user": ["input_user", "input user", "user", "username"],
    "approved_by": ["approved_by", "approved by"],
    "scam_type": ["scam_type", "scam type", "type", "category"],
    "search_for": ["search_for", "search for"],
    "upi_bank_account_wallet": ["upi_bank_account_wallet", "upi bank account wallet", "wallet", "payment type"],
    "bank_name": ["bank_name", "bank name"],
    "bank_account_number": ["bank_account_number", "bank account number", "account_number", "account number", "acc_no", "acc no"],
    "upi_vpa": ["upi_vpa", "upi vpa", "upi", "vpa"],
    "ifsc_code": ["ifsc_code", "ifsc code", "ifsc"],
    "ac_holder_name": ["ac_holder_name", "ac holder name", "account holder name", "holder_name"],
    "web_contact_no": ["web_contact_no", "web contact no", "contact number", "phone", "mobile"],
    "payment_gateway_url": ["payment_gateway_url", "payment gateway url", "payment url", "gateway url"],
    "website_url": ["website_url", "website url", "url", "website"],
    "screenshot": ["screenshot", "image", "proof"],
    "screenshot_case_report_link": ["screenshot_case_report_link", "screenshot case report link", "case report link"],
    "qc_remarks": ["qc_remarks", "qc remarks"],
    "new_update_remark": ["new_update_remark", "new update remark", "update remark"],
    "qc_status": ["qc_status", "qc status"],
    "feature_type": ["feature_type", "feature type"],
    "qc_assigned_date": ["qc_assigned_date", "qc assigned date", "assigned date"],
    "qc_approved_date": ["qc_approved_date", "qc approved date", "approved date"],
    "total_allotment_date": ["total_allotment_date", "total allotment date", "allotment date"],
    "inserted_date": ["inserted_date", "inserted date"],
}


def _qc_normalize_header(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def _qc_find_import_column(df_columns, target_col):
    normalized = {_qc_normalize_header(c): c for c in df_columns}
    candidates = [target_col] + QC_IMPORT_ALIASES.get(target_col, [])
    for candidate in candidates:
        key = _qc_normalize_header(candidate)
        if key in normalized:
            return normalized[key]
    return None


def _qc_normalize_int_value(value):
    raw = str(value or "").strip()
    if raw.upper() in ("", "NA", "N/A", "NONE", "NULL", "NAN", "UNDEFINED"):
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _qc_build_records_from_df(df, input_user_default, force_input_user=None, reset_qc_update=False):
    df = df.fillna("")
    df.columns = df.columns.astype(str).str.strip()
    source_columns = list(df.columns)
    column_lookup = {
        target: _qc_find_import_column(source_columns, target)
        for target in QC_IMPORT_ALIASES.keys()
    }
    matched_columns = [col for col in column_lookup.values() if col]
    if not matched_columns:
        raise ValueError("No matching QC column names found.")

    today = datetime.now().strftime("%Y-%m-%d")
    records = []
    for _, row in df.iterrows():
        rec = {}
        for target_col, source_col in column_lookup.items():
            if not source_col:
                continue
            val = str(row[source_col]).strip()
            if target_col in QC_INT_COLUMNS:
                rec[target_col] = _qc_normalize_int_value(val)
            elif val.lower() in ("nan", "none", "null", "na", "n/a", "undefined", ""):
                rec[target_col] = None if target_col == "inserted_date" else "NA"
            else:
                rec[target_col] = val

        rec["input_user"] = force_input_user or rec.get("input_user") or input_user_default
        if rec.get("input_user") == "NA":
            rec["input_user"] = force_input_user or input_user_default
        if not rec.get("inserted_date") or rec.get("inserted_date") == "NA":
            rec["inserted_date"] = today
        if reset_qc_update:
            rec["new_update_remark"] = "NA"
            rec["qc_status"] = None
            rec["approved_by"] = "NA"
            rec["qc_assigned_date"] = today
            rec["total_allotment_date"] = today
            rec["qc_approved_date"] = None
        records.append(rec)
    return records, matched_columns


@app.route("/qc-gui", methods=["GET"])
@login_required
def qc_gui():
    allowed_pages = session.get("allowed_pages", [])
    if "qc" not in allowed_pages:
        flash("You don't have access to GUI QC.", "error")
        return redirect_to_allowed_page(allowed_pages)

    qc_search     = request.args.get("qc_search",     "").strip()
    qc_wallet     = request.args.get("qc_wallet",     "").strip()
    qc_date_from  = request.args.get("qc_date_from",  "").strip()
    qc_date_to    = request.args.get("qc_date_to",    "").strip()
    qc_status     = request.args.get("qc_status",     "").strip()
    is_super_or_admin = is_admin_or_above(session)
    page = int(request.args.get("page_num", 1))
    items = []
    total_rows = 0
    total_pages = 1
    today_qc_count = 0
    range_qc_count = 0
    range_qc_label = ""
    try:
        query = supabase.table(QC_TABLE).select("*", count="exact")
        if qc_search:
            lt = f"%{qc_search}%"
            or_parts = ",".join(f"{f}.ilike.{lt}" for f in QC_SEARCH_FIELDS)
            query = query.or_(or_parts)
        if qc_wallet:
            query = query.eq("upi_bank_account_wallet", qc_wallet)
        if qc_date_from:
            query = query.gte("inserted_date", qc_date_from)
        if qc_date_to:
            query = query.lte("inserted_date", qc_date_to)
        # QC Status filter: Pending (qc_status IS NULL - never reviewed) or QC Done (qc_status IS NOT NULL)
        if qc_status == "pending":
            query = query.is_("qc_status", "null")
        elif qc_status == "done":
            query = query.not_.is_("qc_status", "null")
        query = query.order("id", desc=True)
        offset = (page - 1) * PER_PAGE
        query = query.range(offset, offset + PER_PAGE - 1)
        resp = query.execute()
        items = resp.data or []
        total_rows = resp.count or 0
        total_pages = max(1, math.ceil(total_rows / PER_PAGE))

        # Today's QC count for the current user (records they reviewed today)
        current_user_name = get_clean_display_name(session.get("display_name", "User"))
        today_str = datetime.now().strftime("%Y-%m-%d")
        qc_count_resp = (
            supabase.table(QC_TABLE)
            .select("id", count="exact")
            .eq("approved_by", current_user_name)
            .eq("qc_approved_date", today_str)
            .execute()
        )
        today_qc_count = qc_count_resp.count or 0

        # Selected-date-range QC count for the current user (when a date range filter is applied)
        if qc_date_from or qc_date_to:
            rq = (
                supabase.table(QC_TABLE)
                .select("id", count="exact")
                .eq("approved_by", current_user_name)
            )
            if qc_date_from:
                rq = rq.gte("qc_approved_date", qc_date_from)
            if qc_date_to:
                rq = rq.lte("qc_approved_date", qc_date_to)
            rq_resp = rq.execute()
            range_qc_count = rq_resp.count or 0
            if qc_date_from and qc_date_to and qc_date_from != qc_date_to:
                range_qc_label = f"{qc_date_from} → {qc_date_to}"
            else:
                single = qc_date_from or qc_date_to
                range_qc_label = single
    except Exception as e:
        flash(f"Error fetching QC data: {e}", "error")

    clean_display_name = get_clean_display_name(session.get("display_name", "User"))
    return render_template(
        "qc_gui.html",
        items=items,
        qc_search=qc_search,
        qc_wallet=qc_wallet,
        qc_date_from=qc_date_from,
        qc_date_to=qc_date_to,
        qc_status=qc_status,
        page_num=page,
        total_pages=total_pages,
        total_rows=total_rows,
        today_qc_count=today_qc_count,
        range_qc_count=range_qc_count,
        range_qc_label=range_qc_label,
        wallet_options=BS_INVESTMENT_WALLET_OPTIONS,
        search_for_options=QC_SEARCH_FOR_OPTIONS,
        scam_type_options=BS_INVESTMENT_SCAM_TYPE_OPTIONS,
        qc_remarks_options=QC_REMARKS_OPTIONS,
        allowed_pages=allowed_pages,
        display_name=session.get("display_name", "User"),
        clean_display_name=clean_display_name,
        is_admin=is_super_or_admin,
        role=session.get("role", "user"),
    )


@app.route("/qc-gui-import", methods=["POST"])
@login_required
def qc_gui_import():
    if "qc" not in session.get("allowed_pages", []):
        flash("Access denied.", "error")
        return redirect("/qc-gui")
    if not is_admin_or_above(session):
        flash("Import is available for admin users only.", "error")
        return redirect("/qc-gui")
    temp_path = None
    try:
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("No file selected", "error")
            return redirect("/qc-gui")
        if not is_allowed_file(file.filename):
            flash("Unsupported file type.", "error")
            return redirect("/qc-gui")
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        file_ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "csv"
        df = read_data_file(temp_path, file_ext)
        input_user_default = get_clean_display_name(session.get("display_name", "User"))
        records, _ = _qc_build_records_from_df(df, input_user_default)

        # Insert in chunks to avoid payload limits
        CHUNK = 500
        for i in range(0, len(records), CHUNK):
            supabase.table(QC_TABLE).insert(records[i:i + CHUNK]).execute()

        log_activity(
            action_type="import",
            target_table=QC_TABLE,
            extra_info={"file_name": filename, "records_count": len(records)}
        )
        flash(f"QC file imported successfully! {len(records)} records added.", "success")
    except Exception as e:
        flash(f"QC Import Error: {str(e)}", "error")
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
    return redirect("/qc-gui")


@app.route("/qc-gui-users", methods=["GET"])
@login_required
def qc_gui_users():
    if "qc" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    if not is_admin_or_above(session):
        return jsonify({"success": False, "error": "Admin access required."})
    current_user_name = get_clean_display_name(session.get("display_name", "User")).strip()
    try:
        auth_client = get_auth_supabase()
        rows = auth_client.table("dashboard_users").select("display_name,allowed_pages,is_active").eq("is_active", True).execute().data or []
        names = {current_user_name} if current_user_name and current_user_name != "User" else set()
        for row in rows:
            display_name = row.get("display_name")
            if not display_name:
                continue
            allowed_pages = row.get("allowed_pages") or []
            if isinstance(allowed_pages, str):
                try:
                    allowed_pages = json.loads(allowed_pages)
                except Exception:
                    allowed_pages = [p.strip() for p in allowed_pages.split(",") if p.strip()]
            if "qc" in (allowed_pages or []):
                names.add(get_clean_display_name(display_name))

        names = sorted(names)
        return jsonify({"success": True, "users": [{"display_name": name} for name in names]})
    except Exception as e:
        if current_user_name and current_user_name != "User":
            return jsonify({
                "success": True,
                "users": [{"display_name": current_user_name}],
                "warning": str(e),
            })
        return jsonify({"success": False, "error": str(e)})


@app.route("/qc-gui-allotment", methods=["POST"])
@login_required
def qc_gui_allotment():
    if "qc" not in session.get("allowed_pages", []):
        flash("Access denied.", "error")
        return redirect("/qc-gui")
    if not is_admin_or_above(session):
        flash("QC Allotment is available for admin users only.", "error")
        return redirect("/qc-gui")
    temp_path = None
    try:
        allot_user = get_clean_display_name(request.form.get("allot_user", "")).strip()
        if not allot_user:
            flash("QC Allotment Error: Select allotment user.", "error")
            return redirect("/qc-gui")
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("QC Allotment Error: No file selected.", "error")
            return redirect("/qc-gui")
        if not is_allowed_file(file.filename):
            flash("QC Allotment Error: Unsupported file type.", "error")
            return redirect("/qc-gui")

        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), f"qc_allotment_{uuid.uuid4().hex}_{filename}")
        file.save(temp_path)
        file_ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "csv"
        df = read_data_file(temp_path, file_ext)
        input_user_default = get_clean_display_name(session.get("display_name", "User"))
        records, _ = _qc_build_records_from_df(
            df,
            input_user_default=input_user_default,
            force_input_user=allot_user,
            reset_qc_update=True,
        )
        if not records:
            flash("QC Allotment Error: No rows found in file.", "error")
            return redirect("/qc-gui")

        CHUNK = 500
        for i in range(0, len(records), CHUNK):
            supabase.table(QC_TABLE).insert(records[i:i + CHUNK]).execute()

        log_activity(
            action_type="import",
            target_table=QC_TABLE,
            extra_info={"file_name": filename, "records_count": len(records), "allot_user": allot_user}
        )
        flash(f"QC Allotment done! {len(records)} cases allotted to {allot_user}.", "success")
    except Exception as e:
        flash(f"QC Allotment Error: {str(e)}", "error")
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
    return redirect("/qc-gui")


@app.route("/qc-gui-export", methods=["GET"])
@login_required
def qc_gui_export():
    if "qc" not in session.get("allowed_pages", []):
        flash("Access denied.", "error")
        return redirect("/qc-gui")
    try:
        qc_search     = request.args.get("qc_search",     "").strip()
        qc_wallet     = request.args.get("qc_wallet",     "").strip()
        qc_date_from  = request.args.get("qc_date_from",  "").strip()
        qc_date_to    = request.args.get("qc_date_to",    "").strip()

        def _build_qc_query():
            q = supabase.table(QC_TABLE).select(",".join(QC_EXPORT_COLUMNS))
            if qc_search:
                lt = f"%{qc_search}%"
                or_parts = ",".join(f"{f}.ilike.{lt}" for f in QC_SEARCH_FIELDS)
                q = q.or_(or_parts)
            if qc_wallet:
                q = q.eq("upi_bank_account_wallet", qc_wallet)
            if qc_date_from:
                q = q.gte("inserted_date", qc_date_from)
            if qc_date_to:
                q = q.lte("inserted_date", qc_date_to)
            return q

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return _stream_supabase_csv(_build_qc_query, "id", f"gui_qc_{ts}.csv")
    except Exception as e:
        flash(f"Export Error: {e}", "error")
        return redirect("/qc-gui")


@app.route("/qc-gui-template", methods=["GET"])
@login_required
def qc_gui_template():
    if "qc" not in session.get("allowed_pages", []):
        flash("Access denied.", "error")
        return redirect("/qc-gui")
    try:
        output = io.StringIO()
        csv.writer(output).writerow(QC_EXPORT_COLUMNS)
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            download_name="GUI_QC_Import_Template.csv",
            as_attachment=True,
            mimetype="text/csv"
        )
    except Exception as e:
        flash(f"Error generating template: {str(e)}", "error")
        return redirect("/qc-gui")


@app.route("/qc-gui-get-record", methods=["GET"])
@login_required
def qc_gui_get_record():
    if "qc" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        rid = request.args.get("id")
        if not rid:
            return jsonify({"success": False, "error": "No ID"})
        resp = supabase.table(QC_TABLE).select("*").eq("id", rid).limit(1).execute()
        if resp.data:
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Record not found"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/qc-gui-update", methods=["POST"])
@login_required
def qc_gui_update():
    if "qc" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data"})
        rid = data.get("id")
        if rid is None:
            return jsonify({"success": False, "error": "No ID"})

        ALLOWED = [
            "gui_id", "date", "input_user", "approved_by", "scam_type",
            "search_for", "upi_bank_account_wallet", "bank_name",
            "bank_account_number", "upi_vpa", "ifsc_code", "ac_holder_name",
            "web_contact_no", "payment_gateway_url", "website_url",
            "screenshot", "screenshot_case_report_link", "qc_remarks",
            "new_update_remark", "qc_status", "feature_type", "inserted_date"
        ]
        record = {}
        existing_qc_remarks = ""
        incoming_new_remark = None
        if "qc_remarks" in data or "new_update_remark" in data:
            old_resp = supabase.table(QC_TABLE).select("qc_remarks").eq("id", rid).limit(1).execute()
            if old_resp.data:
                existing_qc_remarks = str(old_resp.data[0].get("qc_remarks") or "").strip()
            incoming_new_remark = str(data.get("new_update_remark", data.get("qc_remarks", "")) or "").strip()
        for f in ALLOWED:
            if f not in data:
                continue
            if f == "qc_remarks":
                continue
            if f == "qc_status":
                continue
            val = str(data.get(f, "")).strip()
            if f in QC_INT_COLUMNS:
                record[f] = _qc_normalize_int_value(val)
            elif f in ("date", "inserted_date"):
                record[f] = val if val and val.upper() not in ("NA", "N/A", "") else None
            else:
                record[f] = val if val else "NA"
        if incoming_new_remark is not None:
            record["new_update_remark"] = incoming_new_remark if incoming_new_remark else "NA"
            record["qc_status"] = (existing_qc_remarks == incoming_new_remark)
            record["qc_approved_date"] = datetime.now().strftime("%Y-%m-%d")
        record["approved_by"] = get_clean_display_name(session.get("display_name", "User"))
        if not record:
            return jsonify({"success": False, "error": "No valid fields to update"})

        resp = supabase.table(QC_TABLE).update(record).eq("id", rid).execute()
        if resp.data:
            log_activity(
                action_type="field_update",
                target_table=QC_TABLE,
                target_record_id=rid,
                field_name="edit",
                new_value=str(record)
            )
            return jsonify({"success": True, "record": resp.data[0]})
        return jsonify({"success": False, "error": "Update failed"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/qc-gui-delete", methods=["POST"])
@login_required
def qc_gui_delete():
    if "qc" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    try:
        data = request.get_json()
        rid = data.get("id")
        if not rid:
            return jsonify({"success": False, "error": "No ID"})
        supabase.table(QC_TABLE).delete().eq("id", rid).execute()
        log_activity(
            action_type="field_update",
            target_table=QC_TABLE,
            target_record_id=rid,
            field_name="DELETE",
            old_value="EXISTS",
            new_value="DELETED"
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/qc-gui-tracker-stats", methods=["GET"])
@login_required
def qc_gui_tracker_stats():
    if "qc" not in session.get("allowed_pages", []):
        return jsonify({"success": False, "error": "Access denied."})
    if not is_admin_or_above(session):
        return jsonify({"success": False, "error": "Admin access required."})
    try:
        CHUNK = 1000
        all_rows, offset = [], 0
        while True:
            resp = supabase.table(QC_TABLE) \
                .select("scam_type,search_for,upi_bank_account_wallet,approved_by,input_user,inserted_date,new_update_remark,qc_status,qc_assigned_date,qc_approved_date,total_allotment_date") \
                .order("id", desc=False) \
                .range(offset, offset + CHUNK - 1).execute()
            chunk = resp.data or []
            all_rows.extend(chunk)
            if len(chunk) < CHUNK:
                break
            offset += CHUNK
        scam_counts = {}
        sf_counts = {}
        wallet_counts = {}
        approved_counts = {}
        input_user_counts = {}
        daily_counts = {}
        daily_qc_status = {}
        user_daily_qc_status = {}
        allotment_by_date = {}
        allotment_by_user_date = {}
        for row in all_rows:
            st  = (row.get("scam_type") or "Unknown").strip() or "Unknown"
            sf  = (row.get("search_for") or "Unknown").strip() or "Unknown"
            w   = (row.get("upi_bank_account_wallet") or "Unknown").strip() or "Unknown"
            ab  = (row.get("approved_by") or "Unknown").strip() or "Unknown"
            iu  = (row.get("input_user") or "Unknown").strip() or "Unknown"
            updater = ab if ab.upper() not in ("", "NA", "N/A", "UNKNOWN") else "Unknown"
            dt  = (row.get("inserted_date") or "")[:10]
            qc_dt = (row.get("qc_approved_date") or "")[:10]
            allot_dt = (row.get("total_allotment_date") or row.get("qc_assigned_date") or "")[:10]
            if allot_dt:
                allotment_by_date[allot_dt] = allotment_by_date.get(allot_dt, 0) + 1
                user_allot = allotment_by_user_date.setdefault(iu, {})
                user_allot[allot_dt] = user_allot.get(allot_dt, 0) + 1
            scam_counts[st] = scam_counts.get(st, 0) + 1
            sf_counts[sf] = sf_counts.get(sf, 0) + 1
            wallet_counts[w] = wallet_counts.get(w, 0) + 1
            approved_counts[ab] = approved_counts.get(ab, 0) + 1
            input_user_counts[iu] = input_user_counts.get(iu, 0) + 1
            if dt:
                daily_counts[dt] = daily_counts.get(dt, 0) + 1
            # User-wise daily QC status keyed by qc_approved_date (the date QC was completed)
            if qc_dt and row.get("new_update_remark") and str(row.get("new_update_remark")).strip().upper() not in ("", "NA", "N/A"):
                bucket = daily_qc_status.setdefault(qc_dt, {"total": 0, "true": 0, "false": 0})
                bucket["total"] += 1
                user_bucket = user_daily_qc_status.setdefault(updater, {})
                user_day = user_bucket.setdefault(qc_dt, {"total": 0, "true": 0, "false": 0})
                user_day["total"] += 1
                if row.get("qc_status") is True:
                    bucket["true"] += 1
                    user_day["true"] += 1
                elif row.get("qc_status") is False:
                    bucket["false"] += 1
                    user_day["false"] += 1
        return jsonify({
            "success": True,
            "total": len(all_rows),
            "scam_counts": scam_counts,
            "sf_counts": sf_counts,
            "wallet_counts": wallet_counts,
            "approved_counts": approved_counts,
            "input_user_counts": input_user_counts,
            "daily_counts": {k: daily_counts[k] for k in sorted(daily_counts)},
            "daily_qc_status": {k: daily_qc_status[k] for k in sorted(daily_qc_status)},
            "user_daily_qc_status": {
                user: {date_key: days[date_key] for date_key in sorted(days, reverse=True)}
                for user, days in sorted(user_daily_qc_status.items())
            },
            "allotment_by_date": {k: allotment_by_date[k] for k in sorted(allotment_by_date)},
            "allotment_by_user_date": {
                user: {date_key: counts[date_key] for date_key in sorted(counts, reverse=True)}
                for user, counts in sorted(allotment_by_user_date.items())
            },
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


if __name__ == "__main__":
    EXCEL_FOLDER_PATH.mkdir(exist_ok=True)
    load_config()
    load_excel_data()
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_DEBUG", "False").lower() == "true"
    app.run(debug=debug_mode, host='0.0.0.0', port=port)